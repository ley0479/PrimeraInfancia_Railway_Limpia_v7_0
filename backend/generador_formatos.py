"""
Módulo para generación de formatos ICBF
"""
import os
import json
from modules.dbapi_compat import sqlite3
import re
import unicodedata
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from modules.print_master import aplicar_configuracion_impresion_libro
from modules.plantillas_oficiales import get_plantilla_oficial, generar_desde_plantilla_oficial
from modules.seguridad.tenant_context import current_tenant_id
from modules.runtime_schema import schema_ddl_enabled
from models import ConfiguracionSistema, EstadoUsuario


class GeneradorFormatos:
    """Genera formatos ICBF desde datos de la base de datos"""
    
    def __init__(self, db_path, templates_path, output_path):
        self.db_path = db_path
        self.templates_path = templates_path
        self.output_path = output_path
    
    def get_db_connection(self):
        """Obtiene conexión a BD"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _participantes_unidad(self, unidad):
        """Lee una UDS desde la versión maestra activa, con fallback pre-migración."""
        fid = int(current_tenant_id(default=1) or 1)
        conn = self.get_db_connection()
        try:
            cur = conn.cursor()
            try:
                version = cur.execute(
                    "SELECT id FROM master_versiones WHERE fundacion_id=? AND activa=1 ORDER BY id DESC LIMIT 1",
                    (fid,),
                ).fetchone()
            except Exception:
                version = None
            if version:
                rows = cur.execute(
                    """SELECT *, unidad_servicio AS unidad, documento AS nui
                       FROM master_ninos
                       WHERE activo=1 AND fundacion_id=?
                         AND LOWER(TRIM(COALESCE(unidad_servicio,'')))=LOWER(TRIM(?))
                       ORDER BY nombre_completo, documento""",
                    (fid, unidad),
                ).fetchall()
            else:
                rows = cur.execute(
                    """SELECT * FROM beneficiarios
                       WHERE COALESCE(fundacion_id,1)=?
                         AND LOWER(TRIM(COALESCE(unidad,'')))=LOWER(TRIM(?))
                         AND UPPER(COALESCE(estado,'ACTIVO')) NOT IN ('INACTIVO','RETIRADO','FALLECIDO')
                       ORDER BY COALESCE(NULLIF(nombres,''),documento), documento""",
                    (fid, unidad),
                ).fetchall()
            salida = []
            for row in rows:
                data = dict(row)
                try:
                    extra = json.loads(data.get('datos_json') or '{}')
                    if isinstance(extra, dict):
                        data = {**extra, **data}
                except Exception:
                    pass
                salida.append(data)
            return salida
        finally:
            conn.close()

    def _aplicar_impresion_y_guardar(self, ruta, tipo_formato):
        """Aplica la tabla maestra de impresión a un Excel ya generado."""
        try:
            wb = load_workbook(ruta)
            aplicar_configuracion_impresion_libro(wb, tipo_formato, source_name=os.path.basename(ruta))
            wb.save(ruta)
        except Exception as exc:
            print(f"No se pudo aplicar configuración de impresión {tipo_formato} a {ruta}: {exc}")


    def _sincronizar_calendario_entrega(self, tipo_formato, titulo, mes, año, unidad, ruta=None):
        """Marca en el Calendario Inteligente que un formato fue generado/entregado.

        No modifica el archivo oficial. Solo actualiza o crea el entregable operativo
        correspondiente al periodo del formato.
        """
        try:
            from modules.calendario_inteligente.repository import CalendarioInteligenteRepository
            repo = CalendarioInteligenteRepository(self.db_path, self.output_path)
            if schema_ddl_enabled():
                repo.init_schema()
            repo.sincronizar_entrega({
                'titulo': titulo,
                'fecha_entrega': f"{int(año):04d}-{int(mes):02d}-01",
                'modulo': tipo_formato,
                'tipo_formato': tipo_formato,
                'unidad': unidad or '',
                'archivo_evidencia': os.path.basename(ruta) if ruta else None,
                'observaciones': 'Sincronizado automáticamente al generar formato desde la plataforma.'
            })
        except Exception as exc:
            print(f"No se pudo sincronizar Calendario Inteligente ({tipo_formato}): {exc}")

    def _plantilla_oficial_disponible(self, tipo_formato):
        try:
            info = get_plantilla_oficial(self.templates_path, tipo_formato)
            return bool(info and info.get('existe'))
        except Exception:
            return False

    def _usuario_oficial(self, b):
        """Normaliza un participante sin inventar valores ausentes."""
        return {
            'Id': b.get('id') or '',
            'NUI': b.get('nui') or '',
            'NumeroDocumento': b.get('numero_documento') or b.get('documento') or '',
            'Documento': b.get('documento') or b.get('numero_documento') or '',
            'TipoDocumento': b.get('tipo_documento') or b.get('tipo_doc') or '',
            'Nombre': b.get('nombre') or b.get('nombre_completo') or b.get('nombres') or '',
            'Nombres': b.get('nombres') or b.get('nombre_completo') or '',
            'Apellidos': b.get('apellidos') or '',
            'PrimerNombre': b.get('primer_nombre') or '',
            'SegundoNombre': b.get('segundo_nombre') or '',
            'PrimerApellido': b.get('primer_apellido') or '',
            'SegundoApellido': b.get('segundo_apellido') or '',
            'FechaNacimiento': b.get('fecha_nacimiento') or '',
            'FechaIngreso': b.get('fecha_ingreso') or '',
            'FechaRetiro': b.get('fecha_retiro') or '',
            'CausaRetiro': b.get('causa_retiro') or b.get('motivo_retiro') or '',
            'Acudiente': b.get('nombre_acudiente') or '',
            'DocumentoAcudiente': b.get('documento_acudiente') or '',
            'Parentesco': b.get('parentesco') or '',
            'Telefono': b.get('telefono') or b.get('celular') or '',
            'EdadMeses': b.get('edad_meses') if b.get('edad_meses') not in (None, '') else '',
            'GrupoEdad': b.get('grupo_edad') or '',
            'TipoBeneficiario': b.get('tipo_beneficiario') or '',
            'Unidad': b.get('unidad') or b.get('unidad_servicio') or '',
            'Estado': b.get('estado') or '',
        }

    def _metadata_oficial(self, mes, año, unidad, coordinador=None):
        """Construye encabezados oficiales desde la base de datos.

        Alpha16 corrige el problema de encabezados desactualizados: el formato
        no puede heredar UNIDAD DEMO 04, LUIS u otro dato de la plantilla. Los valores
        se toman por UDS exacta desde beneficiarios, unidades y talento humano.
        """
        meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

        def norm(valor):
            texto = str(valor or '').strip().lower()
            texto = unicodedata.normalize('NFKD', texto)
            texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
            texto = texto.replace('ñ', 'n')
            texto = re.sub(r'[^a-z0-9]+', ' ', texto)
            return ' '.join(texto.split())

        def limpiar(valor):
            return str(valor or '').strip()

        def row_value(row, *keys):
            if not row:
                return ''
            for key in keys:
                try:
                    if key in row.keys() and row[key] not in (None, ''):
                        return row[key]
                except Exception:
                    try:
                        if key in row and row.get(key) not in (None, ''):
                            return row.get(key)
                    except Exception:
                        pass
            return ''

        def full_name(row):
            if not row:
                return ''
            direct = row_value(row, 'nombre', 'nombre_completo', 'nombres_y_apellidos', 'Nombre', 'NOMBRE')
            if direct:
                return limpiar(direct).upper()
            return ' '.join(limpiar(row_value(row, k)) for k in ('nombres', 'apellidos') if limpiar(row_value(row, k))).upper()

        def unidades_de_row(row):
            valores = []
            for key in ('unidad', 'nombre', 'Nombre UDS', 'nombre_uds', 'unidad_servicio', 'comunidad'):
                val = row_value(row, key)
                if val:
                    valores.append(val)
            raw_unidades = row_value(row, 'unidades')
            if raw_unidades:
                try:
                    parsed = json.loads(raw_unidades)
                    if isinstance(parsed, list):
                        valores.extend(parsed)
                    elif isinstance(parsed, dict):
                        valores.extend(parsed.values())
                    else:
                        valores.append(parsed)
                except Exception:
                    valores.extend(re.split(r'[;,|/]+', str(raw_unidades)))
            return [limpiar(v).upper() for v in valores if limpiar(v)]

        unidad_norm = norm(unidad)

        def coincide_unidad(row):
            return any(norm(u) == unidad_norm for u in unidades_de_row(row))

        beneficiario_ref = {}
        unidad_db = {}
        fundacion_db = {}
        talentos = []
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()

            try:
                fundacion_id = int(current_tenant_id(default=1) or 1)
                fila_f = cursor.execute("SELECT * FROM fundaciones WHERE id=? LIMIT 1", (fundacion_id,)).fetchone()
                fundacion_db = dict(fila_f) if fila_f else {}
            except Exception:
                fundacion_db = {}

            participantes = self._participantes_unidad(unidad)
            beneficiario_ref = participantes[0] if participantes else {}

            try:
                for fila in cursor.execute("SELECT * FROM unidades").fetchall():
                    data = dict(fila)
                    if coincide_unidad(data):
                        unidad_db = data
                        break
            except Exception:
                unidad_db = {}

            for tabla in ('master_talento_humano', 'coordinadores', 'th_personas'):
                try:
                    if tabla == 'master_talento_humano':
                        filas = cursor.execute(
                            "SELECT * FROM master_talento_humano WHERE activo=1 AND fundacion_id=?",
                            (fundacion_id,),
                        ).fetchall()
                    else:
                        filas = cursor.execute(
                            f"SELECT * FROM {tabla} WHERE COALESCE(fundacion_id,1)=?",
                            (fundacion_id,),
                        ).fetchall()
                except Exception:
                    continue
                for fila in filas:
                    data = dict(fila)
                    estado = norm(row_value(data, 'estado'))
                    if estado in {'inactivo', 'retirado'}:
                        continue
                    activo = row_value(data, 'activo')
                    if str(activo).strip() in {'0', 'False', 'false'}:
                        continue
                    if coincide_unidad(data):
                        talentos.append(data)
            conn.close()
        except Exception:
            pass

        def score_talento(row):
            cargo = norm(row_value(row, 'cargo', 'tipo_equipo', 'perfil'))
            if 'agente' in cargo or 'docente' in cargo or 'educativo' in cargo:
                return 0
            if 'coordin' in cargo:
                return 1
            if 'suplente' in cargo or 'apoyo' in cargo or 'auxiliar' in cargo:
                return 2
            return 3

        talentos.sort(key=lambda item: (score_talento(item), norm(full_name(item))))
        responsable_row = talentos[0] if talentos else {}
        suplente_row = next((t for t in talentos if any(k in norm(row_value(t, 'cargo', 'tipo_equipo', 'perfil')) for k in ['suplente', 'apoyo', 'auxiliar'])), {})
        coordinador_row = next((t for t in talentos if 'coordin' in norm(row_value(t, 'cargo', 'tipo_equipo', 'perfil'))), {})

        responsable = full_name(responsable_row)
        if not responsable and coordinador:
            try:
                responsable = f"{coordinador['nombres']} {coordinador['apellidos']}".strip().upper()
            except Exception:
                responsable = ''

        regional = row_value(beneficiario_ref, 'regional', 'Regional') or row_value(fundacion_db, 'departamento', 'regional')
        centro_zonal = row_value(beneficiario_ref, 'centro_zonal', 'CentroZonal', 'Centro Zonal') or row_value(fundacion_db, 'centro_zonal')
        municipio = row_value(beneficiario_ref, 'municipio', 'Municipio') or row_value(unidad_db, 'municipio') or row_value(fundacion_db, 'municipio')
        modalidad = row_value(beneficiario_ref, 'modalidad', 'Modalidad') or row_value(unidad_db, 'modalidad') or 'EDUCACIÓN INICIAL PROPIA DIARIA - PROPIA E INTERCULTURAL'
        servicio = row_value(beneficiario_ref, 'servicio_atencion', 'ServicioAtencion', 'servicio') or modalidad
        eas = row_value(beneficiario_ref, 'nombre_eas', 'NombreEAS', 'entidad_administradora') or row_value(fundacion_db, 'nombre')
        codigo_uds = row_value(beneficiario_ref, 'codigo_unidad_servicio', 'codigo_uds', 'CodigoUnidadServicio', 'codigo_unidad') or row_value(unidad_db, 'codigo_unidad_servicio', 'codigo_uds', 'codigo')
        direccion = row_value(unidad_db, 'direccion', 'Direccion') or row_value(beneficiario_ref, 'direccion_unidad', 'DireccionUnidad') or row_value(responsable_row, 'direccion')
        telefono_unidad = row_value(unidad_db, 'telefono', 'Telefono') or row_value(responsable_row, 'telefono', 'celular')
        contrato = row_value(beneficiario_ref, 'numero_contrato', 'NumeroContrato', 'contrato') or row_value(unidad_db, 'contrato') or row_value(responsable_row, 'contrato')
        barrio = row_value(beneficiario_ref, 'barrio', 'Barrio') or row_value(unidad_db, 'barrio')
        unidad_origen = row_value(beneficiario_ref, 'nombre_unidad_origen', 'NombreUnidadOrigen', 'nombre_punto_entrega_origen') or unidad
        codigo_origen = row_value(beneficiario_ref, 'codigo_unidad_origen', 'CodigoUnidadOrigen') or codigo_uds

        mes_nombre = meses[int(mes) - 1] if 1 <= int(mes) <= 12 else str(mes).upper()
        metadata = {
            'regional': limpiar(regional).upper(),
            'Regional': limpiar(regional).upper(),
            'centro_zonal': limpiar(centro_zonal).upper(),
            'CentroZonal': limpiar(centro_zonal).upper(),
            'municipio': limpiar(municipio).upper(),
            'Municipio': limpiar(municipio).upper(),
            'modalidad': limpiar(modalidad).upper(),
            'Modalidad': limpiar(modalidad).upper(),
            'servicio_atencion': limpiar(servicio).upper(),
            'ServicioAtencion': limpiar(servicio).upper(),
            'eas': limpiar(eas).upper(),
            'eas_pds': limpiar(eas).upper(),
            'NombreEAS': limpiar(eas).upper(),
            'nit': limpiar(row_value(fundacion_db, 'nit')),
            'unidad': limpiar(unidad).upper(),
            'Unidad': limpiar(unidad).upper(),
            'unidad_origen': limpiar(unidad_origen).upper(),
            'NombreUnidadOrigen': limpiar(unidad_origen).upper(),
            'codigo_unidad': limpiar(codigo_uds),
            'codigo_uds': limpiar(codigo_uds),
            'codigo_cuentame': limpiar(codigo_uds),
            'CodigoUnidadServicio': limpiar(codigo_uds),
            'codigo_origen': limpiar(codigo_origen),
            'CodigoUnidadOrigen': limpiar(codigo_origen),
            'direccion': limpiar(direccion).upper(),
            'direccion_unidad': limpiar(direccion).upper(),
            'direccion_uds': limpiar(direccion).upper(),
            'DireccionUnidad': limpiar(direccion).upper(),
            'barrio': limpiar(barrio).upper(),
            'Barrio': limpiar(barrio).upper(),
            'telefono': limpiar(telefono_unidad),
            'telefono_uds': limpiar(telefono_unidad),
            'Telefono': limpiar(telefono_unidad),
            'telefono_docente': limpiar(row_value(responsable_row, 'telefono', 'celular')) or limpiar(telefono_unidad),
            'responsable': responsable,
            'docente': responsable,
            'agente_educativo': responsable,
            'cedula_docente': limpiar(row_value(responsable_row, 'documento', 'cedula', 'identificacion')),
            'documento_agente': limpiar(row_value(responsable_row, 'documento', 'cedula', 'identificacion')),
            'suplente': full_name(suplente_row),
            'telefono_suplente': limpiar(row_value(suplente_row, 'telefono', 'celular')),
            'coordinador': full_name(coordinador_row) or limpiar(row_value(responsable_row, 'coordinador')),
            'contrato': limpiar(contrato),
            'NumeroContrato': limpiar(contrato),
            'mes': mes_nombre,
            'mes_nombre': mes_nombre,
            'mes_numero': int(mes),
            'Mes': mes_nombre,
            'anio': año,
            'año': año,
            'year': año,
            'fecha_entrega': datetime(año, mes, 1).strftime('%d/%m/%Y'),
            'FechaEntrega': datetime(año, mes, 1).strftime('%d/%m/%Y'),
        }
        return metadata
    
    # ==================== ASISTENCIA / RAM OFICIAL ====================
    def generar_asistencia(self, mes, año, unidad):
        """Genera el listado oficial RAM V3 con participantes reales de la UDS.

        Conserva la plantilla protegida, pagina 20 participantes por hoja y deja
        vacías las marcas diarias cuando no existe un registro electrónico
        verificable de asistencia.
        """
        beneficiarios = self._participantes_unidad(unidad)

        metadata = self._metadata_oficial(mes, año, unidad)
        metadata.update({'mes_numero': int(mes), 'mes_nombre': metadata.get('mes_nombre') or metadata.get('mes')})
        nombre_archivo = f"RAM_V3_{re.sub(r'[^A-Za-z0-9_-]+', '_', str(unidad)).strip('_')}_{int(año):04d}_{int(mes):02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        resultado = generar_desde_plantilla_oficial(
            'ram',
            {'metadata': metadata, 'usuarios': [self._usuario_oficial(b) for b in beneficiarios]},
            ruta,
            self.templates_path,
        )
        self._sincronizar_calendario_entrega('ram', 'Listado de asistencia RAM', mes, año, unidad, resultado)
        return resultado

    # ==================== BIENESTARINA ====================
    def generar_bienestarina(self, mes, año, unidad, bolsas_por_beneficiario=1):
        """Genera formato de entrega de Bienestarina"""
        beneficiarios = self._participantes_unidad(unidad)
        coordinador = None

        if self._plantilla_oficial_disponible('bienestarina'):
            nombre_archivo = f"BIENESTARINA_{unidad}_{año}{mes:02d}.xlsx"
            ruta = os.path.join(self.output_path, nombre_archivo)
            datos = {
                'metadata': self._metadata_oficial(mes, año, unidad, coordinador),
                'usuarios': [self._usuario_oficial(b) for b in beneficiarios],
            }
            return generar_desde_plantilla_oficial('bienestarina', datos, ruta, self.templates_path)
        
        # Crear DataFrame
        df = pd.DataFrame()
        df['BENEFICIARIO'] = [b.get('nombres') or b.get('nombre_completo') or '' for b in beneficiarios]
        df['NUI'] = [b.get('nui', '') for b in beneficiarios]
        df['DOCUMENTO'] = [b['documento'] for b in beneficiarios]
        df['RESPONSABLE'] = ''
        df['PARENTESCO'] = ''
        df['PRIMER NOMBRE'] = [b.get('primer_nombre', '') for b in beneficiarios]
        df['SEGUNDO NOMBRE'] = [b.get('segundo_nombre', '') for b in beneficiarios]
        df['PRIMER APELLIDO'] = [b.get('primer_apellido', '') for b in beneficiarios]
        df['SEGUNDO APELLIDO'] = [b.get('segundo_apellido', '') for b in beneficiarios]
        df['ACUDIENTE'] = [b.get('nombre_acudiente', '') for b in beneficiarios]
        df['DOC ACUDIENTE'] = [b.get('documento_acudiente', '') for b in beneficiarios]
        df['PARENTESCO'] = [b.get('parentesco', '') for b in beneficiarios]
        df['MES ENTREGA'] = datetime(año, mes, 1).strftime('%B').upper()
        df['BOLSAS'] = [bolsas_por_beneficiario] * len(beneficiarios)
        df['FIRMA'] = ''
        df['HUELLA'] = ''
        df['OBSERVACIONES'] = ''
        
        # Metadatos
        fecha_str = datetime(año, mes, 1).strftime('%B %Y')
        
        # Guardar
        nombre_archivo = f"BIENESTARINA_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        
        # Crear workbook con estilos
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='ENTREGA', index=False)
            self._escribir_y_resaltar(df, writer, 'ENTREGA', beneficiarios)
            
            workbook = writer.book
            worksheet = writer.sheets['ENTREGA']
            
            # Agregar encabezado
            worksheet.insert_rows(1, 5)
            worksheet['A1'] = 'FORMATO DE ENTREGA - BIENESTARINA'
            worksheet['A1'].font = Font(bold=True, size=12)
            worksheet['A2'] = f'Período: {fecha_str}'
            worksheet['A3'] = f'Unidad: {unidad}'
            if coordinador:
                worksheet['A4'] = f"Coordinador: {coordinador['nombres']} {coordinador['apellidos']}"

            aplicar_configuracion_impresion_libro(workbook, 'bienestarina', source_name=nombre_archivo)
            
        return ruta

    def _escribir_y_resaltar(self, df, writer, sheet_name, raw_data):
        """Escribe el DataFrame sin alterar colores oficiales.

        Las versiones anteriores pintaban cambios recientes en verde. Eso era útil
        para auditoría, pero en formatos oficiales ICBF cambia la presentación y
        puede borrar/alterar colores exigidos. La trazabilidad se conserva en los
        datos y reportes, no en rellenos del Excel oficial.
        """
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _guardar_con_resaltado(self, df, ruta, sheet_name, raw_data):
        """Versión para to_excel directo (Asistencia)"""
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            self._escribir_y_resaltar(df, writer, sheet_name, raw_data)
    
    # ==================== RAN (Registro Asistencia Nutrición) ====================
    def generar_ran(self, mes, año, unidad):
        """Genera RAN (Registro Asistencia Nutrición)"""
        beneficiarios = self._participantes_unidad(unidad)
        resultados = {'GESTANTES': 0, '0-5 MESES': 0, '6-11 MESES': 0, '1-2 AÑOS': 0, '3-5 AÑOS': 0}
        for b in beneficiarios:
            tipo = str(b.get('tipo_beneficiario') or b.get('grupo_etario') or '').upper()
            try: edad = int(float(b.get('edad_meses') or 0))
            except Exception: edad = 0
            if 'GESTANTE' in tipo: grupo = 'GESTANTES'
            elif edad <= 5: grupo = '0-5 MESES'
            elif edad <= 11: grupo = '6-11 MESES'
            elif edad <= 35: grupo = '1-2 AÑOS'
            else: grupo = '3-5 AÑOS'
            resultados[grupo] += 1
        
        # Crear DataFrame
        df = pd.DataFrame(list(resultados.items()), columns=['GRUPO_ETARIO', 'CANTIDAD'])
        df['FECHA'] = datetime(año, mes, 1).strftime('%d/%m/%Y')
        df['RESPONSABLE'] = ''
        
        # Guardar
        nombre_archivo = f"RAN_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='RAN', index=False)
        self._aplicar_impresion_y_guardar(ruta, 'ram_ran')
        
        return ruta
    
    # ==================== RPP (Registro Procedencia Procedimiento) ====================
    def generar_rpp(self, mes, año, unidad):
        """Genera RPP desde plantilla oficial si está disponible."""
        beneficiarios = self._participantes_unidad(unidad)
        coordinador = None

        if self._plantilla_oficial_disponible('rpp'):
            nombre_archivo = f"RPP_{unidad}_{año}{mes:02d}.xlsx"
            ruta = os.path.join(self.output_path, nombre_archivo)
            datos = {
                'metadata': self._metadata_oficial(mes, año, unidad, coordinador),
                'usuarios': [self._usuario_oficial(b) for b in beneficiarios],
            }
            return generar_desde_plantilla_oficial('rpp', datos, ruta, self.templates_path)

        total = len(beneficiarios)
        df = pd.DataFrame({
            'CONCEPTO': ['BENEFICIARIOS ASISTIDOS', 'RACIONES ENTREGADAS', 'TOTAL KILOGRAMOS'],
            'CANTIDAD': [total, total, total * 0.5],
            'FECHA': datetime(año, mes, 1).strftime('%d/%m/%Y')
        })

        nombre_archivo = f"RPP_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='RPP', index=False)
        self._aplicar_impresion_y_guardar(ruta, 'rpp')
        return ruta

    # ==================== NUTRICIÓN ====================
    def generar_nutricion(self, mes, año, unidad):
        """Genera reporte desde valoraciones identificadas por documento/tenant."""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener mediciones del período
        fid = int(current_tenant_id(1) or 1)
        cursor.execute("""
            SELECT nombre_completo AS nombres, documento,
                   peso_kg AS peso, talla_cm AS talla,
                   diagnostico_global AS estado_nutricional
            FROM sn_valoraciones v
            WHERE v.activo=1 AND COALESCE(v.fundacion_id,1)=?
              AND v.unidad=? AND v.periodo=?
              AND v.id=(SELECT v2.id FROM sn_valoraciones v2
                        WHERE v2.activo=1 AND COALESCE(v2.fundacion_id,1)=COALESCE(v.fundacion_id,1)
                          AND v2.documento=v.documento AND v2.periodo=v.periodo
                        ORDER BY v2.fecha_valoracion DESC,v2.id DESC LIMIT 1)
            ORDER BY nombre_completo, fecha_valoracion DESC
        """, (fid, unidad, f"{int(año):04d}-{int(mes):02d}"))
        
        registros = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        df = pd.DataFrame(registros)
        
        if not df.empty:
            df.rename(columns={
                'nombres': 'BENEFICIARIO',
                'documento': 'DOCUMENTO',
                'peso': 'PESO (KG)',
                'talla': 'TALLA (CM)',
                'estado_nutricional': 'ESTADO'
            }, inplace=True)
        
        # Guardar
        nombre_archivo = f"NUTRICION_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='NUTRICION', index=False)
        
        return ruta
    
    # ==================== INFORME PEDAGÓGICO ====================
    def generar_informe_pedagogico(self, docente_id, mes, año):
        """Genera informe pedagógico del docente"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener informe
        cursor.execute("""
            SELECT * FROM informes_pedagogicos
            WHERE docente_id = ? AND mes = ? AND año = ?
        """, (docente_id, mes, año))
        
        informe = cursor.fetchone()
        
        if not informe:
            conn.close()
            return None
        
        # Obtener datos del docente
        cursor.execute("SELECT * FROM docentes WHERE id = ?", (docente_id,))
        docente = cursor.fetchone()
        
        # Obtener evidencias
        cursor.execute("""
            SELECT * FROM evidencias
            WHERE informe_id = ?
            ORDER BY fecha_carga
        """, (informe['id'],))
        
        evidencias = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        # Crear informe en formato
        df = pd.DataFrame({
            'CAMPO': [
                'Docente', 'Unidad', 'Período',
                'Tema del Mes', 'Objetivos', 'Actividades Realizadas',
                'Resultados', 'Participación Familiar', 'Logros',
                'Dificultades', 'Recomendaciones', 'Total Evidencias'
            ],
            'CONTENIDO': [
                f"{docente['nombres']} {docente['apellidos']}",
                docente['unidad'],
                f"{mes}/{año}",
                informe['tema_mes'] or '',
                informe['objetivos'] or '',
                informe['actividades'] or '',
                informe['resultados'] or '',
                informe['participacion_familiar'] or '',
                informe['logros'] or '',
                informe['dificultades'] or '',
                informe['recomendaciones'] or '',
                len(evidencias)
            ]
        })
        
        # Guardar
        nombre_archivo = f"INFORME_PEDAGOGICO_DOCENTE{docente_id}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='INFORME', index=False)
        
        return ruta
    
    # ==================== GENERACIÓN MASIVA ====================
    def generar_mes_completo(self, mes, año, unidad):
        """Genera todos los formatos para un mes y unidad"""
        archivos_generados = {
            'asistencia': None,
            'bienestarina': None,
            'ran': None,
            'rpp': None,
            'nutricion': None
        }
        
        try:
            archivos_generados['asistencia'] = self.generar_asistencia(mes, año, unidad)
            self._sincronizar_calendario_entrega('RAM/RAN/Asistencia', 'Generación RAM/RAN/Asistencia', mes, año, unidad, archivos_generados['asistencia'])

            archivos_generados['bienestarina'] = self.generar_bienestarina(mes, año, unidad)
            self._sincronizar_calendario_entrega('Bienestarina', 'Generación Bienestarina', mes, año, unidad, archivos_generados['bienestarina'])

            archivos_generados['ran'] = self.generar_ran(mes, año, unidad)
            self._sincronizar_calendario_entrega('RAM/RAN/Asistencia', 'Generación RAN', mes, año, unidad, archivos_generados['ran'])

            archivos_generados['rpp'] = self.generar_rpp(mes, año, unidad)
            self._sincronizar_calendario_entrega('RPP', 'Generación RPP', mes, año, unidad, archivos_generados['rpp'])

            archivos_generados['nutricion'] = self.generar_nutricion(mes, año, unidad)
            self._sincronizar_calendario_entrega('Nutrición', 'Generación reporte nutricional', mes, año, unidad, archivos_generados['nutricion'])
        except Exception as e:
            print(f"Error generando formatos: {e}")
        
        return archivos_generados
