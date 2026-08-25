from __future__ import annotations

import json
import os
from datetime import datetime

from flask import Blueprint, Response, g, jsonify, request, send_from_directory
from modules.seguridad.tenant_context import tenant_path
from werkzeug.utils import secure_filename

from .repository import CruceBasesRepository, now_iso
from .services import (
    allowed_units_for_user,
    comparar_bases,
    docente_por_unidad,
    generar_excel_resultado,
    generar_pdf_resultado,
    insertar_alertas_salud_nutricion,
    normalizar_base,
    normalize_unidad,
    read_tabular_file,
)

ALLOWED = {'.xlsx', '.xls', '.xlsm', '.ods', '.csv', '.txt', '.tsv', '.tab', '.dat', '.html', '.htm', '.json', '.docx', '.pdf'}


def allowed_file(filename: str) -> bool:
    return os.path.splitext((filename or '').lower())[1] in ALLOWED


def save_upload(file, folder: str, prefix: str) -> dict:
    os.makedirs(folder, exist_ok=True)
    nombre_original = file.filename or 'archivo'
    nombre = secure_filename(nombre_original)
    nombre_guardado = f"{prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nombre}"
    ruta = os.path.join(folder, nombre_guardado)
    file.save(ruta)
    return {'nombre_original': nombre_original, 'nombre_guardado': nombre_guardado, 'ruta': ruta}


def user_ctx() -> dict:
    user = getattr(g, 'current_user', {}) or {}
    return {
        'usuario_id': user.get('id'),
        'usuario': user.get('username') or user.get('email') or 'sistema',
        'fundacion_id': user.get('fundacion_id') or 1,
        'rol': user.get('rol') or 'SUPERADMIN',
        'raw': user,
    }


def register_cruce_bases(app, database_path: str, upload_folder: str, output_folder: str) -> None:
    repo = CruceBasesRepository(database_path)
    repo.init_schema()
    bp = Blueprint('cruce_bases', __name__, url_prefix='/api/cruce-bases')
    module_upload = tenant_path(upload_folder, 'cruce_bases')
    module_reports = tenant_path(output_folder, 'cruce_bases')
    os.makedirs(module_upload, exist_ok=True)
    os.makedirs(module_reports, exist_ok=True)
    try:
        from .informe_estadistico import ensure_logo_schema
        ensure_logo_schema(database_path)
    except Exception:
        pass

    @bp.before_request
    def _ensure_schema():
        repo.init_schema()

    @bp.route('/comparar', methods=['POST'])
    def comparar():
        if 'base_anterior' not in request.files or 'base_actual' not in request.files:
            return jsonify({'error': 'Debes cargar base_anterior y base_actual.'}), 400
        anterior_file = request.files['base_anterior']
        actual_file = request.files['base_actual']
        if not anterior_file.filename or not actual_file.filename:
            return jsonify({'error': 'Selecciona ambas bases para comparar.'}), 400
        if not allowed_file(anterior_file.filename) or not allowed_file(actual_file.filename):
            return jsonify({'error': 'Formato no permitido. Usa Excel, CSV, TXT, JSON, DOCX, PDF tabular u otro formato tabular permitido.'}), 400

        ctx = user_ctx()
        mes = request.form.get('mes', type=int) or datetime.now().month
        anio = request.form.get('anio', type=int) or request.form.get('año', type=int) or datetime.now().year
        periodo = f'{anio}-{mes:02d}'
        saved_ant = save_upload(anterior_file, module_upload, 'BASE_ANTERIOR')
        saved_act = save_upload(actual_file, module_upload, 'BASE_ACTUAL')
        try:
            ant_rows, ant_errors = normalizar_base(read_tabular_file(saved_ant['ruta']))
            act_rows, act_errors = normalizar_base(read_tabular_file(saved_act['ruta']))
            errores = [{'base': 'anterior', **e} for e in ant_errors] + [{'base': 'actual', **e} for e in act_errors]
            if not act_rows:
                return jsonify({'error': 'La base actual no contiene registros válidos.', 'errores': errores[:100]}), 400
            resultado = comparar_bases(ant_rows, act_rows)
            resultado['errores'] = errores
            resultado['periodo'] = periodo
            resultado['fecha_cruce'] = now_iso()
        except Exception as exc:
            return jsonify({'error': f'No se pudo comparar las bases: {exc}'}), 400

        metadata = {
            'fundacion_id': ctx['fundacion_id'], 'usuario_id': ctx['usuario_id'], 'usuario': ctx['usuario'],
            'mes': mes, 'anio': anio, 'periodo': periodo,
            'archivo_anterior': saved_ant['nombre_original'], 'archivo_actual': saved_act['nombre_original'],
            'ruta_anterior': saved_ant['ruta'], 'ruta_actual': saved_act['ruta'],
        }
        cruce_id = repo.guardar_cruce(resultado, metadata)
        try:
            insertar_alertas_salud_nutricion(database_path, resultado, ctx['fundacion_id'])
        except Exception:
            pass
        excel_path = generar_excel_resultado(resultado, module_reports, f'CRUCE_BASES_{cruce_id}')
        pdf_path = generar_pdf_resultado(resultado, module_reports, 'resumen', f'CRUCE_BASES_{cruce_id}')
        repo.actualizar_reportes(cruce_id, os.path.basename(excel_path), os.path.basename(pdf_path))
        return jsonify({
            'message': 'Cruce mensual generado correctamente.',
            'cruce_id': cruce_id,
            'resumen': resultado['resumen'],
            'resultado': resultado,
            'errores': resultado.get('errores', [])[:100],
            'reporte_excel': os.path.basename(excel_path),
            'reporte_pdf': os.path.basename(pdf_path),
        }), 201

    @bp.route('/ultimo', methods=['GET'])
    def ultimo():
        ctx = user_ctx()
        row = repo.ultimo_cruce(ctx['fundacion_id'], superadmin=ctx['rol'] == 'SUPERADMIN')
        if not row:
            return jsonify({'cruce': None, 'resumen': {}, 'resultado': {}}), 200
        resultado = json.loads(row.get('resultado_json') or '{}')
        return jsonify({'cruce': row, 'resumen': row, 'resultado': resultado}), 200

    @bp.route('/historial', methods=['GET'])
    def historial():
        ctx = user_ctx()
        if ctx['rol'] == 'SUPERADMIN':
            rows = repo.fetch_all("SELECT * FROM cb_cruces ORDER BY fecha_cruce DESC, id DESC LIMIT 100")
        else:
            rows = repo.fetch_all("SELECT * FROM cb_cruces WHERE fundacion_id = ? ORDER BY fecha_cruce DESC, id DESC LIMIT 100", (ctx['fundacion_id'],))
        return jsonify({'historial': rows}), 200

    @bp.route('/detalle/<tipo>', methods=['GET'])
    def detalle(tipo: str):
        ctx = user_ctx()
        cruce_id = request.args.get('cruce_id', type=int)
        if cruce_id:
            row = repo.fetch_one("SELECT * FROM cb_cruces WHERE id = ?", (cruce_id,))
            if row and ctx['rol'] != 'SUPERADMIN' and int(row.get('fundacion_id') or 0) != int(ctx['fundacion_id'] or 0):
                row = None
        else:
            row = repo.ultimo_cruce(ctx['fundacion_id'], superadmin=ctx['rol'] == 'SUPERADMIN')
        if not row:
            return jsonify({'items': [], 'tipo': tipo}), 200
        resultado = json.loads(row.get('resultado_json') or '{}')
        return jsonify({'items': resultado.get(tipo, []), 'tipo': tipo, 'cruce_id': row['id']}), 200

    @bp.route('/descargar/<int:cruce_id>/<tipo>/<formato>', methods=['GET'])
    def descargar(cruce_id: int, tipo: str, formato: str):
        ctx = user_ctx()
        row = repo.fetch_one("SELECT * FROM cb_cruces WHERE id = ?", (cruce_id,))
        if not row:
            return jsonify({'error': 'Cruce no encontrado.'}), 404
        if ctx['rol'] != 'SUPERADMIN' and int(row.get('fundacion_id') or 0) != int(ctx['fundacion_id'] or 0):
            return jsonify({'error': 'No tienes permiso para descargar este cruce.'}), 403
        resultado = json.loads(row.get('resultado_json') or '{}')
        if tipo != 'resumen':
            resultado = {'resumen': row, tipo: resultado.get(tipo, [])}
        if formato.lower() == 'pdf':
            path = generar_pdf_resultado(resultado, module_reports, tipo, f'CRUCE_{cruce_id}')
        else:
            path = generar_excel_resultado(resultado, module_reports, f'CRUCE_{cruce_id}_{tipo}')
        return send_from_directory(module_reports, os.path.basename(path), as_attachment=True)

    @bp.route('/opciones-informe', methods=['GET'])
    def opciones_informe_estadistico():
        """Opciones disponibles para filtrar el informe Word/PDF sin tocar la Base Maestra."""
        ctx = user_ctx()
        try:
            from .informe_estadistico import obtener_opciones_informe
            opciones = obtener_opciones_informe(
                database_path,
                fundacion_id=ctx['fundacion_id'],
                superadmin=ctx['rol'] == 'SUPERADMIN',
            )
            return jsonify(opciones), 200
        except Exception as exc:
            return jsonify({'error': f'No se pudieron cargar las opciones del informe: {exc}'}), 500

    @bp.route('/informe-estadistico/<formato>', methods=['GET'])
    def generar_informe_estadistico_descargable(formato: str):
        """Genera informe institucional descargable en Word o PDF después del cruce.

        No modifica datos; lee el último cruce o el cruce_id indicado y consulta la
        Base Maestra consolidada para estadísticas, gráficas y anexos.
        """
        formato = (formato or 'docx').lower().strip()
        if formato not in {'docx', 'word', 'pdf'}:
            return jsonify({'error': 'Formato no permitido. Usa docx o pdf.'}), 400

        ctx = user_ctx()
        cruce_id = request.args.get('cruce_id', type=int)
        if cruce_id:
            row = repo.fetch_one("SELECT * FROM cb_cruces WHERE id = ?", (cruce_id,))
        else:
            row = repo.ultimo_cruce(ctx['fundacion_id'], superadmin=ctx['rol'] == 'SUPERADMIN')
        if not row:
            return jsonify({'error': 'No hay cruce de bases disponible. Ejecuta primero el cruce mensual.'}), 404
        if ctx['rol'] != 'SUPERADMIN' and int(row.get('fundacion_id') or 0) != int(ctx['fundacion_id'] or 0):
            return jsonify({'error': 'No tienes permiso para generar informe de este cruce.'}), 403

        try:
            resultado = json.loads(row.get('resultado_json') or '{}')
        except Exception:
            resultado = {}

        filtros = {
            'alcance': request.args.get('alcance') or request.args.get('tipo') or 'general',
            'unidad': request.args.get('unidad') or '',
            'coordinador': request.args.get('coordinador') or '',
            'grupo_etario': request.args.get('grupo_etario') or request.args.get('grupo') or '',
            'estado_nutricional': request.args.get('estado_nutricional') or request.args.get('diagnostico') or '',
            'alertas': request.args.get('alertas') or '',
            'faltantes': request.args.get('faltantes') or request.args.get('datos_faltantes') or '',
        }

        try:
            from .informe_estadistico import crear_informe_estadistico
            path = crear_informe_estadistico(
                database_path=database_path,
                output_folder=module_reports,
                row_cruce=row,
                resultado=resultado,
                filtros=filtros,
                usuario_ctx=ctx,
                formato='docx' if formato == 'word' else formato,
            )
            return send_from_directory(module_reports, os.path.basename(path), as_attachment=True)
        except Exception as exc:
            return jsonify({'error': f'No se pudo generar el informe estadístico: {exc}'}), 500

    def usuarios_por_unidad(unidad: str, unidades_param: str | None = None) -> list[dict]:
        ctx = user_ctx()
        user = ctx['raw']
        allowed_units = allowed_units_for_user(database_path, user)
        unidades = []
        if unidades_param:
            unidades = [normalize_unidad(u) for u in unidades_param.split('|') if normalize_unidad(u)]
        elif unidad and unidad != '__all__':
            unidades = [normalize_unidad(unidad)]
        if allowed_units is not None:
            if unidades:
                unidades = [u for u in unidades if u in allowed_units]
            else:
                unidades = list(allowed_units)
        # La versión publicada de la Base Maestra es la fuente autoritativa.
        # Una UDS vacía en esa versión debe permanecer vacía: volver a
        # ``usuarios`` en ese caso reintroduce personas retiradas y duplicados.
        version_activa = repo.fetch_one(
            'SELECT id FROM master_versiones '
            'WHERE fundacion_id=? AND activa=1 ORDER BY id DESC LIMIT 1',
            [ctx['fundacion_id']],
        )
        fuente = 'master_ninos' if version_activa else 'usuarios'
        cols = repo.columns(fuente)
        where = []
        params = []
        if fuente == 'master_ninos':
            where.extend(['fundacion_id = ?', 'activo = 1'])
            params.append(ctx['fundacion_id'])
        elif 'fundacion_id' in cols and ctx['rol'] != 'SUPERADMIN':
            where.append('(fundacion_id = ? OR fundacion_id IS NULL)')
            params.append(ctx['fundacion_id'])
        if unidades:
            placeholders = ','.join(['?'] * len(unidades))
            campo_unidad = 'unidad_servicio' if fuente == 'master_ninos' else 'unidad'
            where.append(f'{campo_unidad} IN ({placeholders})')
            params.extend(unidades)
        if fuente == 'master_ninos':
            sql = (
                'SELECT *, unidad_servicio AS unidad, nombre_completo AS nombre '
                'FROM master_ninos'
            )
        else:
            sql = 'SELECT * FROM usuarios'
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += (
            ' ORDER BY unidad_servicio, nombre_completo'
            if fuente == 'master_ninos' else
            ' ORDER BY unidad, nombre'
        )
        rows = repo.fetch_all(sql, params)
        for r in rows:
            docente = docente_por_unidad(database_path, r.get('unidad'), ctx['fundacion_id'])
            r['docente_asignado'] = r.get('docente') or docente.get('nombre') or 'Sin docente asignado'
            r['coordinador'] = r.get('coordinador') or docente.get('coordinador') or ''
            try:
                meses = int(r.get('edad_meses') or 0)
                r['edad'] = f"{meses // 12} años {meses % 12} meses" if meses else ''
            except Exception:
                r['edad'] = r.get('edad_completa') or ''
        return rows

    def exportar_usuarios_excel(rows: list[dict], output_folder: str, unidad_label: str) -> str:
        """Exporta usuarios por UDS con colores por grupo etario, leyenda y resumen.
        Alpha70: mejora visual no invasiva; no modifica datos ni rutas.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import Counter, defaultdict

        def valor(row, *keys):
            for key in keys:
                val = row.get(key)
                if val not in (None, ''):
                    return val
            return ''

        def normalizar_grupo(row):
            raw = str(valor(row, 'grupo_etario', 'grupo_edad', 'GrupoEdad', 'tipo_beneficiario') or '').strip()
            edad_meses = 0
            try:
                edad_meses = int(float(valor(row, 'edad_meses', 'EdadMeses') or 0))
            except Exception:
                edad_meses = 0
            raw_norm = (raw or '').lower()
            if 'gestante' in raw_norm:
                return '0 a 6 meses y gestantes'
            if '6' in raw_norm and '11' in raw_norm:
                return '6 a 11 meses'
            if '1' in raw_norm and '2' in raw_norm:
                return '1 a 2 años'
            if '3' in raw_norm and '5' in raw_norm:
                return '3 a 5 años'
            if edad_meses:
                if edad_meses <= 6:
                    return '0 a 6 meses y gestantes'
                if edad_meses <= 11:
                    return '6 a 11 meses'
                if edad_meses <= 35:
                    return '1 a 2 años'
                if edad_meses <= 71:
                    return '3 a 5 años'
            return 'SIN GRUPO'

        def nombre_completo(row):
            nombre = valor(row, 'nombre_completo', 'nombre', 'Nombre', 'nombres', 'Nombres')
            apellidos = valor(row, 'apellidos', 'Apellidos')
            if nombre and apellidos and str(apellidos).lower() not in str(nombre).lower():
                return f'{nombre} {apellidos}'.strip()
            return str(nombre or '').strip()

        def nombres_y_apellidos(row):
            nombre = valor(row, 'nombres', 'Nombres', 'nombre', 'Nombre', 'nombre_completo')
            apellidos = valor(row, 'apellidos', 'Apellidos')
            return nombre, apellidos

        colores = {
            '0 a 6 meses y gestantes': 'FFF2CC',  # amarillo suave
            '6 a 11 meses': 'D9EAF7',             # azul suave
            '1 a 2 años': 'D9EAD3',               # verde suave
            '3 a 5 años': 'FCE4D6',               # naranja suave
            'SIN GRUPO': 'D9D9D9',                # gris suave
        }
        rangos = {
            '0 a 6 meses y gestantes': '0 a 6 meses y mujeres gestantes/lactantes si aplican',
            '6 a 11 meses': '7 a 11 meses y 29 días',
            '1 a 2 años': '12 a 35 meses',
            '3 a 5 años': '36 a 71 meses',
            'SIN GRUPO': 'Sin edad, grupo o dato suficiente',
        }

        mes = request.args.get('mes') or datetime.now().month
        anio = request.args.get('anio') or request.args.get('año') or datetime.now().year
        wb = Workbook()
        ws = wb.active
        ws.title = 'Usuarios por grupo'
        headers = ['Documento', 'Tipo documento', 'Nombres', 'Apellidos', 'Nombre completo', 'Edad', 'Edad meses', 'Grupo etario', 'UDS', 'Docente', 'Coordinador', 'Estado', 'Observaciones']
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E78')
        header_font = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        vistos = set()
        filas_unicas = []
        for r in rows:
            doc = str(valor(r, 'documento', 'Documento', 'nui', 'NUI') or '').strip()
            key = doc or f"{nombre_completo(r)}|{valor(r, 'unidad', 'Unidad')}"
            if key in vistos:
                continue
            vistos.add(key)
            filas_unicas.append(r)

        conteo_grupo = Counter()
        conteo_uds = Counter()
        conteo_docente = Counter()
        conteo_coord = Counter()

        for r in filas_unicas:
            grupo = normalizar_grupo(r)
            conteo_grupo[grupo] += 1
            uds = valor(r, 'unidad', 'Unidad') or unidad_label or ''
            docente = valor(r, 'docente_asignado', 'docente', 'Docente', 'agente_educativo')
            coord = valor(r, 'coordinador', 'Coordinador')
            conteo_uds[uds or 'SIN UDS'] += 1
            if docente:
                conteo_docente[docente] += 1
            if coord:
                conteo_coord[coord] += 1
            nombres, apellidos = nombres_y_apellidos(r)
            row = [
                valor(r, 'documento', 'Documento', 'nui', 'NUI'),
                valor(r, 'tipo_documento', 'TipoDocumento', 'tipo_doc'),
                nombres,
                apellidos,
                nombre_completo(r),
                valor(r, 'edad', 'Edad', 'edad_completa', 'EdadCompleta'),
                valor(r, 'edad_meses', 'EdadMeses'),
                grupo,
                uds,
                docente,
                coord,
                valor(r, 'estado', 'Estado'),
                valor(r, 'observaciones', 'observacion', 'Observaciones'),
            ]
            ws.append(row)
            fill = PatternFill('solid', fgColor=colores.get(grupo, colores['SIN GRUPO']))
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = [16,16,22,22,30,18,14,24,24,26,24,16,34][col-1]
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Hoja resumen
        ws_res = wb.create_sheet('Resumen')
        ws_res.append(['Resumen de usuarios exportados'])
        ws_res['A1'].font = Font(bold=True, size=14)
        ws_res.append(['Unidad solicitada', unidad_label or 'Todas'])
        ws_res.append(['Mes', mes])
        ws_res.append(['Año', anio])
        ws_res.append(['Fecha de generación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_res.append(['Total usuarios únicos', len(filas_unicas)])
        ws_res.append([])
        ws_res.append(['Grupo etario', 'Cantidad'])
        for cell in ws_res[8]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        for grupo, total in sorted(conteo_grupo.items()):
            ws_res.append([grupo, total])
        base_row = ws_res.max_row + 2
        ws_res.cell(base_row, 1, 'UDS')
        ws_res.cell(base_row, 2, 'Cantidad')
        ws_res.cell(base_row, 1).font = ws_res.cell(base_row, 2).font = Font(bold=True)
        for uds, total in sorted(conteo_uds.items()):
            ws_res.append([uds, total])
        base_row = ws_res.max_row + 2
        ws_res.cell(base_row, 1, 'Docente')
        ws_res.cell(base_row, 2, 'Cantidad')
        ws_res.cell(base_row, 1).font = ws_res.cell(base_row, 2).font = Font(bold=True)
        for docente, total in sorted(conteo_docente.items()):
            ws_res.append([docente, total])
        base_row = ws_res.max_row + 2
        ws_res.cell(base_row, 1, 'Coordinador')
        ws_res.cell(base_row, 2, 'Cantidad')
        ws_res.cell(base_row, 1).font = ws_res.cell(base_row, 2).font = Font(bold=True)
        for coord, total in sorted(conteo_coord.items()):
            ws_res.append([coord, total])
        ws_res.column_dimensions['A'].width = 34
        ws_res.column_dimensions['B'].width = 18

        # Hoja leyenda
        ws_leg = wb.create_sheet('Leyenda grupos etarios')
        ws_leg.append(['Color', 'Grupo etario', 'Rango de edad', 'Cantidad de usuarios'])
        for cell in ws_leg[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        for grupo in ['0 a 6 meses y gestantes', '6 a 11 meses', '1 a 2 años', '3 a 5 años', 'SIN GRUPO']:
            ws_leg.append(['', grupo, rangos[grupo], conteo_grupo.get(grupo, 0)])
            ws_leg.cell(ws_leg.max_row, 1).fill = PatternFill('solid', fgColor=colores[grupo])
            for cell in ws_leg[ws_leg.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws_leg.column_dimensions['A'].width = 12
        ws_leg.column_dimensions['B'].width = 28
        ws_leg.column_dimensions['C'].width = 42
        ws_leg.column_dimensions['D'].width = 18

        os.makedirs(output_folder, exist_ok=True)
        path = os.path.join(output_folder, f"USUARIOS_COLORES_{secure_filename(unidad_label or 'UNIDADES')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
        wb.save(path)
        return path

    def exportar_usuarios_pdf(rows: list[dict], output_folder: str, unidad_label: str) -> str:
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        path = os.path.join(output_folder, f"USUARIOS_{secure_filename(unidad_label or 'UNIDADES')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
        doc = SimpleDocTemplate(path, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [Paragraph(f'Usuarios por unidad: {unidad_label}', styles['Title']), Spacer(1, 8)]
        headers = ['Unidad', 'Docente', 'Documento', 'Nombre', 'Acudiente', 'Teléfono', 'Estado']
        data = [headers]
        for r in rows[:500]:
            data.append([r.get('unidad',''), r.get('docente_asignado',''), r.get('documento',''), r.get('nombre',''), r.get('nombre_acudiente',''), r.get('telefono',''), r.get('estado','')])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), .25, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(table)
        doc.build(story)
        return path

    @bp.route('/usuarios-docente/<path:docente>/<formato>', methods=['GET'])
    def descargar_usuarios_docente(docente: str, formato: str):
        rows = [r for r in usuarios_por_unidad('__all__', None) if (docente or '').lower() in (r.get('docente_asignado') or '').lower()]
        etiqueta = f'DOCENTE_{docente}'
        if formato.lower() == 'pdf':
            path = exportar_usuarios_pdf(rows, module_reports, etiqueta)
        else:
            path = exportar_usuarios_excel(rows, module_reports, etiqueta)
        return send_from_directory(module_reports, os.path.basename(path), as_attachment=True)

    @bp.route('/usuarios-coordinador/<path:coordinador>/<formato>', methods=['GET'])
    def descargar_usuarios_coordinador(coordinador: str, formato: str):
        rows = [r for r in usuarios_por_unidad('__all__', None) if (coordinador or '').lower() in (r.get('coordinador') or '').lower()]
        etiqueta = f'COORDINADOR_{coordinador}'
        if formato.lower() == 'pdf':
            path = exportar_usuarios_pdf(rows, module_reports, etiqueta)
        else:
            path = exportar_usuarios_excel(rows, module_reports, etiqueta)
        return send_from_directory(module_reports, os.path.basename(path), as_attachment=True)

    @bp.route('/usuarios-unidad/<path:unidad>/<formato>', methods=['GET'])
    def descargar_usuarios_unidad(unidad: str, formato: str):
        if formato.lower() in {'imprimir', 'print', 'html'}:
            return imprimir_usuarios_unidad(unidad)
        rows = usuarios_por_unidad(unidad, request.args.get('unidades'))
        if formato.lower() == 'pdf':
            path = exportar_usuarios_pdf(rows, module_reports, unidad)
        else:
            path = exportar_usuarios_excel(rows, module_reports, unidad)
        return send_from_directory(module_reports, os.path.basename(path), as_attachment=True)

    @bp.route('/usuarios-unidad/<path:unidad>/imprimir', methods=['GET'])
    def imprimir_usuarios_unidad(unidad: str):
        rows = usuarios_por_unidad(unidad, request.args.get('unidades'))
        filas = ''.join(
            f"<tr><td>{r.get('unidad','')}</td><td>{r.get('docente_asignado','')}</td><td>{r.get('documento','')}</td><td>{r.get('nombre','')}</td><td>{r.get('fecha_nacimiento','')}</td><td>{r.get('sexo','')}</td><td>{r.get('nombre_acudiente','')}</td><td>{r.get('telefono','')}</td><td>{r.get('estado','')}</td></tr>"
            for r in rows
        )
        html = f"""
        <!doctype html><html><head><meta charset='utf-8'><title>Usuarios {unidad}</title>
        <style>body{{font-family:Arial,sans-serif;margin:24px;color:#111}} h1{{font-size:20px}} table{{width:100%;border-collapse:collapse;font-size:11px}} th,td{{border:1px solid #999;padding:5px;text-align:left}} th{{background:#eee}} @media print{{button{{display:none}}}}</style>
        </head><body><button onclick='window.print()'>Imprimir</button><h1>Usuarios por unidad: {unidad}</h1><p>Total: {len(rows)}</p><table><thead><tr><th>Unidad</th><th>Docente</th><th>Documento</th><th>Nombre</th><th>Fecha nac.</th><th>Sexo</th><th>Acudiente</th><th>Teléfono</th><th>Estado</th></tr></thead><tbody>{filas}</tbody></table></body></html>
        """
        return Response(html, mimetype='text/html')


    # ALPHA32: Logo institucional por corporación para informes Word/PDF.
    # Ruta independiente y no destructiva; no toca login, menú, RPP ni Bienestarina.
    bp_logo = Blueprint('corporaciones_logo_alpha32', __name__, url_prefix='/api/corporaciones')

    @bp_logo.route('/logo', methods=['POST'])
    def subir_logo_corporacion():
        ctx = user_ctx()
        logo_file = request.files.get('logo') or request.files.get('file')
        try:
            from .informe_estadistico import registrar_logo_corporacion
            result = registrar_logo_corporacion(database_path, logo_file, upload_folder, ctx)
            return jsonify(result), 200
        except Exception as exc:
            return jsonify({'error': str(exc)}), 400

    @bp_logo.route('/logo', methods=['GET'])
    def obtener_logo_corporacion_actual():
        ctx = user_ctx()
        try:
            from .informe_estadistico import obtener_corporacion_contexto, obtener_logo_corporacion
            corp = obtener_corporacion_contexto(database_path, ctx.get('fundacion_id') or 1, ctx.get('corporacion_id'))
            logo = obtener_logo_corporacion(database_path, ctx.get('fundacion_id') or 1, ctx.get('corporacion_id'), module_reports)
            return jsonify({'corporacion': corp, 'logo_path': logo}), 200
        except Exception as exc:
            return jsonify({'error': str(exc)}), 500

    app.register_blueprint(bp)
    try:
        app.register_blueprint(bp_logo)
    except ValueError:
        # Evita fallo si el módulo se registra dos veces en pruebas de fábrica.
        pass
