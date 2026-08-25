"""Persistencia e integración del Centro Inteligente de Planeación."""
from __future__ import annotations

import json
import mimetypes
import os
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from modules.dbapi_compat import sqlite3
from modules.seguridad.tenant_context import tenant_storage_root

from .schema import SCHEMA_SQL, SCHEMA_VERSION
from .services import (
    COMPLETED_STATES,
    COORDINATION_ROLES,
    file_sha256,
    json_dump,
    normalize,
    now_iso,
    parse_json,
    semaforo,
    source_key,
    unit_key,
)


DEFAULT_RULES = [
    {
        "codigo": "ACTIVIDAD_GENERAL",
        "nombre": "Actividad operativa general",
        "descripcion": "Agenda y recordatorios básicos.",
        "componente": "ADMINISTRATIVO_GESTION",
        "tipo_actividad": "GENERAL",
        "rol_responsable": None,
        "dias_recordatorio": [7, 2, 0],
        "documentos": ["AGENDA"],
        "evidencias": [],
        "prioridad_base": "MEDIA",
    },
    {
        "codigo": "JORNADA_SALUD",
        "nombre": "Jornada de Salud y Nutrición",
        "descripcion": "Prepara agenda, acta y listado de asistencia.",
        "componente": "SALUD_NUTRICION",
        "tipo_actividad": "JORNADA",
        "rol_responsable": "NUTRICIONISTA",
        "dias_recordatorio": [7, 2, 0],
        "documentos": ["AGENDA", "ACTA", "LISTADO_ASISTENCIA", "INFORME"],
        "evidencias": ["ACTA", "LISTADO", "EVIDENCIA_TECNICA"],
        "prioridad_base": "ALTA",
    },
    {
        "codigo": "ACTIVIDAD_PSICOSOCIAL",
        "nombre": "Actividad psicosocial y familiar",
        "descripcion": "Visitas, escuelas de familia y encuentros comunitarios.",
        "componente": "FAMILIA_COMUNIDAD_REDES",
        "tipo_actividad": "PSICOSOCIAL",
        "rol_responsable": "PSICOSOCIAL",
        "dias_recordatorio": [7, 2, 0],
        "documentos": ["AGENDA", "ACTA", "LISTADO_ASISTENCIA", "INFORME"],
        "evidencias": ["ACTA", "LISTADO", "EVIDENCIA"],
        "prioridad_base": "ALTA",
    },
    {
        "codigo": "ACTIVIDAD_PEDAGOGICA",
        "nombre": "Actividad pedagógica",
        "descripcion": "Encuentros y actividades del componente pedagógico.",
        "componente": "PROCESO_PEDAGOGICO",
        "tipo_actividad": "PEDAGOGICA",
        "rol_responsable": "DOCENTE",
        "dias_recordatorio": [7, 2, 0],
        "documentos": ["AGENDA", "ACTA", "LISTADO_ASISTENCIA"],
        "evidencias": ["ACTA", "LISTADO", "EVIDENCIA_PEDAGOGICA"],
        "prioridad_base": "MEDIA",
    },
]


class CentroPlaneacionRepository:
    def __init__(self, database_path: str, data_dir: str, output_folder: str):
        self.database_path = str(database_path)
        self.data_dir = Path(data_dir).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def connect(self):
        conn = sqlite3.connect(self.database_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn

    @staticmethod
    def _table_exists(conn, table: str) -> bool:
        return bool(conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone())

    @staticmethod
    def _columns(conn, table: str) -> set[str]:
        if not CentroPlaneacionRepository._table_exists(conn, table):
            return set()
        return {str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()}

    def init_schema(self) -> None:
        with self.connect() as conn:
            # El calendario histórico es la tabla canónica de fecha/estado.
            if not self._table_exists(conn, "calendario_entregables"):
                conn.execute(
                    """
                    CREATE TABLE calendario_entregables (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        titulo TEXT NOT NULL, descripcion TEXT, fecha_inicio TEXT,
                        fecha_limite TEXT NOT NULL, modulo TEXT, tipo_formato TEXT,
                        responsable_id INTEGER, responsable_nombre TEXT, coordinador TEXT,
                        unidad TEXT, municipio TEXT, estado TEXT DEFAULT 'pendiente',
                        prioridad TEXT DEFAULT 'Media', color TEXT DEFAULT 'azul',
                        requiere_evidencia INTEGER DEFAULT 0, archivo_evidencia TEXT,
                        fecha_entrega TEXT, observaciones TEXT, creado_por TEXT,
                        fecha_creacion TEXT, actualizado_en TEXT, fundacion_id INTEGER DEFAULT 1,
                        usuario_creador_id INTEGER, clave_unica TEXT, origen TEXT DEFAULT 'manual'
                    )
                    """
                )
            # Completa de forma idempotente las columnas que pueden faltar en
            # calendarios creados por versiones históricas. No cambia ni borra datos.
            calendar_columns = self._columns(conn, "calendario_entregables")
            required_calendar_columns = {
                "descripcion": "TEXT", "fecha_inicio": "TEXT", "fecha_limite": "TEXT",
                "modulo": "TEXT", "tipo_formato": "TEXT", "responsable_id": "INTEGER",
                "responsable_nombre": "TEXT", "coordinador": "TEXT", "unidad": "TEXT",
                "municipio": "TEXT", "estado": "TEXT DEFAULT 'pendiente'",
                "prioridad": "TEXT DEFAULT 'Media'", "color": "TEXT DEFAULT 'azul'",
                "requiere_evidencia": "INTEGER DEFAULT 0", "archivo_evidencia": "TEXT",
                "fecha_entrega": "TEXT", "observaciones": "TEXT", "creado_por": "TEXT",
                "fecha_creacion": "TEXT", "actualizado_en": "TEXT",
                "fundacion_id": "INTEGER DEFAULT 1", "usuario_creador_id": "INTEGER",
                "clave_unica": "TEXT", "origen": "TEXT DEFAULT 'manual'",
            }
            for column, definition in required_calendar_columns.items():
                if column not in calendar_columns:
                    conn.execute(f'ALTER TABLE calendario_entregables ADD COLUMN "{column}" {definition}')
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_calendario_entregables_clave ON calendario_entregables(clave_unica)")
            conn.executescript(SCHEMA_SQL)
            conn.execute(
                "INSERT INTO cpo_schema_version(id,version,fecha_actualizacion) VALUES(1,?,?) "
                "ON CONFLICT(id) DO UPDATE SET version=excluded.version,fecha_actualizacion=excluded.fecha_actualizacion",
                (SCHEMA_VERSION, now_iso()),
            )
            conn.commit()

    def audit(self, fundacion_id: int, user: dict[str, Any], action: str, entity: str, entity_id: int | None, detail: dict[str, Any] | None = None, conn=None) -> None:
        own = conn is None
        target = conn or self.connect()
        try:
            target.execute(
                "INSERT INTO cpo_auditoria(fundacion_id,usuario_id,usuario,accion,entidad,entidad_id,detalle_json,fecha) VALUES(?,?,?,?,?,?,?,?)",
                (fundacion_id, user.get("id"), user.get("username") or user.get("email") or "sistema", action, entity, entity_id, json_dump(detail or {}), now_iso()),
            )
            if own:
                target.commit()
        finally:
            if own:
                target.close()

    def ensure_rules(self, fundacion_id: int, user_id: int | None = None) -> None:
        now = now_iso()
        with self.connect() as conn:
            for item in DEFAULT_RULES:
                conn.execute(
                    """
                    INSERT INTO cpo_reglas_operativas
                    (fundacion_id,codigo,nombre,descripcion,componente,tipo_actividad,rol_responsable,
                     dias_recordatorio_json,documentos_json,evidencias_json,condicion_cierre_json,
                     prioridad_base,activa,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?, ?,1,?,?,?,?)
                    ON CONFLICT(fundacion_id,codigo) DO NOTHING
                    """,
                    (
                        fundacion_id, item["codigo"], item["nombre"], item["descripcion"], item["componente"],
                        item["tipo_actividad"], item["rol_responsable"], json_dump(item["dias_recordatorio"]),
                        json_dump(item["documentos"]), json_dump(item["evidencias"]), json_dump({"revision_humana": True}),
                        item["prioridad_base"], user_id, user_id, now, now,
                    ),
                )
            conn.commit()

    def list_rules(self, fundacion_id: int, active_only: bool = False) -> list[dict[str, Any]]:
        with self.connect() as conn:
            where = "fundacion_id=?" + (" AND activa=1" if active_only else "")
            rows = conn.execute(f"SELECT * FROM cpo_reglas_operativas WHERE {where} ORDER BY activa DESC, componente, codigo", (fundacion_id,)).fetchall()
        return [self._rule(dict(row)) for row in rows]

    @staticmethod
    def _rule(data: dict[str, Any]) -> dict[str, Any]:
        for key in ("dias_recordatorio_json", "documentos_json", "evidencias_json", "condicion_cierre_json"):
            data[key.replace("_json", "")] = parse_json(data.get(key), [] if key != "condicion_cierre_json" else {})
        return data

    def create_rule(self, fundacion_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        code = normalize(data.get("codigo"))
        name = str(data.get("nombre") or "").strip()
        if not code or not name:
            raise ValueError("Código y nombre son obligatorios.")
        now = now_iso()
        with self.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO cpo_reglas_operativas
                (fundacion_id,codigo,nombre,descripcion,componente,tipo_actividad,rol_responsable,
                 dias_recordatorio_json,documentos_json,evidencias_json,condicion_cierre_json,prioridad_base,
                 activa,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (fundacion_id, code, name, data.get("descripcion"), normalize(data.get("componente")), normalize(data.get("tipo_actividad")),
                 normalize(data.get("rol_responsable")) or None, json_dump(data.get("dias_recordatorio") or [7,2,0]),
                 json_dump(data.get("documentos") or []), json_dump(data.get("evidencias") or []),
                 json_dump(data.get("condicion_cierre") or {"revision_humana": True}), normalize(data.get("prioridad_base")) or "MEDIA",
                 1 if data.get("activa", True) else 0, user.get("id"), user.get("id"), now, now),
            )
            rule_id = int(cur.lastrowid)
            self.audit(fundacion_id, user, "CREAR_REGLA", "cpo_reglas_operativas", rule_id, {"codigo": code}, conn)
            conn.commit()
        return next(row for row in self.list_rules(fundacion_id) if row["id"] == rule_id)

    def _rule_for(self, conn, fundacion_id: int, component: str, activity_type: str, source_table: str) -> dict[str, Any] | None:
        component = normalize(component)
        activity_type = normalize(activity_type)
        code = "ACTIVIDAD_GENERAL"
        if component == "SALUD_NUTRICION" or source_table.startswith("sn_"):
            code = "JORNADA_SALUD"
        elif component == "FAMILIA_COMUNIDAD_REDES" or source_table.startswith(("fcr_", "ps_")):
            code = "ACTIVIDAD_PSICOSOCIAL"
        elif component == "PROCESO_PEDAGOGICO" or source_table.startswith(("gp_", "pp_")):
            code = "ACTIVIDAD_PEDAGOGICA"
        row = conn.execute("SELECT * FROM cpo_reglas_operativas WHERE fundacion_id=? AND codigo=? AND activa=1", (fundacion_id, code)).fetchone()
        if not row:
            row = conn.execute("SELECT * FROM cpo_reglas_operativas WHERE fundacion_id=? AND activa=1 ORDER BY id LIMIT 1", (fundacion_id,)).fetchone()
        return self._rule(dict(row)) if row else None

    @staticmethod
    def _pick(row: dict[str, Any], *names: str, default=None):
        for name in names:
            if name in row and row.get(name) not in (None, ""):
                return row.get(name)
        return default

    def _source_specs(self) -> list[dict[str, Any]]:
        return [
            {
                "table": "mgp_tareas", "module": "MOTOR_GESTION", "id": "id", "title": ("titulo",),
                "description": ("descripcion",), "start": ("fecha_inicio",), "due": ("fecha_limite",),
                "state": ("estado",), "priority": ("prioridad",), "responsible_id": ("responsable_id",),
                "responsible_name": ("responsable_nombre",), "unit": ("unidad_nombre",), "unit_id": ("unidad_id",),
                "component": ("componente",), "type": ("tipo_tarea",), "expediente": ("expediente_id",),
                "source_key": ("fuente_clave",), "active": ("activa",),
            },
            {
                "table": "sn_actividades_integrales", "module": "SALUD_NUTRICION", "id": "id", "title": ("titulo",),
                "description": ("objetivo", "metodologia"), "start": ("fecha_programada",), "due": ("fecha_programada",),
                "state": ("estado",), "priority": (), "responsible_id": ("responsable_id",), "responsible_name": ("responsable_nombre",),
                "unit": ("unidad_nombre",), "unit_id": (), "component": (), "type": ("tipo_actividad",), "expediente": ("expediente_uca_id",),
            },
            {
                "table": "fcr_actividades", "module": "FAMILIAS_REDES", "id": "id", "title": ("titulo",),
                "description": ("objetivo", "metodologia"), "start": ("fecha_programada",), "due": ("fecha_limite_cierre", "fecha_programada"),
                "state": ("estado",), "priority": (), "responsible_id": ("profesional_id",), "responsible_name": ("profesional_nombre",),
                "unit": ("unidad_nombre",), "unit_id": ("unidad_id",), "component": (), "type": ("tipo",), "expediente": ("expediente_uca_id",),
            },
            {
                "table": "ps_acciones_plan", "module": "PSICOSOCIAL", "id": "id", "title": ("titulo",),
                "description": ("descripcion",), "start": ("fecha_inicio",), "due": ("fecha_limite",),
                "state": ("estado",), "priority": ("prioridad",), "responsible_id": ("responsable_id",), "responsible_name": ("responsable_nombre",),
                "unit": ("unidad_nombre",), "unit_id": (), "component": (), "type": (), "expediente": ("expediente_uca_id",),
            },
            {
                "table": "gp_calendario_eventos", "module": "GESTION_PEDAGOGICA", "id": "id", "title": ("titulo",),
                "description": ("descripcion",), "start": ("fecha",), "due": ("fecha",), "state": ("estado",),
                "priority": (), "responsible_id": ("coordinador_id",), "responsible_name": (), "unit": (), "unit_id": (),
                "component": (), "type": ("tipo",), "expediente": (),
            },
            {
                "table": "pp_actividades", "module": "PLANEACION_PEDAGOGICA", "id": "id", "title": ("titulo", "nombre", "actividad"),
                "description": ("descripcion", "objetivo"), "start": ("fecha", "fecha_programada"), "due": ("fecha", "fecha_programada"),
                "state": ("estado",), "priority": (), "responsible_id": ("responsable_id", "docente_id"), "responsible_name": ("responsable",),
                "unit": ("unidad", "unidad_nombre"), "unit_id": (), "component": (), "type": ("tipo_actividad",), "expediente": (),
            },
        ]

    def _component_for(self, spec: dict[str, Any], row: dict[str, Any]) -> str:
        value = self._pick(row, *spec.get("component", ()))
        if value:
            return normalize(value)
        table = spec["table"]
        if table.startswith("sn_"):
            return "SALUD_NUTRICION"
        if table.startswith(("fcr_", "ps_")):
            return "FAMILIA_COMUNIDAD_REDES"
        if table.startswith(("gp_", "pp_")):
            return "PROCESO_PEDAGOGICO"
        return "ADMINISTRATIVO_GESTION"

    def _read_sources(self, conn, fundacion_id: int) -> Iterable[tuple[dict[str, Any], dict[str, Any]]]:
        for spec in self._source_specs():
            table = spec["table"]
            if not self._table_exists(conn, table):
                continue
            cols = self._columns(conn, table)
            where = []
            params: list[Any] = []
            if "fundacion_id" in cols:
                where.append("fundacion_id=?")
                params.append(fundacion_id)
            if spec.get("active") and any(name in cols for name in spec["active"]):
                active_col = next(name for name in spec["active"] if name in cols)
                where.append(f"COALESCE({active_col},1)=1")
            sql = f'SELECT * FROM "{table}"' + (" WHERE " + " AND ".join(where) if where else "")
            direct_source_tables = {item["table"] for item in self._source_specs() if item["table"] != "mgp_tareas"}
            for raw in conn.execute(sql, params).fetchall():
                row = dict(raw)
                if not row.get(spec["id"]):
                    continue
                # El Motor de Gestión ya contiene referencias a actividades de
                # otros módulos. Cuando la fuente misional también se consulta
                # directamente, se descarta la referencia del motor para evitar
                # mostrar dos veces la misma actividad.
                if table == "mgp_tareas":
                    source_table = str(row.get("fuente_tabla") or "").strip()
                    source_id = row.get("fuente_id")
                    if source_table in direct_source_tables and source_id and self._table_exists(conn, source_table):
                        source_columns = self._columns(conn, source_table)
                        source_where = ["id=?"]
                        source_params: list[Any] = [source_id]
                        if "fundacion_id" in source_columns:
                            source_where.append("fundacion_id=?")
                            source_params.append(fundacion_id)
                        if conn.execute(
                            f'SELECT 1 FROM "{source_table}" WHERE ' + " AND ".join(source_where) + " LIMIT 1",
                            source_params,
                        ).fetchone():
                            continue
                yield spec, row

    def synchronize(self, fundacion_id: int, user: dict[str, Any]) -> dict[str, Any]:
        self.ensure_rules(fundacion_id, user.get("id"))
        created = updated = sources = 0
        now = now_iso()
        with self.connect() as conn:
            for spec, row in self._read_sources(conn, fundacion_id):
                sources += 1
                table = spec["table"]
                source_id = int(row[spec["id"]])
                component = self._component_for(spec, row)
                activity_type = normalize(self._pick(row, *spec.get("type", ()), default="GENERAL")) or "GENERAL"
                rule = self._rule_for(conn, fundacion_id, component, activity_type, table)
                unique = self._pick(row, *spec.get("source_key", ())) or source_key(table, source_id)
                calendar_key = f"cpo:{fundacion_id}:{table}:{unique}"
                title = str(self._pick(row, *spec.get("title", ()), default=f"Actividad {table} #{source_id}")).strip()
                description = str(self._pick(row, *spec.get("description", ()), default="") or "").strip()
                start = str(self._pick(row, *spec.get("start", ()), default="") or "")[:10] or None
                due = str(self._pick(row, *spec.get("due", ()), default=start or date.today().isoformat()) or date.today().isoformat())[:10]
                state = normalize(self._pick(row, *spec.get("state", ()), default="PENDIENTE")) or "PENDIENTE"
                priority = normalize(self._pick(row, *spec.get("priority", ()), default=(rule or {}).get("prioridad_base", "MEDIA"))) or "MEDIA"
                responsible_id = self._pick(row, *spec.get("responsible_id", ()))
                responsible_name = self._pick(row, *spec.get("responsible_name", ()))
                unit_name = str(self._pick(row, *spec.get("unit", ()), default="") or "").strip() or None
                unit_id = self._pick(row, *spec.get("unit_id", ()))
                expediente_id = self._pick(row, *spec.get("expediente", ()))
                existing = conn.execute("SELECT id FROM calendario_entregables WHERE fundacion_id=? AND clave_unica=?", (fundacion_id, calendar_key)).fetchone()
                calendar_values = (
                    title, description, start, due, component, activity_type, responsible_id, responsible_name,
                    unit_name, state.lower(), priority.title(), 1 if (rule and rule.get("evidencias")) else 0,
                    user.get("username") or "sistema", now, now, fundacion_id, user.get("id"), calendar_key, spec["module"],
                )
                if existing:
                    entregable_id = int(existing[0])
                    conn.execute(
                        """UPDATE calendario_entregables SET titulo=?,descripcion=?,fecha_inicio=?,fecha_limite=?,modulo=?,tipo_formato=?,
                        responsable_id=?,responsable_nombre=?,unidad=?,estado=?,prioridad=?,requiere_evidencia=?,creado_por=COALESCE(creado_por,?),
                        actualizado_en=?,fundacion_id=?,usuario_creador_id=COALESCE(usuario_creador_id,?),origen=? WHERE id=?""",
                        calendar_values[:13] + (now, fundacion_id, user.get("id"), spec["module"], entregable_id),
                    )
                    updated += 1
                else:
                    cur = conn.execute(
                        """INSERT INTO calendario_entregables
                        (titulo,descripcion,fecha_inicio,fecha_limite,modulo,tipo_formato,responsable_id,responsable_nombre,unidad,estado,prioridad,
                         requiere_evidencia,creado_por,fecha_creacion,actualizado_en,fundacion_id,usuario_creador_id,clave_unica,origen)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        calendar_values,
                    )
                    entregable_id = int(cur.lastrowid)
                    created += 1
                documents = (rule or {}).get("documentos") or ["AGENDA"]
                evidence = (rule or {}).get("evidencias") or []
                role = (rule or {}).get("rol_responsable")
                meta_values = (
                    fundacion_id, entregable_id, expediente_id, unit_id, unit_name, unit_key(unit_name), component, activity_type,
                    spec["module"], table, source_id, str(unique), role, (rule or {}).get("id"), state,
                    100 if state in COMPLETED_STATES else 0, 0, None,
                    1 if "AGENDA" in documents else 0, 1 if "ACTA" in documents else 0,
                    1 if "LISTADO_ASISTENCIA" in documents else 0, 1 if "INFORME" in documents else 0,
                    1 if evidence else 0, json_dump({"source": row, "documents": documents, "evidence": evidence}),
                    user.get("id"), user.get("id"), now, now,
                )
                conn.execute(
                    """INSERT INTO cpo_actividad_metadata
                    (fundacion_id,entregable_id,expediente_uca_id,unidad_id,unidad_nombre,unidad_clave,componente,tipo_actividad,
                     fuente_modulo,fuente_tabla,fuente_id,fuente_clave,rol_responsable,regla_id,estado_flujo,porcentaje,bloqueada,motivo_bloqueo,
                     requiere_agenda,requiere_acta,requiere_listado,requiere_informe,requiere_evidencias,metadata_json,
                     creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(fundacion_id,entregable_id) DO UPDATE SET
                     expediente_uca_id=excluded.expediente_uca_id,unidad_id=excluded.unidad_id,unidad_nombre=excluded.unidad_nombre,
                     unidad_clave=excluded.unidad_clave,componente=excluded.componente,tipo_actividad=excluded.tipo_actividad,
                     fuente_modulo=excluded.fuente_modulo,fuente_tabla=excluded.fuente_tabla,fuente_id=excluded.fuente_id,
                     fuente_clave=excluded.fuente_clave,rol_responsable=excluded.rol_responsable,regla_id=excluded.regla_id,
                     estado_flujo=excluded.estado_flujo,porcentaje=excluded.porcentaje,requiere_agenda=excluded.requiere_agenda,
                     requiere_acta=excluded.requiere_acta,requiere_listado=excluded.requiere_listado,requiere_informe=excluded.requiere_informe,
                     requiere_evidencias=excluded.requiere_evidencias,metadata_json=excluded.metadata_json,activa=1,
                     actualizada_por=excluded.actualizada_por,fecha_actualizacion=excluded.fecha_actualizacion""",
                    meta_values,
                )
            self._recalculate_dependencies(conn, fundacion_id)
            self._generate_notifications(conn, fundacion_id)
            self.audit(fundacion_id, user, "SINCRONIZAR", "centro_planeacion", None, {"fuentes": sources, "creadas": created, "actualizadas": updated}, conn)
            conn.commit()
        return {"fuentes_leidas": sources, "creadas": created, "actualizadas": updated, "mensaje": "Sincronización idempotente completada."}

    def _recalculate_dependencies(self, conn, fundacion_id: int) -> None:
        rows = conn.execute(
            """SELECT m.id,m.estado_flujo,d.depende_de_actividad_id,p.estado_flujo AS padre_estado
               FROM cpo_actividad_metadata m
               LEFT JOIN cpo_dependencias d ON d.fundacion_id=m.fundacion_id AND d.actividad_id=m.id AND d.obligatoria=1
               LEFT JOIN cpo_actividad_metadata p ON p.id=d.depende_de_actividad_id
               WHERE m.fundacion_id=? AND m.activa=1""", (fundacion_id,)
        ).fetchall()
        grouped: dict[int, list[str]] = defaultdict(list)
        for row in rows:
            if row["depende_de_actividad_id"]:
                grouped[int(row["id"])].append(normalize(row["padre_estado"]))
        for activity_id, states in grouped.items():
            blocked = any(state not in COMPLETED_STATES for state in states)
            conn.execute("UPDATE cpo_actividad_metadata SET bloqueada=?,motivo_bloqueo=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
                         (1 if blocked else 0, "Dependencia obligatoria pendiente" if blocked else None, now_iso(), fundacion_id, activity_id))

    def _generate_notifications(self, conn, fundacion_id: int) -> None:
        rows = conn.execute(
            """SELECT m.*,e.titulo,e.fecha_limite,e.responsable_id,e.responsable_nombre,e.estado
               FROM cpo_actividad_metadata m JOIN calendario_entregables e ON e.id=m.entregable_id
               WHERE m.fundacion_id=? AND m.activa=1""", (fundacion_id,)
        ).fetchall()
        today = date.today()
        for raw in rows:
            row = dict(raw)
            if normalize(row.get("estado_flujo") or row.get("estado")) in COMPLETED_STATES:
                continue
            due_text = str(row.get("fecha_limite") or "")[:10]
            try:
                due = date.fromisoformat(due_text)
            except Exception:
                continue
            rule = None
            if row.get("regla_id"):
                r = conn.execute("SELECT * FROM cpo_reglas_operativas WHERE fundacion_id=? AND id=?", (fundacion_id, row["regla_id"])).fetchone()
                rule = self._rule(dict(r)) if r else None
            days_list = (rule or {}).get("dias_recordatorio") or [7,2,0]
            remaining = (due - today).days
            for offset in days_list:
                try:
                    offset_i = int(offset)
                except Exception:
                    continue
                if remaining != offset_i and not (remaining < 0 and offset_i == 0):
                    continue
                level = "CRITICO" if remaining <= 0 else ("ALTO" if remaining <= 2 else "INFO")
                scheduled = due_text if remaining >= 0 else today.isoformat()
                conn.execute(
                    """INSERT INTO cpo_notificaciones
                    (fundacion_id,actividad_id,destinatario_id,destinatario_rol,nivel,tipo,titulo,mensaje,fecha_programada,estado,leida,fecha_creacion)
                    VALUES(?,?,?,?,?,'VENCIMIENTO',?,?,?,?,0,?)
                    ON CONFLICT(fundacion_id,actividad_id,destinatario_id,destinatario_rol,tipo,fecha_programada) DO NOTHING""",
                    (fundacion_id, row["id"], row.get("responsable_id"), row.get("rol_responsable"), level,
                     f"Actividad {('vencida' if remaining < 0 else 'próxima a vencer')}: {row.get('titulo')}",
                     f"Fecha límite {due_text}. Días restantes: {remaining}.", scheduled, "PENDIENTE", now_iso()),
                )

    def add_dependency(self, fundacion_id: int, activity_id: int, parent_id: int, user: dict[str, Any]) -> dict[str, Any]:
        if activity_id == parent_id:
            raise ValueError("Una actividad no puede depender de sí misma.")
        now = now_iso()
        with self.connect() as conn:
            for item_id in (activity_id, parent_id):
                if not conn.execute("SELECT 1 FROM cpo_actividad_metadata WHERE fundacion_id=? AND id=?", (fundacion_id, item_id)).fetchone():
                    raise LookupError("Actividad no encontrada.")
            cur = conn.execute(
                "INSERT INTO cpo_dependencias(fundacion_id,actividad_id,depende_de_actividad_id,obligatoria,tipo,observaciones,creada_por,fecha_creacion) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(fundacion_id,actividad_id,depende_de_actividad_id) DO NOTHING",
                (fundacion_id, activity_id, parent_id, 1, "FIN_A_INICIO", None, user.get("id"), now),
            )
            self._recalculate_dependencies(conn, fundacion_id)
            self.audit(fundacion_id, user, "AGREGAR_DEPENDENCIA", "cpo_dependencias", int(cur.lastrowid or 0), {"actividad": activity_id, "depende_de": parent_id}, conn)
            conn.commit()
        return self.activity(fundacion_id, activity_id)

    def _activity_query(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> tuple[str, list[Any]]:
        filters = filters or {}
        where = ["m.fundacion_id=?", "m.activa=1"]
        params: list[Any] = [fundacion_id]
        for key, column in (("unidad", "m.unidad_nombre"), ("componente", "m.componente"), ("estado", "m.estado_flujo"), ("rol", "m.rol_responsable")):
            value = filters.get(key)
            if value:
                where.append(f"UPPER(COALESCE({column},'')) LIKE UPPER(?)")
                params.append(f"%{value}%")
        if filters.get("periodo"):
            where.append("substr(e.fecha_limite,1,7)=?")
            params.append(str(filters["periodo"])[:7])
        if filters.get("responsable_id"):
            where.append("e.responsable_id=?")
            params.append(int(filters["responsable_id"]))
        return " AND ".join(where), params

    def list_activities(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        where, params = self._activity_query(fundacion_id, filters)
        with self.connect() as conn:
            rows = conn.execute(
                f"""SELECT m.*,e.titulo,e.descripcion,e.fecha_inicio,e.fecha_limite,e.estado AS calendario_estado,e.prioridad,
                e.responsable_id,e.responsable_nombre,e.archivo_evidencia,e.fecha_entrega,e.observaciones AS calendario_observaciones,
                r.codigo AS regla_codigo,r.nombre AS regla_nombre
                FROM cpo_actividad_metadata m
                JOIN calendario_entregables e ON e.id=m.entregable_id
                LEFT JOIN cpo_reglas_operativas r ON r.id=m.regla_id
                WHERE {where} ORDER BY e.fecha_limite ASC,m.componente,m.unidad_nombre""", params
            ).fetchall()
            result = [self._decorate_activity(conn, dict(row)) for row in rows]
        return result

    def _decorate_activity(self, conn, data: dict[str, Any]) -> dict[str, Any]:
        data["metadata"] = parse_json(data.get("metadata_json"), {})
        data["semaforo"] = semaforo(data.get("fecha_limite"), data.get("estado_flujo") or data.get("calendario_estado"), data.get("bloqueada"))
        deps = conn.execute(
            """SELECT d.*,p.id AS padre_meta_id,e.titulo AS padre_titulo,p.estado_flujo AS padre_estado
               FROM cpo_dependencias d JOIN cpo_actividad_metadata p ON p.id=d.depende_de_actividad_id
               JOIN calendario_entregables e ON e.id=p.entregable_id
               WHERE d.fundacion_id=? AND d.actividad_id=?""", (data["fundacion_id"], data["id"])
        ).fetchall()
        data["dependencias"] = [dict(row) for row in deps]
        data["documentos"] = [dict(row) for row in conn.execute("SELECT * FROM cpo_documentos_preparados WHERE fundacion_id=? AND actividad_id=? ORDER BY fecha_generacion DESC", (data["fundacion_id"], data["id"])).fetchall()]
        return data

    def activity(self, fundacion_id: int, activity_id: int) -> dict[str, Any]:
        rows = self.list_activities(fundacion_id, {})
        for row in rows:
            if int(row["id"]) == int(activity_id):
                return row
        raise LookupError("Actividad no encontrada.")

    def update_activity(self, fundacion_id: int, activity_id: int, data: dict[str, Any], user: dict[str, Any], allow_approve: bool = False) -> dict[str, Any]:
        current = self.activity(fundacion_id, activity_id)
        state = normalize(data.get("estado_flujo") or current.get("estado_flujo"))
        if state in {"APROBADA", "APROBADO", "CERRADA", "CERRADO"} and not allow_approve:
            raise PermissionError("La aprobación o cierre requiere coordinación.")
        if current.get("bloqueada") and state in COMPLETED_STATES:
            raise ValueError("La actividad tiene dependencias obligatorias pendientes.")
        if state in COMPLETED_STATES and current.get("requiere_evidencias") and not (current.get("archivo_evidencia") or current.get("documentos")):
            raise ValueError("La actividad requiere evidencias o documentos antes del cierre.")
        now = now_iso()
        with self.connect() as conn:
            conn.execute(
                """UPDATE cpo_actividad_metadata SET estado_flujo=?,porcentaje=?,revisor_id=COALESCE(?,revisor_id),
                revisor_nombre=COALESCE(?,revisor_nombre),aprobador_id=CASE WHEN ? IN ('APROBADA','APROBADO','CERRADA','CERRADO') THEN ? ELSE aprobador_id END,
                aprobador_nombre=CASE WHEN ? IN ('APROBADA','APROBADO','CERRADA','CERRADO') THEN ? ELSE aprobador_nombre END,
                actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?""",
                (state, float(data.get("porcentaje", 100 if state in COMPLETED_STATES else current.get("porcentaje") or 0)),
                 data.get("revisor_id"), data.get("revisor_nombre"), state, user.get("id"), state,
                 user.get("nombre_completo") or user.get("username"), user.get("id"), now, fundacion_id, activity_id),
            )
            conn.execute(
                "UPDATE calendario_entregables SET estado=?,fecha_entrega=?,observaciones=COALESCE(?,observaciones),actualizado_en=? WHERE fundacion_id=? AND id=?",
                (state.lower(), date.today().isoformat() if state in COMPLETED_STATES else current.get("fecha_entrega"), data.get("observaciones"), now, fundacion_id, current["entregable_id"]),
            )
            self._recalculate_dependencies(conn, fundacion_id)
            self.audit(fundacion_id, user, "ACTUALIZAR_ACTIVIDAD", "cpo_actividad_metadata", activity_id, {"estado": state}, conn)
            conn.commit()
        return self.activity(fundacion_id, activity_id)

    def dashboard(self, fundacion_id: int, user: dict[str, Any], filters: dict[str, Any] | None = None) -> dict[str, Any]:
        filters = dict(filters or {})
        is_coordinator = user.get("rol") in COORDINATION_ROLES
        coordination = bool(is_coordinator and filters.get("vista") != "rol")
        if not coordination:
            filters.setdefault("responsable_id", user.get("id"))
        rows = self.list_activities(fundacion_id, filters)
        by_component: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "verde": 0, "amarillo": 0, "rojo": 0})
        by_unit: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "pendientes": 0, "vencidas": 0, "bloqueadas": 0})
        for item in rows:
            component = item.get("componente") or "SIN_COMPONENTE"
            color = str(item["semaforo"]["color"]).lower()
            by_component[component]["total"] += 1
            by_component[component][color] += 1
            unit = item.get("unidad_nombre") or "Sin UCA"
            by_unit[unit]["total"] += 1
            if normalize(item.get("estado_flujo")) not in COMPLETED_STATES:
                by_unit[unit]["pendientes"] += 1
            if item["semaforo"]["nivel"] == "VENCIDA":
                by_unit[unit]["vencidas"] += 1
            if item.get("bloqueada"):
                by_unit[unit]["bloqueadas"] += 1
        notifications = self.list_notifications(fundacion_id, user)
        return {
            "vista": "COORDINACION" if coordination else "ROL",
            "rol": user.get("rol"),
            "resumen": {
                "total": len(rows),
                "pendientes": sum(1 for x in rows if normalize(x.get("estado_flujo")) not in COMPLETED_STATES),
                "vencidas": sum(1 for x in rows if x["semaforo"]["nivel"] == "VENCIDA"),
                "criticas": sum(1 for x in rows if x["semaforo"]["color"] == "ROJO"),
                "bloqueadas": sum(1 for x in rows if x.get("bloqueada")),
                "documentos_borrador": sum(1 for x in rows for d in x.get("documentos", []) if normalize(d.get("estado")) == "BORRADOR"),
                "notificaciones": len(notifications),
            },
            "por_componente": [{"componente": key, **value} for key, value in sorted(by_component.items())],
            "por_uca": [{"unidad": key, **value} for key, value in sorted(by_unit.items())],
            "actividades": rows[:300],
            "notificaciones": notifications[:100],
        }

    def list_notifications(self, fundacion_id: int, user: dict[str, Any]) -> list[dict[str, Any]]:
        where = ["fundacion_id=?", "estado='PENDIENTE'"]
        params: list[Any] = [fundacion_id]
        if user.get("rol") not in COORDINATION_ROLES:
            where.append("(destinatario_id=? OR destinatario_rol=?)")
            params.extend([user.get("id"), user.get("rol")])
        with self.connect() as conn:
            rows = conn.execute("SELECT * FROM cpo_notificaciones WHERE " + " AND ".join(where) + " ORDER BY nivel DESC,fecha_programada", params).fetchall()
        return [dict(row) for row in rows]

    def mark_notification_read(self, fundacion_id: int, notification_id: int, user: dict[str, Any]) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM cpo_notificaciones WHERE fundacion_id=? AND id=?", (fundacion_id, notification_id)).fetchone()
            if not row:
                raise LookupError("Notificación no encontrada.")
            conn.execute("UPDATE cpo_notificaciones SET leida=1,estado='LEIDA',fecha_lectura=? WHERE fundacion_id=? AND id=?", (now_iso(), fundacion_id, notification_id))
            self.audit(fundacion_id, user, "LEER_NOTIFICACION", "cpo_notificaciones", notification_id, {}, conn)
            conn.commit()
        return dict(row) | {"leida": 1, "estado": "LEIDA"}

    def _tenant_folder(self, fundacion_id: int, *parts: str) -> Path:
        root = tenant_storage_root(self.data_dir, fundacion_id).resolve()
        path = root.joinpath("centro_planeacion", *parts).resolve()
        path.mkdir(parents=True, exist_ok=True)
        return path

    def prepare_documents(self, fundacion_id: int, activity_id: int, types: Iterable[str], user: dict[str, Any]) -> list[dict[str, Any]]:
        activity = self.activity(fundacion_id, activity_id)
        types_norm = [normalize(value) for value in types if normalize(value)]
        if not types_norm:
            types_norm = ["AGENDA"]
        folder = self._tenant_folder(fundacion_id, "documentos")
        token = datetime.now().strftime("%Y%m%d_%H%M%S")
        generated: list[dict[str, Any]] = []
        for doc_type in types_norm:
            if doc_type == "LISTADO_ASISTENCIA":
                path = folder / f"listado_asistencia_actividad_{activity_id}_{token}.xlsx"
                self._write_attendance(path, activity)
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                safe = doc_type.lower().replace("_", "-")
                path = folder / f"{safe}_actividad_{activity_id}_{token}.pdf"
                self._write_pdf(path, activity, doc_type)
                mime = "application/pdf"
            generated.append(self._save_document(fundacion_id, activity_id, doc_type, path, mime, user))
        return generated

    def _save_document(self, fundacion_id: int, activity_id: int, doc_type: str, path: Path, mime: str, user: dict[str, Any]) -> dict[str, Any]:
        now = now_iso()
        digest = file_sha256(path)
        with self.connect() as conn:
            cur = conn.execute(
                """INSERT INTO cpo_documentos_preparados
                (fundacion_id,actividad_id,tipo_documento,nombre_archivo,ruta_archivo,mime_type,tamano_bytes,sha256,estado,
                 plantilla_codigo,plantilla_version,generado_por,fecha_generacion)
                VALUES(?,?,?,?,?,?,?,?, 'BORRADOR',CASE WHEN ?='LISTADO_ASISTENCIA' THEN 'F27.MT1.PP' ELSE 'CPO-GENERICA' END,CASE WHEN ?='LISTADO_ASISTENCIA' THEN '3' ELSE '1' END,?,?)""",
                (fundacion_id, activity_id, doc_type, path.name, str(path), mime, path.stat().st_size, digest, doc_type, doc_type, user.get("id"), now),
            )
            doc_id = int(cur.lastrowid)
            self.audit(fundacion_id, user, "GENERAR_DOCUMENTO", "cpo_documentos_preparados", doc_id, {"tipo": doc_type, "sha256": digest}, conn)
            conn.commit()
        return self.document(fundacion_id, doc_id)

    def _write_attendance(self, path: Path, activity: dict[str, Any]) -> None:
        """Genera el formato oficial RAM V3, no una hoja genérica."""
        from generador_formatos import GeneradorFormatos
        from modules.plantillas_oficiales import generar_desde_plantilla_oficial

        unit = str(activity.get("unidad_nombre") or activity.get("unidad") or "").strip()
        if not unit:
            raise ValueError("La actividad no tiene UDS/UCA asignada para generar el listado oficial.")
        raw_date = str(activity.get("fecha_limite") or activity.get("fecha_inicio") or date.today().isoformat())[:10]
        try:
            period_date = date.fromisoformat(raw_date)
        except ValueError:
            period_date = date.today()
        year, month = period_date.year, period_date.month
        templates_dir = self.data_dir / "templates_originales"
        packaged_templates = Path(__file__).resolve().parents[2] / "seed_data" / "templates_originales"
        if not (templates_dir / "oficiales" / "plantilla_ram_oficial_v3.xlsx").is_file():
            templates_dir = packaged_templates
        generator = GeneradorFormatos(
            self.database_path,
            str(templates_dir),
            str(path.parent),
        )
        with self.connect() as conn:
            rows, _ = self._ram_participant_rows(conn, int(activity.get("fundacion_id") or 1), unit)
        users = [generator._usuario_oficial(dict(row)) for row in rows]
        metadata = generator._metadata_oficial(month, year, unit)
        metadata.update({"mes_numero": month, "mes_nombre": metadata.get("mes_nombre") or metadata.get("mes")})
        generar_desde_plantilla_oficial(
            "ram",
            {"metadata": metadata, "usuarios": users},
            str(path),
            str(templates_dir),
        )

    def _ram_participant_rows(self, conn, fundacion_id: int, unit: str):
        """Base Maestra primero; legado solo si aún no existe versión publicada."""
        active_version = False
        if self._table_exists(conn, "master_versiones"):
            active_version = bool(conn.execute(
                "SELECT 1 FROM master_versiones WHERE fundacion_id=? AND activa=1 LIMIT 1",
                (fundacion_id,),
            ).fetchone())
        if self._table_exists(conn, "master_ninos"):
            rows = conn.execute(
                """
                SELECT n.* FROM master_ninos n
                 WHERE COALESCE(n.fundacion_id,1)=?
                   AND LOWER(TRIM(COALESCE(n.unidad_servicio,'')))=LOWER(TRIM(?))
                   AND COALESCE(n.activo,1)=1
                   AND UPPER(COALESCE(n.estado,'ACTIVO')) NOT IN ('INACTIVO','RETIRADO','FALLECIDO')
                 ORDER BY COALESCE(n.nombre_completo,n.documento)
                """,
                (fundacion_id, unit),
            ).fetchall()
            if rows or active_version:
                return rows, "master_ninos"
        if self._table_exists(conn, "beneficiarios"):
            rows = conn.execute(
                """
                SELECT b.* FROM beneficiarios b
                 WHERE COALESCE(b.fundacion_id,1)=?
                   AND LOWER(TRIM(COALESCE(b.unidad,'')))=LOWER(TRIM(?))
                   AND UPPER(COALESCE(b.estado,'ACTIVO')) NOT IN ('INACTIVO','RETIRADO','FALLECIDO')
                 ORDER BY COALESCE(NULLIF(b.primer_nombre,''), b.nombres, b.documento),
                          COALESCE(NULLIF(b.primer_apellido,''), b.apellidos, '')
                """,
                (fundacion_id, unit),
            ).fetchall()
            return rows, "beneficiarios_compatibilidad_sin_version_maestra"
        return [], "sin_fuente"

    def _write_pdf(self, path: Path, activity: dict[str, Any], doc_type: str) -> None:
        styles = getSampleStyleSheet(); story = []
        title = f"{doc_type.replace('_',' ').title()} — BORRADOR PARA REVISIÓN HUMANA"
        story.append(Paragraph(title, styles["Title"])); story.append(Spacer(1, 0.3*cm))
        data = [
            ["Actividad", activity.get("titulo") or ""], ["Componente", activity.get("componente") or ""],
            ["UCA", activity.get("unidad_nombre") or ""], ["Fecha", activity.get("fecha_limite") or ""],
            ["Responsable", activity.get("responsable_nombre") or activity.get("rol_responsable") or ""],
            ["Estado", activity.get("estado_flujo") or ""], ["Fuente", f"{activity.get('fuente_modulo')} / {activity.get('fuente_tabla')}"],
        ]
        table = Table(data, colWidths=[4*cm, 12*cm]); table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#0f766e")), ("TEXTCOLOR",(0,0),(0,-1),colors.white),
            ("GRID",(0,0),(-1,-1),0.4,colors.grey), ("VALIGN",(0,0),(-1,-1),"TOP"), ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ])); story.append(table); story.append(Spacer(1,0.4*cm))
        if doc_type == "ACTA":
            sections = ["Objetivo", "Desarrollo de la actividad", "Resultados observados", "Compromisos", "Responsables", "Observaciones y firmas"]
        elif doc_type == "INFORME":
            sections = ["Resumen ejecutivo", "Metodología", "Resultados", "Indicadores", "Evidencias relacionadas", "Conclusiones profesionales", "Recomendaciones"]
        else:
            sections = ["Objetivo", "Agenda", "Recursos", "Documentos requeridos", "Evidencias esperadas", "Responsables"]
        for section in sections:
            story.append(Paragraph(f"<b>{section}</b>", styles["Heading3"])); story.append(Paragraph("Pendiente de diligenciamiento y validación por el profesional responsable.", styles["BodyText"])); story.append(Spacer(1,0.25*cm))
        SimpleDocTemplate(str(path), pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm).build(story)

    def document(self, fundacion_id: int, document_id: int) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute("SELECT d.*,m.unidad_nombre FROM cpo_documentos_preparados d JOIN cpo_actividad_metadata m ON m.id=d.actividad_id WHERE d.fundacion_id=? AND d.id=?", (fundacion_id, document_id)).fetchone()
        if not row:
            raise LookupError("Documento no encontrado.")
        return dict(row)

    def document_path(self, fundacion_id: int, document_id: int) -> tuple[Path, str, str] | None:
        row = self.document(fundacion_id, document_id)
        path = Path(row["ruta_archivo"]).resolve()
        root = tenant_storage_root(self.data_dir, fundacion_id).resolve()
        try:
            path.relative_to(root)
        except ValueError:
            return None
        if not path.is_file() or file_sha256(path) != row["sha256"]:
            return None
        return path, row["nombre_archivo"], row.get("mime_type") or mimetypes.guess_type(path.name)[0] or "application/octet-stream"

    def review_document(self, fundacion_id: int, document_id: int, action: str, user: dict[str, Any]) -> dict[str, Any]:
        action_norm = normalize(action)
        if action_norm not in {"REVISAR", "APROBAR", "DEVOLVER"}:
            raise ValueError("Acción documental inválida.")
        state = {"REVISAR": "REVISADO", "APROBAR": "APROBADO", "DEVOLVER": "DEVUELTO"}[action_norm]
        now = now_iso()
        with self.connect() as conn:
            if not conn.execute("SELECT 1 FROM cpo_documentos_preparados WHERE fundacion_id=? AND id=?", (fundacion_id, document_id)).fetchone():
                raise LookupError("Documento no encontrado.")
            if action_norm == "APROBAR":
                conn.execute("UPDATE cpo_documentos_preparados SET estado=?,aprobado_por=?,fecha_aprobacion=? WHERE fundacion_id=? AND id=?", (state,user.get("id"),now,fundacion_id,document_id))
            else:
                conn.execute("UPDATE cpo_documentos_preparados SET estado=?,revisado_por=?,fecha_revision=? WHERE fundacion_id=? AND id=?", (state,user.get("id"),now,fundacion_id,document_id))
            self.audit(fundacion_id,user,action_norm,"cpo_documentos_preparados",document_id,{},conn); conn.commit()
        return self.document(fundacion_id, document_id)

    def export_monthly_package(self, fundacion_id: int, user: dict[str, Any], period: str, filters: dict[str, Any] | None = None) -> dict[str, Any]:
        filters = dict(filters or {}); filters["periodo"] = period
        data = self.dashboard(fundacion_id, user, filters)
        folder = self._tenant_folder(fundacion_id, "paquetes")
        token = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx = folder / f"planeacion_{period}_{token}.xlsx"
        zip_path = folder / f"paquete_planeacion_{period}_{token}.zip"
        wb=Workbook(); ws=wb.active; ws.title="Actividades"; ws.append(["ID","Componente","Tipo","Título","UCA","Responsable","Inicio","Límite","Estado","Semáforo","Fuente"])
        for item in data.get("actividades",[]):
            ws.append([item.get("id"),item.get("componente"),item.get("tipo_actividad"),item.get("titulo"),item.get("unidad_nombre"),item.get("responsable_nombre") or item.get("rol_responsable"),item.get("fecha_inicio"),item.get("fecha_limite"),item.get("estado_flujo"),item.get("semaforo",{}).get("color"),item.get("fuente_tabla")])
        for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="334155")
        wb.save(xlsx)
        with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as archive:
            archive.write(xlsx,xlsx.name)
            archive.writestr("00_RESUMEN.json",json.dumps(data.get("resumen",{}),ensure_ascii=False,indent=2,default=str))
            archive.writestr("01_POR_COMPONENTE.json",json.dumps(data.get("por_componente",[]),ensure_ascii=False,indent=2,default=str))
            archive.writestr("02_POR_UCA.json",json.dumps(data.get("por_uca",[]),ensure_ascii=False,indent=2,default=str))
            archive.writestr("LEEME.txt","Paquete operativo generado como borrador. Requiere revisión y aprobación humana.\n")
        self.audit(fundacion_id,user,"GENERAR_PAQUETE","centro_planeacion",None,{"periodo":period,"sha256":file_sha256(zip_path)})
        return {"ruta":str(zip_path),"nombre_archivo":zip_path.name,"sha256":file_sha256(zip_path),"tamano_bytes":zip_path.stat().st_size}

    def package_path(self, fundacion_id: int, path_value: str) -> Path | None:
        path=Path(path_value).resolve(); root=tenant_storage_root(self.data_dir,fundacion_id).resolve()
        try: path.relative_to(root)
        except ValueError: return None
        return path if path.is_file() else None
