from __future__ import annotations

import json
import os
import re
import shutil
from modules.dbapi_compat import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .schema import PM_SCHEMA_SQL, CATEGORIAS_PAQUETE
from services.relacion_mes_service import cantidades, consolidar_por_unidad, docente_mas_frecuente


MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE',
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def safe_filename(text: str) -> str:
    text = str(text or '').strip()
    text = re.sub(r'[^\w\-.]+', '_', text, flags=re.UNICODE)
    text = re.sub(r'_+', '_', text).strip('_')
    return text or 'archivo'


def normalizar_texto(text: str) -> str:
    import unicodedata
    text = str(text or '').strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return ' '.join(text.split())


def periodo_key(anio: int, mes: int) -> str:
    return f'{anio}-{mes:02d}'


def ensure_dir(path: str | Path) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


class PaqueteMensualService:
    def __init__(self, database_path: str, output_folder: str, base_dir: str | None = None):
        self.database_path = database_path
        self._output_folder = output_folder
        self.base_dir = Path(base_dir or Path(database_path).parent)
        self.templates_folder = self.base_dir / 'templates_originales'

    @property
    def output_folder(self) -> Path:
        return ensure_dir(Path(os.fspath(self._output_folder)))

    @property
    def paquetes_folder(self) -> Path:
        return ensure_dir(self.output_folder / 'paquete_mensual')

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.database_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_schema(self) -> None:
        conn = self.connect()
        conn.executescript(PM_SCHEMA_SQL)
        self._ensure_runtime_columns(conn)
        conn.commit()
        conn.close()

    def _ensure_runtime_columns(self, conn: sqlite3.Connection) -> None:
        """Migra instalaciones existentes sin borrar historial del paquete mensual."""
        cur = conn.cursor()
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(pm_paquetes)").fetchall()}
        extras = {
            'ruta_carpeta': 'TEXT',
            'componentes_json': 'TEXT',
            'errores_json': 'TEXT',
            'fecha_actualizacion': 'TEXT',
        }
        for col, ddl in extras.items():
            if col not in cols:
                cur.execute(f"ALTER TABLE pm_paquetes ADD COLUMN {col} {ddl}")

    def table_exists(self, cur: sqlite3.Cursor, table: str) -> bool:
        row = cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
        return bool(row)

    def fetch_all(self, table: str, sql: str, params: tuple = ()) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, table):
            conn.close()
            return []
        rows = [dict(r) for r in cur.execute(sql, params).fetchall()]
        conn.close()
        return rows

    def fetch_one(self, table: str, sql: str, params: tuple = ()) -> dict[str, Any] | None:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, table):
            conn.close()
            return None
        row = cur.execute(sql, params).fetchone()
        conn.close()
        return dict(row) if row else None

    def get_fundacion_nombre(self, fundacion_id: int | None) -> str:
        row = self.fetch_one('fundaciones', "SELECT nombre FROM fundaciones WHERE id=?", (fundacion_id or 1,))
        return row.get('nombre') if row else 'Fundación Principal'

    def get_beneficiarios(self, fundacion_id: int | None = None) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, 'master_ninos'):
            conn.close()
            return []
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(master_ninos)").fetchall()}
        where = "1=1"
        params: list[Any] = []
        if fundacion_id and 'fundacion_id' in cols:
            where += " AND COALESCE(fundacion_id, ?) = ?"
            params.extend([fundacion_id, fundacion_id])
        rows = [dict(r) for r in cur.execute(f"SELECT *, unidad_servicio AS unidad, documento AS nui FROM master_ninos WHERE activo=1 AND {where}", tuple(params)).fetchall()]
        conn.close()
        return rows

    def get_talento(self, fundacion_id: int | None = None) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, 'master_talento_humano'):
            conn.close()
            return []
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(master_talento_humano)").fetchall()}
        where = "1=1"
        params: list[Any] = []
        if fundacion_id and 'fundacion_id' in cols:
            where += " AND COALESCE(fundacion_id, ?) = ?"
            params.extend([fundacion_id, fundacion_id])
        rows = [dict(r) for r in cur.execute(f"SELECT *, nombre_completo AS nombre, unidad_servicio AS unidad FROM master_talento_humano WHERE activo=1 AND {where} ORDER BY unidad_servicio, cargo, nombre_completo", tuple(params)).fetchall()]
        conn.close()
        return rows

    def docente_unidad(self, unidad: str, talento: list[dict[str, Any]]) -> str:
        unidad_norm = normalizar_texto(unidad)
        for t in talento:
            cargo = normalizar_texto(t.get('rol_normalizado') or t.get('cargo') or t.get('tipo_equipo') or '')
            unidad_t = normalizar_texto(t.get('unidad') or '')
            if unidad_norm and unidad_norm == unidad_t and ('docente' in cargo or 'agente' in cargo):
                return t.get('nombre') or 'Sin docente'
        for t in talento:
            unidad_t = normalizar_texto(t.get('unidad') or '')
            if unidad_norm and unidad_norm == unidad_t:
                return t.get('nombre') or 'Sin docente'
        return 'Sin docente asignado'

    def coordinador_unidad(self, unidad: str, talento: list[dict[str, Any]]) -> str:
        unidad_norm = normalizar_texto(unidad)
        for t in talento:
            cargo = normalizar_texto(t.get('cargo') or t.get('tipo_equipo') or '')
            if 'coord' in cargo and unidad_norm and unidad_norm == normalizar_texto(t.get('unidad') or ''):
                return t.get('nombre') or ''
        for t in talento:
            coord = t.get('coordinador')
            if coord and unidad_norm and unidad_norm == normalizar_texto(t.get('unidad') or ''):
                return coord
        return ''

    def edad_texto(self, edad_meses: Any) -> str:
        try:
            total = int(float(edad_meses or 0))
        except Exception:
            total = 0
        anios = total // 12
        meses = total % 12
        if anios and meses:
            return f'{anios} años y {meses} meses'
        if anios:
            return f'{anios} años'
        return f'{meses} meses'

    def grupos_por_unidad(self, beneficiarios: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        return consolidar_por_unidad(beneficiarios)

    def write_excel(self, path: Path, title: str, sheets: dict[str, dict[str, Any]]) -> None:
        wb = Workbook()
        ws = wb.active
        first = True
        header_fill = PatternFill('solid', fgColor='D9EAF7')
        title_fill = PatternFill('solid', fgColor='1F2937')
        title_font = Font(bold=True, color='FFFFFF', size=13)
        thin = Side(style='thin', color='CBD5E1')
        for sheet_name, cfg in sheets.items():
            if first:
                ws.title = sheet_name[:31] or 'Reporte'
                first = False
            else:
                ws = wb.create_sheet(sheet_name[:31] or 'Reporte')
            headers = cfg.get('headers') or []
            rows = cfg.get('rows') or []
            meta = cfg.get('meta') or []
            max_cols = max(1, len(headers), *(len(r) for r in rows)) if rows else max(1, len(headers))
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
            ws.cell(1, 1).value = title
            ws.cell(1, 1).fill = title_fill
            ws.cell(1, 1).font = title_font
            ws.cell(1, 1).alignment = Alignment(horizontal='center')
            row_idx = 3
            for item in meta:
                ws.cell(row_idx, 1).value = item[0]
                ws.cell(row_idx, 2).value = item[1]
                ws.cell(row_idx, 1).font = Font(bold=True)
                row_idx += 1
            if meta:
                row_idx += 1
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for r, row in enumerate(rows, row_idx + 1):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(r, c)
                    cell.value = val
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in range(1, max_cols + 1):
                ws.column_dimensions[get_column_letter(col)].width = 22
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def write_pdf(self, path: Path, title: str, headers: list[str], rows: list[list[Any]], meta: list[tuple[str, str]] | None = None) -> None:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

            doc = SimpleDocTemplate(str(path), pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
            styles = getSampleStyleSheet()
            elems = [Paragraph(title, styles['Title']), Spacer(1, 8)]
            for k, v in meta or []:
                elems.append(Paragraph(f'<b>{k}:</b> {v}', styles['Normal']))
            if meta:
                elems.append(Spacer(1, 8))
            safe_rows = rows[:80]
            data = [headers] + [[str(x or '')[:90] for x in row] for row in safe_rows]
            if len(data) == 1:
                data.append(['Sin registros'] + [''] * (len(headers) - 1))
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
            ]))
            elems.append(table)
            if len(rows) > len(safe_rows):
                elems.append(Spacer(1, 8))
                elems.append(Paragraph(f'Se muestran {len(safe_rows)} de {len(rows)} registros. Consulte el Excel para el detalle completo.', styles['Italic']))
            doc.build(elems)
        except Exception:
            # Fallback mínimo: evita que falte reportlab o una fuente bloquee el ZIP completo.
            text_path = Path(path)
            text_path.write_text(
                title + '\n\n' + '\n'.join(f'{k}: {v}' for k, v in (meta or [])) + '\n\n' +
                '\t'.join(str(h) for h in headers) + '\n' +
                '\n'.join('\t'.join(str(x or '') for x in row) for row in rows[:500]),
                encoding='utf-8'
            )

    def add_file_record(self, conn: sqlite3.Connection, paquete_id: int, categoria: str, file_path: Path, tipo: str, estado: str = 'GENERADO', observaciones: str = '') -> None:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO pm_archivos
            (paquete_id, categoria, nombre_archivo, ruta_archivo, tipo, tamano_bytes, estado, observaciones, fecha_creacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            paquete_id, categoria, file_path.name, str(file_path), tipo, file_path.stat().st_size if file_path.exists() else 0,
            estado, observaciones, now_iso()
        ))

    def insert_package(self, periodo: str, mes: int, anio: int, user: dict[str, Any], observaciones: str = '') -> int:
        conn = self.connect()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO pm_paquetes
            (fundacion_id, usuario_id, periodo, mes, anio, estado, observaciones, fecha_creacion)
            VALUES (?, ?, ?, ?, ?, 'EN_PROCESO', ?, ?)
        """, (user.get('fundacion_id') or 1, user.get('id'), periodo, mes, anio, observaciones, now_iso()))
        new_id = cur.lastrowid
        cur.execute("""
            INSERT INTO pm_auditoria (paquete_id, accion, detalle, usuario_id, fundacion_id, fecha_accion)
            VALUES (?, 'CREAR_PAQUETE', ?, ?, ?, ?)
        """, (new_id, f'Inicio de generación del paquete mensual {periodo}', user.get('id'), user.get('fundacion_id') or 1, now_iso()))
        conn.commit()
        conn.close()
        return int(new_id)

    def update_package(self, paquete_id: int, **kwargs) -> None:
        if not kwargs:
            return
        conn = self.connect()
        self._ensure_runtime_columns(conn)
        kwargs.setdefault('fecha_actualizacion', now_iso())
        cols = ', '.join([f"{k}=?" for k in kwargs])
        vals = list(kwargs.values())
        vals.append(paquete_id)
        conn.execute(f"UPDATE pm_paquetes SET {cols} WHERE id=?", tuple(vals))
        conn.commit()
        conn.close()

    def _safe_write_text(self, path: Path, content: str) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding='utf-8')
        return path

    def _record_placeholder(self, conn: sqlite3.Connection, paquete_id: int, categoria: str, message: str, filename: str = 'PENDIENTE.txt') -> dict[str, Any]:
        folder = ensure_dir(Path(categoria)) if Path(categoria).is_absolute() else None
        # El caller normalmente pasa la carpeta final; si solo llega nombre de categoría se crea luego allí.
        raise RuntimeError('_record_placeholder requiere ruta absoluta en Alpha17')

    def unidades_activas(self, fundacion_id: int | None = None) -> list[str]:
        """Devuelve unidades/UDS con beneficiarios activos para generar el paquete mensual actualizado."""
        conn = self.connect()
        cur = conn.cursor()
        unidades: set[str] = set()
        if self.table_exists(cur, 'master_unidades'):
            cols = {r['name'] for r in cur.execute("PRAGMA table_info(master_unidades)").fetchall()}
            where = "activo=1 AND nombre IS NOT NULL AND TRIM(nombre) <> ''"
            params: list[Any] = []
            if 'estado' in cols:
                where += " AND (estado IS NULL OR LOWER(estado) IN ('activo', 'activa'))"
            if fundacion_id and 'fundacion_id' in cols:
                where += " AND COALESCE(fundacion_id, ?) = ?"
                params.extend([fundacion_id, fundacion_id])
            for row in cur.execute(f"SELECT DISTINCT nombre AS unidad FROM master_unidades WHERE {where} ORDER BY nombre", tuple(params)).fetchall():
                unidad = str(row['unidad'] or '').strip()
                if unidad:
                    unidades.add(unidad)
        conn.close()
        return sorted(unidades, key=normalizar_texto)

    def _add_generated_if_exists(self, conn: sqlite3.Connection, paquete_id: int, categoria: str, path: str | Path, tipo: str | None = None, estado: str = 'GENERADO') -> dict[str, Any] | None:
        if not path:
            return None
        p = Path(path)
        if not p.exists() or not p.is_file():
            return None
        self.add_file_record(conn, paquete_id, categoria, p, tipo or p.suffix.lstrip('.') or 'archivo', estado)
        return {'categoria': categoria, 'archivo': p.name, 'estado': estado, 'tamano': p.stat().st_size}

    def generar_formatos_operativos_actualizados(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """Genera dentro del paquete los formatos operativos actualizados.

        A diferencia de versiones anteriores, esta rutina no depende únicamente de
        archivos viejos ya generados. Genera Bienestarina, RPP y RAM
        por cada UDS activa usando los datos actuales de la base. Si una UDS falla,
        registra el error y continúa para que el ZIP completo siempre se pueda bajar.
        """
        generated: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []
        unidades = self.unidades_activas(fundacion_id)
        folder_bien = ensure_dir(package_dir / '01_Bienestarina')
        folder_rpp = ensure_dir(package_dir / '02_RPP')
        folder_ram = ensure_dir(package_dir / '03_RAM_RAN_RRAN')

        if not unidades:
            msg = 'No se encontraron unidades/UDS activas en beneficiarios o talento humano. Cargue o sincronice las bases antes de generar formatos operativos.'
            for folder, categoria in [(folder_bien, '01_Bienestarina'), (folder_rpp, '02_RPP'), (folder_ram, '03_RAM_RAN_RRAN')]:
                placeholder = self._safe_write_text(folder / 'SIN_UNIDADES_ACTIVAS.txt', msg)
                self.add_file_record(conn, paquete_id, categoria, placeholder, 'txt', 'PENDIENTE', msg)
                generated.append({'categoria': categoria, 'archivo': placeholder.name, 'estado': 'PENDIENTE'})
            return generated, errores

        try:
            from generador_formatos import GeneradorFormatos
        except Exception as exc:
            msg = f'No se pudo cargar el generador de formatos: {exc}'
            for folder, categoria in [(folder_bien, '01_Bienestarina'), (folder_rpp, '02_RPP'), (folder_ram, '03_RAM_RAN_RRAN')]:
                err = self._safe_write_text(folder / 'ERROR_GENERADOR_FORMATOS.txt', msg)
                self.add_file_record(conn, paquete_id, categoria, err, 'txt', 'ERROR', msg)
                errores.append({'categoria': categoria, 'error': msg})
            return generated, errores

        def run_one(unidad: str, categoria: str, folder: Path, metodo: str, label: str):
            try:
                gen = GeneradorFormatos(self.database_path, str(self.templates_folder), str(folder))
                ruta = getattr(gen, metodo)(mes, anio, unidad)
                item = self._add_generated_if_exists(conn, paquete_id, categoria, ruta)
                if item:
                    generated.append(item)
                else:
                    msg = f'{label} para {unidad} no devolvió archivo físico.'
                    placeholder = self._safe_write_text(folder / f'PENDIENTE_{safe_filename(unidad)}_{safe_filename(label)}.txt', msg)
                    self.add_file_record(conn, paquete_id, categoria, placeholder, 'txt', 'PENDIENTE', msg)
                    generated.append({'categoria': categoria, 'archivo': placeholder.name, 'estado': 'PENDIENTE'})
            except Exception as exc:
                msg = f'Error generando {label} para {unidad}: {exc}'
                placeholder = self._safe_write_text(folder / f'ERROR_{safe_filename(unidad)}_{safe_filename(label)}.txt', msg)
                self.add_file_record(conn, paquete_id, categoria, placeholder, 'txt', 'ERROR', msg)
                generated.append({'categoria': categoria, 'archivo': placeholder.name, 'estado': 'ERROR'})
                errores.append({'categoria': categoria, 'unidad': unidad, 'componente': label, 'error': str(exc)})

        for unidad in unidades:
            run_one(unidad, '01_Bienestarina', folder_bien, 'generar_bienestarina', 'Bienestarina')
            run_one(unidad, '02_RPP', folder_rpp, 'generar_rpp', 'RPP')
            run_one(unidad, '03_RAM_RAN_RRAN', folder_ram, 'generar_asistencia', 'RAM')
        return generated, errores

    def copy_existing_formats(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, anio: int, mes: int, only_if_empty: bool = True) -> list[dict[str, Any]]:
        """Copia formatos existentes como respaldo, sin fallar por rutas faltantes.

        Se usa solo cuando una carpeta quedó vacía, para no mezclar formatos viejos con
        los recién generados en Alpha17.
        """
        categories = {
            '01_Bienestarina': ['bienestarina', 'bienesterina'],
            '02_RPP': ['rpp'],
            # Clave interna histórica; solo se admiten archivos RAM/asistencia.
            '03_RAM_RAN_RRAN': ['ram', 'asistencia'],
        }
        copied: list[dict[str, Any]] = []
        period_tokens = {
            f'{anio}{mes:02d}',
            f'{anio}-{mes:02d}',
            str(anio),
            MESES_ES.get(mes, '').lower(),
            normalizar_texto(MESES_ES.get(mes, '')),
        }
        allowed_suffixes = {'.xlsx', '.xlsm', '.xls', '.pdf', '.csv', '.txt', '.zip'}
        for folder, tokens in categories.items():
            dest_dir = ensure_dir(package_dir / folder)
            if only_if_empty and any(p.is_file() for p in dest_dir.iterdir()):
                continue
            candidates: list[tuple[float, float, Path]] = []
            try:
                iterator = self.output_folder.rglob('*') if self.output_folder.exists() else []
                for p in iterator:
                    try:
                        if not p.is_file() or p.suffix.lower() not in allowed_suffixes:
                            continue
                        if 'paquete_mensual' in p.parts:
                            continue
                        name_norm = normalizar_texto(p.name)
                        name_low = p.name.lower()
                        if not any(tok in name_norm or tok in name_low for tok in tokens):
                            continue
                        score = 0.0
                        if any(tok and (tok in name_low or tok in name_norm) for tok in period_tokens):
                            score += 10_000_000
                        score += p.stat().st_mtime
                        candidates.append((score, p.stat().st_mtime, p))
                    except OSError:
                        continue
            except Exception:
                candidates = []
            candidates.sort(reverse=True)
            selected = [p for _, _, p in candidates[:300]]
            if not selected:
                readme = dest_dir / 'SIN_ARCHIVOS_GENERADOS_PREVIOS.txt'
                msg = 'No se encontraron archivos generados previamente para esta categoría. Se dejó la carpeta creada para revisión.'
                readme.write_text(msg + '\n', encoding='utf-8')
                self.add_file_record(conn, paquete_id, folder, readme, 'txt', 'PENDIENTE', msg)
                copied.append({'categoria': folder, 'archivo': readme.name, 'estado': 'PENDIENTE'})
                continue
            for src in selected:
                try:
                    dst = dest_dir / src.name
                    if dst.exists():
                        dst = dest_dir / f"{src.stem}_{int(src.stat().st_mtime)}{src.suffix}"
                    shutil.copy2(src, dst)
                    self.add_file_record(conn, paquete_id, folder, dst, dst.suffix.lstrip('.'), 'COPIADO')
                    copied.append({'categoria': folder, 'archivo': dst.name, 'estado': 'COPIADO'})
                except Exception as exc:
                    msg = f'No se pudo copiar {src}: {exc}'
                    err = dest_dir / f'ERROR_COPIA_{safe_filename(src.name)}.txt'
                    err.write_text(msg, encoding='utf-8')
                    self.add_file_record(conn, paquete_id, folder, err, 'txt', 'ERROR', msg)
                    copied.append({'categoria': folder, 'archivo': err.name, 'estado': 'ERROR'})
        return copied

    def reporte_relacion_mes(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_nombre: str, fundacion_id: int | None) -> None:
        beneficiarios = self.get_beneficiarios(fundacion_id)
        talento = self.get_talento(fundacion_id)
        resumen = consolidar_por_unidad(beneficiarios, anio, mes)
        rows = []
        pdf_rows = []
        first_excel_row = 7  # dos metadatos, línea en blanco y encabezado
        for index, unidad in enumerate(sorted(resumen)):
            d = resumen[unidad]
            qty = cantidades(d)
            excel_row = first_excel_row + index
            docente = docente_mas_frecuente(d) or self.docente_unidad(unidad, talento)
            rows.append([
                unidad, docente, d['gestantes'], d['menores_6'], d['seis_11'], d['uno_2'], d['tres_5'], d['sin_clasificar'],
                f'=SUM(C{excel_row}:H{excel_row})', f'=(C{excel_row}+D{excel_row}+F{excel_row}+G{excel_row}+H{excel_row})*30',
                f'=E{excel_row}*15', f'=SUM(J{excel_row}:K{excel_row})', f'=ROUNDUP(L{excel_row}/30,0)',
                f'=QUOTIENT(M{excel_row},7)', f'=MOD(M{excel_row},7)', d['verduras_dobles'],
                f'=I{excel_row}+P{excel_row}', f'=IF(I{excel_row}>0,1,0)', f'=I{excel_row}'
            ])
            pdf_rows.append([
                unidad, docente, d['gestantes'], d['menores_6'], d['seis_11'], d['uno_2'], d['tres_5'], d['sin_clasificar'],
                qty['total'], qty['huevos_30'], qty['huevos_15'], qty['total_huevos'], qty['cubetas_30'],
                qty['panales_7'], qty['cubetas_sueltas'], d['verduras_dobles'], qty['verduras'],
                qty['olla_comunitaria'], qty['bienestarina']
            ])
        if rows:
            total_row = first_excel_row + len(rows)
            rows.append(['TOTAL GENERAL', ''] + [f'=SUM({get_column_letter(col)}{first_excel_row}:{get_column_letter(col)}{total_row - 1})' for col in range(3, 20)])
        headers = [
            'Unidad', 'Docente', 'Gestantes', 'Menores 6 meses', '6 a 11 meses', '1 a 2 años 11 meses',
            '3 a 5 años 11 meses', 'Sin clasificar / revisar', 'Total usuarios', 'Huevos grupos de 30',
            'Huevos 6 a 11 (15)', 'Total huevos (unidades)', 'Cubetas de 30',
            'Panales completos (7 cubetas)', 'Cubetas sueltas', 'Gestantes/lactantes con doble verdura',
            'Total verduras', 'Olla comunitaria', 'Bienestarina'
        ]
        folder = ensure_dir(package_dir / '04_Relacion_Mes')
        xlsx = folder / f'RELACION_MES_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'RELACION_MES_{anio}_{mes:02d}.pdf'
        meta = [
            ('Fundación', fundacion_nombre),
            ('Periodo', f'{MESES_ES[mes]} {anio}. Regla huevos: 30 por usuario; 6 a 11 meses recibe 15. Panal = 7 cubetas de 30; entrega exacta en panales completos más cubetas sueltas.'),
        ]
        self.write_excel(xlsx, f'Relación del mes - {MESES_ES[mes]} {anio}', {'Relación': {'headers': headers, 'rows': rows, 'meta': meta}})
        self.write_pdf(pdf, f'Relación del mes - {MESES_ES[mes]} {anio}', headers, pdf_rows, meta)
        self.add_file_record(conn, paquete_id, '04_Relacion_Mes', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '04_Relacion_Mes', pdf, 'pdf')

    def generar_cuentas_cobro(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> None:
        folder = ensure_dir(package_dir / '05_Cuentas_Cobro')
        periodo = periodo_key(anio, mes)
        cur = conn.cursor()
        rows = []
        if self.table_exists(cur, 'cuentas_cobro_generadas'):
            cols = {r['name'] for r in cur.execute("PRAGMA table_info(cuentas_cobro_generadas)").fetchall()}
            where = "periodo=?"
            params: list[Any] = [periodo]
            if fundacion_id and 'fundacion_id' in cols:
                where += " AND COALESCE(fundacion_id, ?) = ?"
                params += [fundacion_id, fundacion_id]
            rows_gen = [dict(r) for r in cur.execute(f"SELECT * FROM cuentas_cobro_generadas WHERE {where}", tuple(params)).fetchall()]
            for r in rows_gen:
                ruta = Path(r.get('ruta_archivo') or '')
                if ruta.exists():
                    dst = folder / ruta.name
                    shutil.copy2(ruta, dst)
                    self.add_file_record(conn, paquete_id, '05_Cuentas_Cobro', dst, dst.suffix.lstrip('.'))
                rows.append([r.get('docente_nombre'), r.get('documento'), r.get('unidad'), r.get('periodo'), r.get('numero_cuenta'), r.get('nombre_archivo')])
        if not rows:
            talento = self.get_talento(fundacion_id)
            for idx, t in enumerate(talento, 1):
                cargo = normalizar_texto(t.get('cargo') or t.get('tipo_equipo') or '')
                if any(x in cargo for x in ['docente', 'agente', 'pedagoga', 'psicosocial', 'enfermera', 'nutricionista', 'administrativo']):
                    rows.append([t.get('nombre'), t.get('documento'), t.get('unidad'), periodo, idx, 'Pendiente de generar desde plantilla DOCX'])
        headers = ['Docente / Talento', 'Documento', 'Unidad', 'Periodo', 'Número cuenta', 'Archivo']
        xlsx = folder / f'CUENTAS_COBRO_RESUMEN_{periodo}.xlsx'
        pdf = folder / f'CUENTAS_COBRO_RESUMEN_{periodo}.pdf'
        self.write_excel(xlsx, f'Cuentas de cobro - {periodo}', {'Cuentas': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Cuentas de cobro - {periodo}', headers, rows)
        self.add_file_record(conn, paquete_id, '05_Cuentas_Cobro', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '05_Cuentas_Cobro', pdf, 'pdf')

    def reporte_nutricional(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> None:
        folder = ensure_dir(package_dir / '06_Informe_Nutricional')
        cur = conn.cursor()
        rows = []
        if self.table_exists(cur, 'master_salud_nutricion'):
            cols = {r['name'] for r in cur.execute("PRAGMA table_info(master_salud_nutricion)").fetchall()}
            where = "1=1"
            params = []
            if 'fundacion_id' in cols and fundacion_id:
                where += " AND COALESCE(s.fundacion_id, ?) = ?"
                params += [fundacion_id, fundacion_id]
            sql = f"""
                SELECT n.unidad_servicio AS unidad, n.nombre_completo, s.documento,
                       n.edad_meses AS edad_texto, n.sexo, s.peso AS peso_kg,
                       s.talla AS talla_cm, NULL AS imc,
                       s.diagnostico_nutricional AS diagnostico_global,
                       s.estado_nutricional AS nivel_alerta, s.fecha_toma AS fecha_valoracion,
                       NULL AS proximo_control, NULL AS trimestre, s.estado_nutricional AS estado_control
                FROM master_salud_nutricion s
                LEFT JOIN master_ninos n ON n.version_id=s.version_id AND n.fundacion_id=s.fundacion_id AND n.documento=s.documento AND n.activo=1
                WHERE s.activo=1 AND {where} ORDER BY n.unidad_servicio, n.nombre_completo LIMIT 5000
            """
            rows = [[r['unidad'], r['nombre_completo'], r['documento'], r['edad_texto'], r['sexo'], r['peso_kg'], r['talla_cm'], r['imc'], r['diagnostico_global'], r['nivel_alerta'], r['fecha_valoracion'], r['proximo_control'], r['trimestre'], r['estado_control']] for r in cur.execute(sql, tuple(params)).fetchall()]
        headers = ['Unidad', 'Nombre', 'Documento', 'Edad', 'Sexo', 'Peso', 'Talla', 'IMC', 'Diagnóstico', 'Alerta', 'Fecha valoración', 'Próximo control', 'Trimestre', 'Estado']
        xlsx = folder / f'INFORME_NUTRICIONAL_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'INFORME_NUTRICIONAL_{anio}_{mes:02d}.pdf'
        self.write_excel(xlsx, f'Informe nutricional - {MESES_ES[mes]} {anio}', {'Nutrición': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Informe nutricional - {MESES_ES[mes]} {anio}', headers, rows)
        self.add_file_record(conn, paquete_id, '06_Informe_Nutricional', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '06_Informe_Nutricional', pdf, 'pdf')

    def reporte_novedades(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> None:
        folder = ensure_dir(package_dir / '07_Informe_Novedades')
        cur = conn.cursor()
        rows = []
        headers = ['Tipo', 'Documento', 'Nombre', 'Unidad anterior', 'Unidad actual', 'Campo', 'Valor anterior', 'Valor actual', 'Observación']
        if self.table_exists(cur, 'cb_detalles'):
            sql = "SELECT * FROM cb_detalles ORDER BY id DESC LIMIT 5000"
            for r in cur.execute(sql).fetchall():
                d = dict(r)
                rows.append([
                    d.get('tipo') or d.get('categoria'),
                    d.get('documento') or d.get('documento_nino'),
                    d.get('nombre') or d.get('nombre_completo'),
                    d.get('unidad_anterior'),
                    d.get('unidad_actual') or d.get('unidad'),
                    d.get('campo'),
                    d.get('valor_anterior'),
                    d.get('valor_actual'),
                    d.get('observacion') or d.get('detalle')
                ])
        else:
            if self.table_exists(cur, 'movimientos'):
                for r in cur.execute("SELECT tipo, documento, nombre, unidad_origen, unidad_destino, detalle FROM movimientos ORDER BY id DESC LIMIT 1000").fetchall():
                    rows.append([r['tipo'], r['documento'], r['nombre'], r['unidad_origen'], r['unidad_destino'], '', '', '', r['detalle']])
        xlsx = folder / f'INFORME_NOVEDADES_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'INFORME_NOVEDADES_{anio}_{mes:02d}.pdf'
        self.write_excel(xlsx, f'Informe de novedades - {MESES_ES[mes]} {anio}', {'Novedades': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Informe de novedades - {MESES_ES[mes]} {anio}', headers, rows)
        self.add_file_record(conn, paquete_id, '07_Informe_Novedades', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '07_Informe_Novedades', pdf, 'pdf')

    def reporte_talento(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> None:
        folder = ensure_dir(package_dir / '08_Talento_Humano')
        talento = self.get_talento(fundacion_id)
        headers = ['Documento', 'Nombre', 'Cargo', 'Tipo equipo', 'Unidad', 'Coordinador', 'Dirección', 'Teléfono', 'Estado']
        rows = [[t.get('documento'), t.get('nombre'), t.get('cargo'), t.get('tipo_equipo'), t.get('unidad'), t.get('coordinador'), t.get('direccion'), t.get('telefono'), t.get('estado')] for t in talento]
        xlsx = folder / f'INFORME_TALENTO_HUMANO_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'INFORME_TALENTO_HUMANO_{anio}_{mes:02d}.pdf'
        self.write_excel(xlsx, f'Informe de talento humano - {MESES_ES[mes]} {anio}', {'Talento Humano': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Informe de talento humano - {MESES_ES[mes]} {anio}', headers, rows)
        self.add_file_record(conn, paquete_id, '08_Talento_Humano', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '08_Talento_Humano', pdf, 'pdf')

    def reporte_gerencial(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_nombre: str, fundacion_id: int | None) -> None:
        """Genera reporte gerencial profesional dentro del paquete mensual.

        Si el módulo independiente de reportes gerenciales no está disponible,
        conserva el comportamiento anterior con un resumen simple.
        """
        folder = ensure_dir(package_dir / '09_Reporte_Gerencial')
        try:
            from modules.reportes_gerenciales.services import ReportesGerencialesService
            rg = ReportesGerencialesService(self.database_path, str(self.output_folder))
            user = {'id': None, 'username': 'paquete_mensual', 'fundacion_id': fundacion_id or 1, 'rol': 'SUPERADMIN'}
            generado = rg.generar_reporte_ejecutivo(mes, anio, user, output_dir=folder, registrar=False)
            xlsx = Path(generado['excel'])
            pdf = Path(generado['pdf'])
            self.add_file_record(conn, paquete_id, '09_Reporte_Gerencial', xlsx, 'xlsx')
            self.add_file_record(conn, paquete_id, '09_Reporte_Gerencial', pdf, 'pdf')
            return
        except Exception as exc:
            print(f'No se pudo usar Reportes Gerenciales profesionales dentro del paquete: {exc}')

        beneficiarios = self.get_beneficiarios(fundacion_id)
        talento = self.get_talento(fundacion_id)
        activos = [b for b in beneficiarios if normalizar_texto(b.get('estado') or '') in {'activo', 'activa'}]
        resumen_unidades = self.grupos_por_unidad(beneficiarios)
        rows = [
            ['Fundación', fundacion_nombre],
            ['Periodo', f'{MESES_ES[mes]} {anio}'],
            ['Beneficiarios activos', len(activos)],
            ['Unidades con participantes', len(resumen_unidades)],
            ['Talento humano registrado', len(talento)],
            ['Docentes / agentes', sum(1 for t in talento if any(x in normalizar_texto(t.get('cargo') or '') for x in ['docente', 'agente']))],
            ['Coordinadores', sum(1 for t in talento if 'coord' in normalizar_texto(t.get('cargo') or ''))],
        ]
        headers = ['Indicador', 'Valor']
        xlsx = folder / f'REPORTE_GERENCIAL_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'REPORTE_GERENCIAL_{anio}_{mes:02d}.pdf'
        self.write_excel(xlsx, f'Reporte gerencial mensual - {MESES_ES[mes]} {anio}', {'Resumen': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Reporte gerencial mensual - {MESES_ES[mes]} {anio}', headers, rows)
        self.add_file_record(conn, paquete_id, '09_Reporte_Gerencial', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '09_Reporte_Gerencial', pdf, 'pdf')

    def reporte_auditoria(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, mes: int, anio: int, fundacion_id: int | None) -> None:
        folder = ensure_dir(package_dir / '10_Auditoria_Mensual')
        cur = conn.cursor()
        rows = []
        headers = ['Fecha', 'Usuario', 'Acción', 'Módulo/Tabla', 'Detalle', 'Archivo']
        if self.table_exists(cur, 'auditoria'):
            for r in cur.execute("SELECT fecha, usuario, accion, tabla, cambios_detectados, archivo FROM auditoria ORDER BY id DESC LIMIT 2000").fetchall():
                rows.append([r['fecha'], r['usuario'], r['accion'], r['tabla'], r['cambios_detectados'], r['archivo']])
        if self.table_exists(cur, 'auditoria_seguridad'):
            # Tolerante con distintas estructuras.
            cols = [c['name'] for c in cur.execute("PRAGMA table_info(auditoria_seguridad)").fetchall()]
            fecha_col = 'fecha_accion' if 'fecha_accion' in cols else ('fecha' if 'fecha' in cols else cols[0])
            accion_col = 'accion' if 'accion' in cols else cols[0]
            for r in cur.execute(f"SELECT * FROM auditoria_seguridad ORDER BY id DESC LIMIT 500").fetchall():
                d = dict(r)
                rows.append([d.get(fecha_col), d.get('usuario') or d.get('username'), d.get(accion_col), 'seguridad', json.dumps(d, ensure_ascii=False)[:500], ''])
        xlsx = folder / f'AUDITORIA_MENSUAL_{anio}_{mes:02d}.xlsx'
        pdf = folder / f'AUDITORIA_MENSUAL_{anio}_{mes:02d}.pdf'
        self.write_excel(xlsx, f'Auditoría mensual - {MESES_ES[mes]} {anio}', {'Auditoría': {'headers': headers, 'rows': rows}})
        self.write_pdf(pdf, f'Auditoría mensual - {MESES_ES[mes]} {anio}', headers, rows)
        self.add_file_record(conn, paquete_id, '10_Auditoria_Mensual', xlsx, 'xlsx')
        self.add_file_record(conn, paquete_id, '10_Auditoria_Mensual', pdf, 'pdf')

    def write_manifest(self, package_dir: Path, periodo: str, user: dict[str, Any], archivos: list[dict[str, Any]]) -> Path:
        manifest = {
            'periodo': periodo,
            'generado_en': now_iso(),
            'generado_por': user.get('username') or user.get('email') or 'sistema',
            'fundacion_id': user.get('fundacion_id') or 1,
            'categorias': [dict(folder=f, nombre=n) for f, n in CATEGORIAS_PAQUETE],
            'archivos': archivos,
            'nota': 'El paquete mensual consolida archivos generados y reportes automáticos. No modifica plantillas oficiales ICBF.'
        }
        path = package_dir / 'manifest.json'
        path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
        return path

    def _component_error(self, package_dir: Path, conn: sqlite3.Connection, paquete_id: int, categoria: str, exc: Exception) -> dict[str, Any]:
        folder = ensure_dir(package_dir / categoria)
        msg = f'No se pudo generar este componente: {exc}'
        path = folder / 'ERROR_COMPONENTE.txt'
        path.write_text(msg, encoding='utf-8')
        self.add_file_record(conn, paquete_id, categoria, path, 'txt', 'ERROR', msg)
        return {'categoria': categoria, 'archivo': path.name, 'estado': 'ERROR', 'error': str(exc)}

    def _build_zip(self, package_dir: Path, zip_path: Path) -> int:
        """Crea/recrea el ZIP final con rutas relativas limpias y tolerancia a archivos faltantes."""
        zip_path.parent.mkdir(parents=True, exist_ok=True)
        if zip_path.exists():
            zip_path.unlink()
        total_files = 0
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for p in sorted(package_dir.rglob('*')):
                if p.is_file():
                    try:
                        zf.write(p, p.relative_to(package_dir))
                        total_files += 1
                    except FileNotFoundError:
                        continue
        return total_files

    def generate_package(self, mes: int, anio: int, user: dict[str, Any], opciones: dict[str, Any] | None = None) -> dict[str, Any]:
        self.init_schema()
        mes = max(1, min(12, int(mes)))
        anio = int(anio)
        periodo = periodo_key(anio, mes)
        fundacion_id = int(user.get('fundacion_id') or 1)
        fundacion_nombre = self.get_fundacion_nombre(fundacion_id)
        paquete_id = self.insert_package(periodo, mes, anio, user, 'Paquete mensual completo actualizado')

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        package_dir = ensure_dir(self.paquetes_folder / f'PAQUETE_MENSUAL_{periodo}_{timestamp}')
        for folder, _ in CATEGORIAS_PAQUETE:
            ensure_dir(package_dir / folder)

        archivos_manifest: list[dict[str, Any]] = []
        componentes: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []
        conn = self.connect()
        try:
            self._ensure_runtime_columns(conn)

            # 1) Formatos operativos actualizados: no depende de archivos antiguos.
            generados, errs = self.generar_formatos_operativos_actualizados(package_dir, conn, paquete_id, mes, anio, fundacion_id)
            archivos_manifest.extend(generados)
            componentes.append({'componente': 'Formatos oficiales operativos', 'estado': 'OK' if not errs else 'PARCIAL', 'total': len(generados)})
            errores.extend(errs)

            # 2) Respaldo de archivos ya generados solo para carpetas que quedaron vacías.
            try:
                copied = self.copy_existing_formats(package_dir, conn, paquete_id, anio, mes, only_if_empty=True)
                archivos_manifest.extend(copied)
                if copied:
                    componentes.append({'componente': 'Formatos existentes de respaldo', 'estado': 'OK', 'total': len(copied)})
            except Exception as exc:
                errores.append({'componente': 'Formatos existentes de respaldo', 'error': str(exc)})

            tasks = [
                ('04_Relacion_Mes', 'Relación del Mes', lambda: self.reporte_relacion_mes(package_dir, conn, paquete_id, mes, anio, fundacion_nombre, fundacion_id)),
                ('05_Cuentas_Cobro', 'Cuentas de Cobro', lambda: self.generar_cuentas_cobro(package_dir, conn, paquete_id, mes, anio, fundacion_id)),
                ('06_Informe_Nutricional', 'Informe Nutricional', lambda: self.reporte_nutricional(package_dir, conn, paquete_id, mes, anio, fundacion_id)),
                ('07_Informe_Novedades', 'Novedades', lambda: self.reporte_novedades(package_dir, conn, paquete_id, mes, anio, fundacion_id)),
                ('08_Talento_Humano', 'Talento Humano', lambda: self.reporte_talento(package_dir, conn, paquete_id, mes, anio, fundacion_id)),
                ('09_Reporte_Gerencial', 'Reporte Gerencial', lambda: self.reporte_gerencial(package_dir, conn, paquete_id, mes, anio, fundacion_nombre, fundacion_id)),
                ('10_Auditoria_Mensual', 'Auditoría Mensual', lambda: self.reporte_auditoria(package_dir, conn, paquete_id, mes, anio, fundacion_id)),
            ]
            for categoria, nombre, fn in tasks:
                before = {p for p in (package_dir / categoria).rglob('*') if p.is_file()} if (package_dir / categoria).exists() else set()
                try:
                    fn()
                    after = {p for p in (package_dir / categoria).rglob('*') if p.is_file()} if (package_dir / categoria).exists() else set()
                    nuevos = after - before
                    componentes.append({'componente': nombre, 'categoria': categoria, 'estado': 'OK', 'total': len(nuevos)})
                except Exception as exc:
                    item = self._component_error(package_dir, conn, paquete_id, categoria, exc)
                    archivos_manifest.append(item)
                    componentes.append({'componente': nombre, 'categoria': categoria, 'estado': 'ERROR', 'error': str(exc)})
                    errores.append({'componente': nombre, 'categoria': categoria, 'error': str(exc)})

            # Asegurar que todas las categorías tengan al menos una trazabilidad.
            for folder, label in CATEGORIAS_PAQUETE:
                folder_path = ensure_dir(package_dir / folder)
                if not any(p.is_file() for p in folder_path.rglob('*')):
                    msg = f'No se generaron archivos para {label}. Revise si hay datos suficientes para el periodo {periodo}.'
                    pendiente = folder_path / 'PENDIENTE_SIN_DATOS.txt'
                    pendiente.write_text(msg, encoding='utf-8')
                    self.add_file_record(conn, paquete_id, folder, pendiente, 'txt', 'PENDIENTE', msg)

            # Captura todos los archivos finales para manifest y zip.
            archivos = []
            for p in sorted(package_dir.rglob('*')):
                if p.is_file():
                    archivos.append({'categoria': p.parent.name, 'archivo': p.name, 'tamano': p.stat().st_size})
            manifest = self.write_manifest(package_dir, periodo, user, archivos)
            self.add_file_record(conn, paquete_id, '00_Manifest', manifest, 'json')

            zip_name = f'PAQUETE_MENSUAL_COMPLETO_{periodo}_{timestamp}.zip'
            zip_path = self.paquetes_folder / zip_name
            total_files = self._build_zip(package_dir, zip_path)

            estado_final = 'GENERADO' if not errores else 'GENERADO_CON_ALERTAS'
            conn.commit()
            self.update_package(
                paquete_id,
                estado=estado_final,
                nombre_archivo=zip_name,
                ruta_zip=str(zip_path.resolve()),
                ruta_carpeta=str(package_dir.resolve()),
                total_archivos=total_files,
                tamano_bytes=zip_path.stat().st_size if zip_path.exists() else 0,
                manifest_json=json.dumps({'periodo': periodo, 'total_archivos': total_files, 'estado': estado_final}, ensure_ascii=False),
                componentes_json=json.dumps(componentes, ensure_ascii=False),
                errores_json=json.dumps(errores, ensure_ascii=False),
            )
            conn.execute("""
                INSERT INTO pm_auditoria (paquete_id, accion, detalle, usuario_id, fundacion_id, fecha_accion)
                VALUES (?, 'PAQUETE_GENERADO', ?, ?, ?, ?)
            """, (paquete_id, f'Paquete mensual {periodo} generado con {total_files} archivo(s). Estado: {estado_final}.', user.get('id'), fundacion_id, now_iso()))
            conn.commit()
            return {
                'id': paquete_id,
                'periodo': periodo,
                'archivo': zip_name,
                'url': f'/api/paquete-mensual/{paquete_id}/descargar',
                'descarga_url': f'/api/paquete-mensual/{paquete_id}/descargar',
                'estado': estado_final,
                'total_archivos': total_files,
                'tamano_bytes': zip_path.stat().st_size if zip_path.exists() else 0,
                'errores': errores,
                'componentes': componentes,
            }
        except Exception as exc:
            conn.rollback()
            # Incluso ante un error crítico se intenta dejar un ZIP diagnóstico descargable.
            try:
                error_file = package_dir / 'ERROR_PAQUETE_MENSUAL.txt'
                error_file.write_text(f'No se pudo completar el paquete mensual: {exc}', encoding='utf-8')
                zip_name = f'PAQUETE_MENSUAL_ERROR_{periodo}_{timestamp}.zip'
                zip_path = self.paquetes_folder / zip_name
                total_files = self._build_zip(package_dir, zip_path)
                self.update_package(
                    paquete_id,
                    estado='ERROR',
                    nombre_archivo=zip_name,
                    ruta_zip=str(zip_path.resolve()),
                    ruta_carpeta=str(package_dir.resolve()),
                    total_archivos=total_files,
                    tamano_bytes=zip_path.stat().st_size if zip_path.exists() else 0,
                    observaciones=str(exc),
                    errores_json=json.dumps([{'error': str(exc)}], ensure_ascii=False),
                )
            except Exception:
                self.update_package(paquete_id, estado='ERROR', observaciones=str(exc))
            raise
        finally:
            conn.close()

    def reconstruir_zip_paquete(self, paquete_id: int) -> Path | None:
        """Reconstruye el ZIP si se perdió la ruta guardada pero existe la carpeta del paquete."""
        paquete = self.get_package(paquete_id)
        if not paquete:
            return None
        ruta_zip = Path(paquete.get('ruta_zip') or '') if paquete.get('ruta_zip') else None
        if ruta_zip and ruta_zip.exists():
            return ruta_zip
        ruta_carpeta = Path(paquete.get('ruta_carpeta') or '') if paquete.get('ruta_carpeta') else None
        if not ruta_carpeta or not ruta_carpeta.exists():
            # Búsqueda tolerante por periodo en carpeta de paquetes.
            periodo = paquete.get('periodo') or ''
            candidates = sorted(self.paquetes_folder.glob(f'PAQUETE_MENSUAL_{periodo}_*'), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
            ruta_carpeta = next((p for p in candidates if p.is_dir()), None)
        if not ruta_carpeta or not ruta_carpeta.exists():
            return None
        zip_name = paquete.get('nombre_archivo') or f"PAQUETE_MENSUAL_COMPLETO_{paquete.get('periodo') or paquete_id}_RECONSTRUIDO.zip"
        zip_path = self.paquetes_folder / safe_filename(zip_name)
        total_files = self._build_zip(ruta_carpeta, zip_path)
        self.update_package(paquete_id, ruta_zip=str(zip_path.resolve()), nombre_archivo=zip_path.name, total_archivos=total_files, tamano_bytes=zip_path.stat().st_size if zip_path.exists() else 0, ruta_carpeta=str(ruta_carpeta.resolve()))
        return zip_path if zip_path.exists() else None

    def list_packages(self, limit: int = 100, fundacion_id: int | None = None, superadmin: bool = False) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        where = "1=1"
        params: list[Any] = []
        if fundacion_id and not superadmin:
            where += " AND COALESCE(fundacion_id, ?) = ?"
            params.extend([fundacion_id, fundacion_id])
        rows = [dict(r) for r in cur.execute(f"SELECT * FROM pm_paquetes WHERE {where} ORDER BY fecha_creacion DESC LIMIT ?", tuple(params + [limit])).fetchall()]
        conn.close()
        return rows

    def get_package(self, paquete_id: int) -> dict[str, Any] | None:
        return self.fetch_one('pm_paquetes', "SELECT * FROM pm_paquetes WHERE id=?", (paquete_id,))

    def list_package_files(self, paquete_id: int) -> list[dict[str, Any]]:
        return self.fetch_all('pm_archivos', "SELECT * FROM pm_archivos WHERE paquete_id=? ORDER BY categoria, nombre_archivo", (paquete_id,))
