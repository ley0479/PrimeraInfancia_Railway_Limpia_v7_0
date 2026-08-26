
"""
Servicio de plantillas oficiales Excel.

Alpha14: RPP y Bienestarina dejan de recrearse desde cero. La plataforma abre
una copia de la plantilla oficial, escribe únicamente valores en las celdas
correspondientes y conserva estilos, fuentes, colores, bordes, anchos, altos,
combinaciones, fórmulas e impresión.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import unicodedata
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from flask import Blueprint, jsonify, request, send_from_directory
except Exception:  # pragma: no cover - permite pruebas de servicio sin Flask instalado
    Blueprint = jsonify = request = send_from_directory = None
from openpyxl import load_workbook
try:
    from werkzeug.utils import secure_filename
except Exception:  # pragma: no cover
    def secure_filename(name):
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(name or "")).strip("._")

try:
    from modules.print_master import infer_print_format
except Exception:  # pragma: no cover
    infer_print_format = None

OFICIALES_SUBDIR = "oficiales"
MANIFEST_FILENAME = "templates_manifest.json"
ALLOWED_OFFICIAL_TEMPLATE_EXTENSIONS = {".xlsx", ".xlsm"}

DEFAULT_MANIFEST: dict[str, dict[str, Any]] = {
    "rpp": {
        "nombre": "Formato RPP Oficial",
        "archivo": "plantilla_rpp_oficial.xlsx",
        "archivo_versionado": "plantilla_rpp_oficial_v2026.xlsx",
        "hoja": "PLANTILLA O FORMADE RPP",
        "tipo": "excel",
        "preservar_estilos": True,
        "preservar_impresion": True,
        "version": "2026",
        "area_impresion_si_falta": "A1:AA42",
        "filas_usuarios": [12, 31],
    },
    "bienestarina": {
        "nombre": "Formato Bienestarina Oficial",
        "archivo": "plantilla_bienestarina_oficial.xlsx",
        "archivo_versionado": "plantilla_bienestarina_oficial_v2026.xlsx",
        "hoja": "plantilla de bienestarina ",
        "tipo": "excel",
        "preservar_estilos": True,
        "preservar_impresion": True,
        "version": "2026",
        "area_impresion_si_falta": "A1:T50",
        "filas_usuarios": [[10, 23], [31, 46]],
    },
    "ram": {
        "nombre": "Formato Asistencia Registro Mensual RAM V3",
        "archivo": "plantilla_ram_oficial_v3.xlsx",
        "archivo_versionado": "plantilla_ram_oficial_v3.xlsx",
        "hoja": "FORMATO RAM",
        "hoja_instrucciones": "INSTRUCCIONES DILIGENCIAMIENTO",
        "tipo": "excel",
        "preservar_estilos": True,
        "preservar_impresion": True,
        "version": "3",
        "codigo": "F27.MT1.PP",
        "fecha_vigencia": "2026-08-01",
        "hash_sha256": "a6b4c9412f7c72a19b9d5e842fa5ffd4b876c7d0f0c3d5c8e140b5287d700753",
        "area_impresion_si_falta": "A1:AO41",
        "filas_usuarios": [15, 34],
        "capacidad_por_pagina": 20,
        "versiones": [
            {
                "nombre": "Formato Registro Asistencia Mensual RAM V2 histórico sanitizado",
                "archivo": "plantilla_ram_oficial_v2_historica.xlsx",
                "archivo_versionado": "plantilla_ram_oficial_v2_historica.xlsx",
                "hoja": "FORMATO RAM V2 HISTORICO",
                "hoja_instrucciones": None,
                "version": "2",
                "codigo": "F27.MT1.PP",
                "fecha_vigencia": "2000-01-01",
                "fecha_vigencia_fin": "2026-07-31",
                "hash_sha256": "69fabc7db8460c11a75a56eb7a382279c2673b49bdce602146fde7293444f666",
                "area_impresion_si_falta": "A1:AK41",
                "filas_usuarios": [15, 34],
                "capacidad_por_pagina": 20,
                "sanitizada": True,
                "contiene_datos_reales": False,
            },
            {
                "nombre": "Formato Asistencia Registro Mensual RAM V3",
                "archivo": "plantilla_ram_oficial_v3.xlsx",
                "archivo_versionado": "plantilla_ram_oficial_v3.xlsx",
                "hoja": "FORMATO RAM",
                "hoja_instrucciones": "INSTRUCCIONES DILIGENCIAMIENTO",
                "version": "3",
                "codigo": "F27.MT1.PP",
                "fecha_vigencia": "2026-08-01",
                "fecha_vigencia_fin": "",
                "hash_sha256": "a6b4c9412f7c72a19b9d5e842fa5ffd4b876c7d0f0c3d5c8e140b5287d700753",
                "area_impresion_si_falta": "A1:AO41",
                "filas_usuarios": [15, 34],
                "capacidad_por_pagina": 20,
                "sanitizada": True,
                "contiene_datos_reales": False,
            },
        ],
    },
}


def normalizar_texto(valor: Any) -> str:
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.replace("º", "o").replace("°", "o")
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return " ".join(texto.split())


def oficiales_dir(templates_folder: str | os.PathLike[str]) -> Path:
    path = Path(templates_folder) / OFICIALES_SUBDIR
    path.mkdir(parents=True, exist_ok=True)
    (path / "backups").mkdir(exist_ok=True)
    return path


def manifest_path(templates_folder: str | os.PathLike[str]) -> Path:
    return oficiales_dir(templates_folder) / MANIFEST_FILENAME


def cargar_manifest(templates_folder: str | os.PathLike[str]) -> dict[str, dict[str, Any]]:
    path = manifest_path(templates_folder)
    if not path.exists():
        path.write_text(json.dumps(DEFAULT_MANIFEST, ensure_ascii=False, indent=2), encoding="utf-8")
        return json.loads(json.dumps(DEFAULT_MANIFEST))
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    merged = json.loads(json.dumps(DEFAULT_MANIFEST))
    for key, value in (data or {}).items():
        if isinstance(value, dict):
            merged.setdefault(key, {}).update(value)
    return merged


def guardar_manifest(templates_folder: str | os.PathLike[str], manifest: dict[str, Any]) -> None:
    manifest_path(templates_folder).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def tipo_normalizado(tipo_formato: str | None) -> str | None:
    raw = normalizar_texto(tipo_formato)
    if raw in {"listado asistencia usuarios", "listado de asistencia de usuarios", "asistencia usuarios"}:
        return "listado_asistencia_usuarios"
    if raw in {"listado usuarios", "listado de usuarios", "listado oficial usuarios"}:
        return "listado_usuarios"
    if raw in {"rpp"} or "rpp" in raw:
        return "rpp"
    if "bienestarina" in raw or "bienesterina" in raw:
        return "bienestarina"
    tokens = set(raw.split())
    if raw == "ram" or "ram" in tokens or "asistencia registro mensual" in raw:
        return "ram"
    if infer_print_format:
        detected = infer_print_format(tipo_formato)
        if detected in {"rpp", "bienestarina", "ram"}:
            return detected
    return raw or None


def get_plantilla_oficial(
    templates_folder: str | os.PathLike[str],
    tipo_formato: str,
    mes: int | None = None,
    anio: int | None = None,
) -> dict[str, Any] | None:
    tipo = tipo_normalizado(tipo_formato)
    if tipo not in {"rpp", "bienestarina", "ram"}:
        return None
    manifest = cargar_manifest(templates_folder)
    raw_info = dict(manifest.get(tipo) or {})
    if not raw_info:
        return None
    info = _select_manifest_version(raw_info, mes=mes, anio=anio)
    if not info:
        return None
    path = oficiales_dir(templates_folder) / str(info.get("archivo") or "")
    info.update({
        "codigo_tipo": tipo,
        "tipo_formato": tipo,
        "ruta": str(path),
        "existe": path.exists(),
        "tamano_bytes": path.stat().st_size if path.exists() else 0,
        "fecha_actualizacion": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if path.exists() else None,
    })
    # Compatibilidad histórica: `codigo` representaba el tipo en RPP/Bienestarina.
    info.setdefault("codigo", tipo)
    return info


def listar_plantillas_oficiales(
    templates_folder: str | os.PathLike[str],
    mes: int | None = None,
    anio: int | None = None,
) -> list[dict[str, Any]]:
    manifest = cargar_manifest(templates_folder)
    result = []
    for key in manifest.keys():
        item = get_plantilla_oficial(templates_folder, key, mes=mes, anio=anio)
        if item:
            result.append(item)
    return result



def _database_path_from_templates(templates_folder: str | os.PathLike[str]) -> Path:
    """Ubica database.sqlite3 sin acoplar este módulo a app.py."""
    return Path(templates_folder).resolve().parent / "database.sqlite3"


def _period_start(mes: int | None = None, anio: int | None = None) -> date:
    now = datetime.now()
    try:
        month = max(1, min(12, int(mes or now.month)))
    except Exception:
        month = now.month
    try:
        year = int(anio or now.year)
    except Exception:
        year = now.year
    return date(year, month, 1)


def _manifest_applies(info: dict[str, Any], mes: int | None = None, anio: int | None = None) -> bool:
    """Valida inicio y fin de vigencia de una plantilla del manifiesto."""
    periodo = _period_start(mes, anio)
    inicio_raw = str(info.get("fecha_vigencia") or "").strip()
    fin_raw = str(info.get("fecha_vigencia_fin") or "").strip()
    try:
        if inicio_raw and date.fromisoformat(inicio_raw[:10]) > periodo:
            return False
        if fin_raw and date.fromisoformat(fin_raw[:10]) < periodo:
            return False
    except Exception:
        return False
    return True


def _versiones_manifest(info: dict[str, Any]) -> list[dict[str, Any]]:
    """Expande una entrada compatible con versiones sin romper manifiestos antiguos."""
    versiones = info.get("versiones")
    if not isinstance(versiones, list) or not versiones:
        return [dict(info)]
    base = {k: v for k, v in info.items() if k != "versiones"}
    result = []
    for version in versiones:
        if not isinstance(version, dict):
            continue
        item = dict(base)
        item.update(version)
        result.append(item)
    return result or [dict(info)]


def _select_manifest_version(
    info: dict[str, Any],
    mes: int | None = None,
    anio: int | None = None,
) -> dict[str, Any] | None:
    aplicables = [item for item in _versiones_manifest(info) if _manifest_applies(item, mes=mes, anio=anio)]
    if not aplicables:
        return None

    def sort_key(item: dict[str, Any]):
        raw = str(item.get("fecha_vigencia") or "0001-01-01")[:10]
        try:
            inicio = date.fromisoformat(raw)
        except Exception:
            inicio = date.min
        return (inicio, str(item.get("version") or ""))

    return dict(sorted(aplicables, key=sort_key, reverse=True)[0])


def _load_versioned_templates_from_db(
    templates_folder: str | os.PathLike[str],
    mes: int | None = None,
    anio: int | None = None,
) -> list[dict[str, Any]]:
    """Obtiene plantillas oficiales aplicables al periodo solicitado.

    RPP y Bienestarina conservan el comportamiento estable de versión vigente.
    RAM selecciona la última versión cuya fecha de vigencia sea anterior o igual
    al primer día del mes reportado, permitiendo generar periodos históricos.
    """
    db_path = _database_path_from_templates(templates_folder)
    if not db_path.exists():
        return []
    conn = None
    try:
        from modules.dbapi_compat import sqlite3
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        rows = []
        rows.extend(conn.execute("""
            SELECT * FROM plantillas_oficiales_versiones
            WHERE UPPER(tipo_formato) IN ('RPP','BIENESTARINA')
              AND LOWER(COALESCE(estado,''))='vigente'
            ORDER BY updated_at DESC, id DESC
        """).fetchall())

        report_date = _period_start(mes, anio).isoformat()
        ram_row = conn.execute("""
            SELECT * FROM plantillas_oficiales_versiones
            WHERE UPPER(tipo_formato)='RAM'
              AND LOWER(COALESCE(estado,'')) IN ('vigente','programado','historico','activa')
              AND (COALESCE(fecha_vigencia,'')='' OR substr(fecha_vigencia,1,10)<=?)
              AND (COALESCE(fecha_vigencia_fin,'')='' OR substr(fecha_vigencia_fin,1,10)>=?)
            ORDER BY CASE WHEN COALESCE(fecha_vigencia,'')='' THEN 1 ELSE 0 END,
                     substr(fecha_vigencia,1,10) DESC, updated_at DESC, id DESC
            LIMIT 1
        """, (report_date, report_date)).fetchone()
        if ram_row:
            rows.append(ram_row)

        result = []
        seen_types = set()
        for row in rows:
            item = dict(row)
            tipo = tipo_normalizado(item.get('tipo_formato'))
            if tipo in seen_types:
                continue
            ruta = item.get('archivo_path') or ''
            if ruta and not os.path.isabs(ruta):
                ruta = str((Path(templates_folder).resolve().parent / ruta).resolve())
            if not ruta or not Path(ruta).exists():
                continue
            seen_types.add(tipo)
            mapeos = [dict(r) for r in conn.execute(
                'SELECT * FROM plantillas_oficiales_mapeos WHERE version_id=? ORDER BY id',
                (item.get('id'),)
            ).fetchall()]
            productos = [dict(r) for r in conn.execute(
                'SELECT * FROM plantillas_oficiales_productos WHERE version_id=? AND COALESCE(activo,1)=1 ORDER BY orden,id',
                (item.get('id'),)
            ).fetchall()]
            try:
                mapeo_json = json.loads(item.get('mapeo_json') or '[]')
            except Exception:
                mapeo_json = []
            try:
                productos_json = json.loads(item.get('productos_json') or '[]')
            except Exception:
                productos_json = []
            result.append({
                'nombre': os.path.basename(ruta),
                'ruta': ruta,
                'tipo': tipo,
                'oficial': True,
                'source': 'motor_plantillas_versionado',
                'version_id': item.get('id'),
                'plantilla_oficial_version_id': item.get('id'),
                'mp_plantilla_id': item.get('mp_plantilla_id'),
                'version': item.get('version'),
                'codigo': item.get('codigo'),
                'estado': item.get('estado'),
                'fecha_vigencia': item.get('fecha_vigencia'),
                'fecha_vigencia_fin': item.get('fecha_vigencia_fin'),
                'hash_sha256': item.get('hash_sha256'),
                'manual_path': item.get('manual_path'),
                'reglas_json': item.get('reglas_json'),
                'mapeo': mapeos or mapeo_json,
                'productos': productos or productos_json,
                'mapeo_json': json.dumps(mapeos or mapeo_json, ensure_ascii=False),
                'productos_json': json.dumps(productos or productos_json, ensure_ascii=False),
                'preservar_estilos': True,
                'preservar_impresion': True,
            })
        return result
    except Exception:
        return []
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


def iter_plantillas_oficiales_para_generacion(
    templates_folder: str | os.PathLike[str],
    mes: int | None = None,
    anio: int | None = None,
) -> list[dict[str, Any]]:
    """Devuelve plantillas oficiales aplicables al mes/año del reporte."""
    plantillas = []
    versionadas = _load_versioned_templates_from_db(templates_folder, mes=mes, anio=anio)
    plantillas.extend(versionadas)
    tipos_versionados = {p.get('tipo') for p in versionadas if p.get('tipo')}
    for info in listar_plantillas_oficiales(templates_folder, mes=mes, anio=anio):
        if info and info.get("existe"):
            tipo = info.get("tipo_formato") or info.get("codigo_tipo") or tipo_normalizado(info.get("codigo"))
            if tipo in tipos_versionados:
                continue
            if not _manifest_applies(info, mes=mes, anio=anio):
                continue
            plantillas.append({
                "nombre": os.path.basename(info["ruta"]),
                "ruta": info["ruta"],
                "tipo": tipo,
                "oficial": True,
                "source": "manifest_oficial",
                "hoja": info.get("hoja"),
                "hoja_instrucciones": info.get("hoja_instrucciones"),
                "version": info.get("version"),
                "codigo": info.get("codigo"),
                "fecha_vigencia": info.get("fecha_vigencia"),
                "fecha_vigencia_fin": info.get("fecha_vigencia_fin"),
                "hash_sha256": info.get("hash_sha256"),
                "preservar_estilos": True,
                "preservar_impresion": True,
            })
    return plantillas


def _hoja_por_nombre(wb: Any, nombre: str | None):
    if nombre and nombre in wb.sheetnames:
        return wb[nombre]
    nombre_norm = normalizar_texto(nombre)
    for ws in wb.worksheets:
        if normalizar_texto(ws.title) == nombre_norm:
            return ws
    return wb.active


def _set_value(cell: Any, value: Any) -> None:
    """Regla crítica: solo se modifica el valor de la celda."""
    cell.value = value


def _cell_value(data: dict[str, Any], *keys: str, default: str = "") -> Any:
    for key in keys:
        if key in data and data.get(key) not in (None, ""):
            return data.get(key)
    return default


def _name_from_user(user: dict[str, Any]) -> str:
    parts = [
        _cell_value(user, "PrimerNombre", "primer_nombre"),
        _cell_value(user, "SegundoNombre", "segundo_nombre"),
        _cell_value(user, "PrimerApellido", "primer_apellido"),
        _cell_value(user, "SegundoApellido", "segundo_apellido"),
    ]
    name = " ".join(str(p).strip() for p in parts if str(p or "").strip())
    return name or str(_cell_value(user, "Nombre", "nombre", "nombres", default="")).strip()


def _tipo_doc(user: dict[str, Any]) -> str:
    raw = str(_cell_value(user, "TipoDocumento", "tipo_documento", default="")).strip().upper()
    if "REGISTRO" in raw or raw in {"REGISTRO CIVIL", "RC"}:
        return "RC"
    if "TARJETA" in raw or raw in {"TI"}:
        return "TI"
    if "CEDULA" in raw or "CÉDULA" in raw or raw in {"CC", "C.C."}:
        return "C.C."
    return raw or "RC"


def _edad_meses(user: dict[str, Any]) -> int | None:
    value = _cell_value(user, "EdadMeses", "edad_meses", default=None)
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _grupo_edad_marker(user: dict[str, Any]) -> str:
    tipo = normalizar_texto(_cell_value(user, "TipoBeneficiario", "tipo_beneficiario", default=""))
    grupo = normalizar_texto(_cell_value(user, "GrupoEdad", "grupo_edad", default=""))
    edad = _edad_meses(user)
    if "gestante" in tipo or "gestante" in grupo or "0 a 6" in grupo or "0 a 5" in grupo or (edad is not None and edad <= 5):
        return "D"
    if "6 a 11" in grupo or (edad is not None and 6 <= edad <= 11):
        return "E"
    if "1 a 2" in grupo or (edad is not None and 12 <= edad <= 35):
        return "F"
    if "3 a 5" in grupo or (edad is not None and 36 <= edad <= 71):
        return "G"
    return ""


def _filas_rango(ranges: Any) -> list[int]:
    result: list[int] = []
    for item in ranges or []:
        if isinstance(item, int):
            result.append(item)
        elif isinstance(item, (list, tuple)) and len(item) == 2:
            result.extend(range(int(item[0]), int(item[1]) + 1))
    return result


def asegurar_area_impresion_si_falta(ws: Any, area: str | None) -> None:
    if not area:
        return
    try:
        actual = ws.print_area
        if actual:
            return
    except Exception:
        pass
    try:
        ws.print_area = area
    except Exception:
        pass


def _clear_cells(ws: Any, row: int, cols: Iterable[str]) -> None:
    for col in cols:
        try:
            _set_value(ws[f"{col}{row}"], "")
        except Exception:
            pass


def _write_bienestarina(ws: Any, usuarios: list[dict[str, Any]], metadata: dict[str, Any], info: dict[str, Any]) -> None:
    # Encabezados: solo valores en celdas de datos, no etiquetas.
    headers = {
        "C2": _cell_value(metadata, "regional", "Regional"),
        "C3": _cell_value(metadata, "centro_zonal", "CentroZonal"),
        "C4": _cell_value(metadata, "municipio", "Municipio"),
        "C5": _cell_value(metadata, "modalidad", "Modalidad"),
        "J1": _cell_value(metadata, "codigo_uds", "codigo_unidad", "CodigoUnidadServicio"),
        "J2": _cell_value(metadata, "unidad", "Unidad"),
        "J3": _cell_value(metadata, "responsable", "agente_educativo", "docente"),
        "Q3": _cell_value(metadata, "suplente", "Suplente"),
        "J4": _cell_value(metadata, "direccion", "direccion_unidad", "DireccionUnidad"),
        "O4": _cell_value(metadata, "barrio", "Barrio"),
        "S4": _cell_value(metadata, "telefono", "telefono_docente", "Telefono"),
        "J5": _cell_value(metadata, "codigo_origen", "CodigoUnidadOrigen", "codigo_uds", "codigo_unidad"),
        "R5": _cell_value(metadata, "unidad_origen", "NombreUnidadOrigen", "unidad", "Unidad"),
        "R1": f"AÑO: {_cell_value(metadata, 'anio', 'año', 'year', default=datetime.now().year)}",
    }
    mes = _cell_value(metadata, "mes", "Mes", default="")
    if mes:
        headers["N1"] = f"MES DE CONSUMO: {mes}"
    for cell, value in headers.items():
        if value not in (None, ""):
            _set_value(ws[cell], value)

    filas = _filas_rango(info.get("filas_usuarios") or [[10, 23], [31, 46]])
    data_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
    fecha = _cell_value(metadata, "fecha_entrega", "FechaEntrega")
    lote = _cell_value(metadata, "lote", "lote_bienestarina", "Lote")
    cantidad = _cell_value(metadata, "cantidad", "cantidad_bienestarina", default=1)

    for idx, row in enumerate(filas):
        _clear_cells(ws, row, data_cols)
        if idx >= len(usuarios):
            # Mantener consecutivo visual existente solo si se desea imprimir en blanco.
            _set_value(ws[f"A{row}"], idx + 1)
            continue
        user = usuarios[idx]
        _set_value(ws[f"A{row}"], idx + 1)
        _set_value(ws[f"B{row}"], _cell_value(user, "PrimerNombre", "primer_nombre"))
        _set_value(ws[f"C{row}"], _cell_value(user, "SegundoNombre", "segundo_nombre"))
        _set_value(ws[f"D{row}"], _cell_value(user, "PrimerApellido", "primer_apellido"))
        _set_value(ws[f"E{row}"], _cell_value(user, "SegundoApellido", "segundo_apellido"))
        _set_value(ws[f"F{row}"], _tipo_doc(user))
        _set_value(ws[f"G{row}"], _cell_value(user, "NUI", "Documento", "documento"))
        _set_value(ws[f"H{row}"], fecha)
        _set_value(ws[f"I{row}"], lote)
        _set_value(ws[f"J{row}"], cantidad)
        acudiente = str(_cell_value(user, "Acudiente", "nombre_acudiente", "NombreAcudiente", default="")).strip()
        doc_acudiente = str(_cell_value(user, "DocumentoAcudiente", "documento_acudiente", default="")).strip()
        _set_value(ws[f"Q{row}"], f"{acudiente} {doc_acudiente}".strip())
        _set_value(ws[f"R{row}"], _cell_value(user, "Parentesco", "parentesco"))
        _set_value(ws[f"S{row}"], "")


def _write_rpp(ws: Any, usuarios: list[dict[str, Any]], metadata: dict[str, Any], info: dict[str, Any]) -> None:
    # Encabezados con texto completo en las celdas oficiales existentes.
    headers = {
        "A4": f"REGIONAL: {_cell_value(metadata, 'regional', 'Regional')}",
        "F4": f"CENTRO ZONAL: {_cell_value(metadata, 'centro_zonal', 'CentroZonal')}",
        "J4": f"MUNICIPIO: {_cell_value(metadata, 'municipio', 'Municipio')}",
        "Q4": _cell_value(metadata, 'servicio_atencion', 'ServicioAtencion', 'modalidad', 'Modalidad'),
        "U4": f"NOMBRE UNIDAD DE SERVICIO / UNIDAD DE ATENCIÓN: {_cell_value(metadata, 'unidad', 'Unidad')}",
        "A5": f"MODALIDAD DE ATENCIÓN: {_cell_value(metadata, 'modalidad', 'Modalidad')}",
        "N5": f"TÉLEFONO: {_cell_value(metadata, 'telefono', 'Telefono', 'telefono_docente')}",
        "A6": f"ENTIDAD ADMINISTRADORA DEL SERVICIO: {_cell_value(metadata, 'eas', 'NombreEAS')}",
        "O6": f"CÓDIGO CUENTAME DE LA UDS/UA/UCA/UE: {_cell_value(metadata, 'codigo_unidad', 'codigo_uds', 'CodigoUnidadServicio')}",
        "T7": f"MES: {_cell_value(metadata, 'mes', 'Mes')}",
    }
    contrato = _cell_value(metadata, "contrato", "NumeroContrato")
    if contrato:
        headers["A8"] = contrato
    for cell, value in headers.items():
        if str(value).strip().split(":")[-1].strip():
            _set_value(ws[cell], value)

    fila_inicio, fila_fin = info.get("filas_usuarios") or [12, 31]
    filas = list(range(int(fila_inicio), int(fila_fin) + 1))
    data_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "V", "W", "X", "Y", "Z", "AA"]
    minuta_cols = ["H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
    patron_minuta = {col: ws[f"{col}{fila_inicio}"].value for col in minuta_cols if ws[f"{col}{fila_inicio}"].value not in (None, "")}

    for idx, row in enumerate(filas):
        _clear_cells(ws, row, data_cols)
        _set_value(ws[f"A{row}"], idx + 1)
        if idx >= len(usuarios):
            continue
        user = usuarios[idx]
        _set_value(ws[f"B{row}"], _cell_value(user, "NUI", "Documento", "documento"))
        _set_value(ws[f"C{row}"], _name_from_user(user))
        marker_col = _grupo_edad_marker(user)
        if marker_col:
            _set_value(ws[f"{marker_col}{row}"], "X")
        for col, value in patron_minuta.items():
            _set_value(ws[f"{col}{row}"], value)
        _set_value(ws[f"V{row}"], "SI")
        _set_value(ws[f"W{row}"], _cell_value(user, "Acudiente", "nombre_acudiente", "NombreAcudiente"))
        _set_value(ws[f"X{row}"], _cell_value(user, "Parentesco", "parentesco"))
        _set_value(ws[f"Y{row}"], _cell_value(user, "DocumentoAcudiente", "documento_acudiente"))
        _set_value(ws[f"Z{row}"], _cell_value(user, "Telefono", "telefono"))
        _set_value(ws[f"AA{row}"], "")  # Firma o huella siempre en blanco.


def generar_desde_plantilla_oficial(
    tipo_formato: str,
    datos: dict[str, Any],
    salida: str | os.PathLike[str],
    templates_folder: str | os.PathLike[str],
) -> str:
    """Genera un Excel desde plantilla oficial sin recrear ni reestilizar.

    `datos` debe tener opcionalmente:
    - metadata: dict con encabezados.
    - usuarios: list[dict] con beneficiarios.
    """
    tipo = tipo_normalizado(tipo_formato)
    metadata = dict(datos.get("metadata") or {})
    report_year = int(metadata.get("anio") or metadata.get("año") or datetime.now().year)
    report_month = int(metadata.get("mes_numero") or metadata.get("mes") or datetime.now().month)
    candidates = iter_plantillas_oficiales_para_generacion(
        templates_folder, mes=report_month, anio=report_year
    )
    selected = next((item for item in candidates if item.get("tipo") == tipo), None)
    info = ({**DEFAULT_MANIFEST.get(tipo or "", {}), **selected} if selected else None)
    if not info or not info.get("ruta") or not Path(str(info.get("ruta"))).is_file():
        raise FileNotFoundError(
            "No se encontró la plantilla oficial de este formato. "
            "Cargue la plantilla desde Administración > Plantillas oficiales."
        )

    wb = load_workbook(info["ruta"], data_only=False, keep_vba=str(info["ruta"]).lower().endswith(".xlsm"))
    ws = _hoja_por_nombre(wb, info.get("hoja"))
    usuarios = list(datos.get("usuarios") or [])

    if tipo == "bienestarina":
        _write_bienestarina(ws, usuarios, metadata, info)
    elif tipo == "rpp":
        _write_rpp(ws, usuarios, metadata, info)
    elif tipo == "ram":
        version = str(info.get("version") or "").strip()
        if version == "3":
            from services.ram_v3_service import generate_ram_v3
            result = generate_ram_v3(
                info["ruta"], salida, usuarios,
                report_year,
                report_month,
                metadata=metadata,
                expected_sha256=info.get("hash_sha256"),
            )
            return str(result["archivo"])
        from services.ram_historical_service import generate_ram_historical
        result = generate_ram_historical(
            info["ruta"], salida, usuarios,
            report_year,
            report_month,
            metadata=metadata,
            expected_sha256=info.get("hash_sha256"),
        )
        return str(result["archivo"])
    else:
        raise ValueError(f"Tipo de plantilla oficial no soportado: {tipo_formato}")

    # Preservar configuración existente. Solo se fija área de impresión si la plantilla no trae una.
    asegurar_area_impresion_si_falta(ws, info.get("area_impresion_si_falta"))
    Path(salida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return str(salida)


def reemplazar_plantilla_oficial(templates_folder: str | os.PathLike[str], tipo_formato: str, uploaded_file: Any) -> dict[str, Any]:
    tipo = tipo_normalizado(tipo_formato)
    if tipo not in {"rpp", "bienestarina", "ram"}:
        raise ValueError("Tipo de plantilla oficial no soportado.")
    filename = secure_filename(uploaded_file.filename or "")
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_OFFICIAL_TEMPLATE_EXTENSIONS:
        raise ValueError("La plantilla oficial debe ser .xlsx o .xlsm.")

    manifest = cargar_manifest(templates_folder)
    info = manifest[tipo]
    base_dir = oficiales_dir(templates_folder)
    destino = base_dir / info["archivo"]
    if destino.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{tipo}_{stamp}_{destino.name}"
        shutil.copy2(destino, base_dir / "backups" / backup_name)
    uploaded_file.save(destino)
    version_name = f"plantilla_{tipo}_oficial_v{info.get('version', '2026')}{ext}"
    shutil.copy2(destino, base_dir / version_name)
    info["archivo_versionado"] = version_name
    info["fecha_actualizacion"] = datetime.now().isoformat(timespec="seconds")
    manifest[tipo] = info
    guardar_manifest(templates_folder, manifest)
    return get_plantilla_oficial(templates_folder, tipo) or {}


def restaurar_ultima_plantilla(templates_folder: str | os.PathLike[str], tipo_formato: str) -> dict[str, Any]:
    tipo = tipo_normalizado(tipo_formato)
    info = get_plantilla_oficial(templates_folder, tipo or "")
    if not info:
        raise FileNotFoundError("No hay manifiesto para la plantilla solicitada.")
    backup_dir = oficiales_dir(templates_folder) / "backups"
    candidates = sorted(backup_dir.glob(f"{tipo}_*_{info['archivo']}"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("No hay copia anterior para restaurar.")
    shutil.copy2(candidates[0], info["ruta"])
    return get_plantilla_oficial(templates_folder, tipo) or {}


def register_plantillas_oficiales(app: Any, templates_folder: str) -> None:
    if Blueprint is None:
        raise RuntimeError("Flask no está disponible para registrar el módulo de plantillas oficiales.")
    bp = Blueprint("plantillas_oficiales", __name__, url_prefix="/api/plantillas-oficiales")

    @bp.route("", methods=["GET"])
    def listar():
        plantillas = listar_plantillas_oficiales(templates_folder)
        from services.listado_usuarios_docx_service import template_info
        plantillas.append(template_info(app.config.get("DATA_DIR") or Path(templates_folder).parent))
        from services.listado_asistencia_usuarios_service import template_info as attendance_template_info
        plantillas.append(attendance_template_info(app.config.get("DATA_DIR") or Path(templates_folder).parent))
        return jsonify({"plantillas": plantillas}), 200

    @bp.route("/<tipo_formato>", methods=["POST"])
    def subir(tipo_formato: str):
        if "file" not in request.files:
            return jsonify({"error": "Falta el archivo de plantilla oficial."}), 400
        try:
            tipo = tipo_normalizado(tipo_formato)
            if tipo == "listado_usuarios":
                from services.listado_usuarios_docx_service import replace_template
                plantilla = replace_template(app.config.get("DATA_DIR") or Path(templates_folder).parent, request.files["file"])
            elif tipo == "listado_asistencia_usuarios":
                from services.listado_asistencia_usuarios_service import replace_template
                plantilla = replace_template(app.config.get("DATA_DIR") or Path(templates_folder).parent, request.files["file"])
            else:
                plantilla = reemplazar_plantilla_oficial(templates_folder, tipo_formato, request.files["file"])
            return jsonify({"message": "Plantilla oficial actualizada.", "plantilla": plantilla}), 200
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/<tipo_formato>/descargar", methods=["GET"])
    def descargar(tipo_formato: str):
        tipo = tipo_normalizado(tipo_formato)
        if tipo == "listado_usuarios":
            from services.listado_usuarios_docx_service import template_info
            info = template_info(app.config.get("DATA_DIR") or Path(templates_folder).parent)
        elif tipo == "listado_asistencia_usuarios":
            from services.listado_asistencia_usuarios_service import template_info
            info = template_info(app.config.get("DATA_DIR") or Path(templates_folder).parent)
        else:
            info = get_plantilla_oficial(templates_folder, tipo_formato)
        if not info or not info.get("existe"):
            return jsonify({"error": "No se encontró la plantilla oficial solicitada."}), 404
        return send_from_directory(str(Path(info["ruta"]).parent), Path(info["ruta"]).name, as_attachment=True)

    @bp.route("/<tipo_formato>/restaurar", methods=["POST"])
    def restaurar(tipo_formato: str):
        try:
            tipo = tipo_normalizado(tipo_formato)
            if tipo == "listado_usuarios":
                from services.listado_usuarios_docx_service import restore_template
                plantilla = restore_template(app.config.get("DATA_DIR") or Path(templates_folder).parent)
            elif tipo == "listado_asistencia_usuarios":
                from services.listado_asistencia_usuarios_service import restore_template
                plantilla = restore_template(app.config.get("DATA_DIR") or Path(templates_folder).parent)
            else:
                plantilla = restaurar_ultima_plantilla(templates_folder, tipo_formato)
            return jsonify({"message": "Plantilla oficial restaurada.", "plantilla": plantilla}), 200
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    app.register_blueprint(bp)
