"""Persistencia del componente Familia, Comunidad y Redes Sociales."""
from __future__ import annotations

import json
import mimetypes
import os
from modules.dbapi_compat import sqlite3
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from services.master_data_provider import MasterDataProvider

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
try:
    from werkzeug.utils import secure_filename
except ImportError:  # Permite ejecutar pruebas aisladas sin cargar Flask/Werkzeug.
    import re as _filename_re
    def secure_filename(value: str) -> str:
        name = os.path.basename(str(value or "archivo"))
        name = _filename_re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("._")
        return name or "archivo"

from .schema import SCHEMA_SQL, SCHEMA_VERSION
from .services import COMPLETED_STATES, file_sha256, json_dump, normalize_text, now_iso, parse_json, safe_state, unit_key
from modules.motor_gestion_proyecto.services import source_key as mgp_source_key, unit_key as mgp_unit_key
from modules.seguridad.tenant_context import tenant_storage_root


class FamiliasRedesRepository:
    def __init__(self, database_path: str, data_dir: str, output_folder: str):
        self.database_path = str(database_path)
        self.data_dir = Path(data_dir).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.database_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn

    @staticmethod
    def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
        return bool(conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone())

    @staticmethod
    def _columns(conn: sqlite3.Connection, table: str) -> set[str]:
        if not FamiliasRedesRepository._table_exists(conn, table):
            return set()
        return {str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()}

    def init_schema(self) -> None:
        with self.connect() as conn:
            conn.executescript(SCHEMA_SQL)
            conn.execute(
                "INSERT INTO fcr_schema_version(id,version,fecha_actualizacion) VALUES(1,?,?) "
                "ON CONFLICT(id) DO UPDATE SET version=excluded.version,fecha_actualizacion=excluded.fecha_actualizacion",
                (SCHEMA_VERSION, now_iso()),
            )
            conn.commit()

    def audit(self, fundacion_id: int, user: dict[str, Any], action: str, entity: str, entity_id: int | None, detail: dict[str, Any] | None = None, conn: sqlite3.Connection | None = None) -> None:
        own = conn is None
        target = conn or self.connect()
        try:
            target.execute(
                "INSERT INTO fcr_auditoria(fundacion_id,usuario_id,usuario,accion,entidad,entidad_id,detalle_json,fecha) VALUES(?,?,?,?,?,?,?,?)",
                (fundacion_id, user.get("id"), user.get("username") or user.get("email") or "sistema", action, entity, entity_id, json_dump(detail or {}), now_iso()),
            )
            if own:
                target.commit()
        finally:
            if own:
                target.close()

    @staticmethod
    def _field(row: dict[str, Any], *names: str, default: Any = None) -> Any:
        for name in names:
            if name in row and row.get(name) not in (None, ""):
                return row.get(name)
        return default

    def _participant_source(self, conn: sqlite3.Connection, fundacion_id: int, source: str, participant_id: int) -> dict[str, Any] | None:
        data, effective_source = MasterDataProvider.resolve_historical_participant(
            conn, fundacion_id, source, participant_id
        )
        if not data:
            return None
        participant_id = int(data.get("id") or participant_id)
        return {
            "id": participant_id,
            "origen": effective_source,
            "documento": self._field(data, "documento", "numero_documento", "identificacion", "num_documento"),
            "nombre": self._field(data, "nombre_completo", "nombre", "nombres", default=f"Participante #{participant_id}"),
            "unidad": self._field(data, "unidad_servicio", "unidad", "uds", "uca"),
            "codigo_unidad": self._field(data, "codigo_unidad", "codigo_uds", "codigo_uca"),
            "telefono": self._field(data, "telefono", "telefono_contacto", "celular"),
            "direccion": self._field(data, "direccion", "direccion_residencia"),
            "cuidador": self._field(data, "nombre_cuidador", "cuidador_principal", "madre_padre_cuidador"),
            "raw": data,
        }

    def _resolve_family(self, conn: sqlite3.Connection, fundacion_id: int, row: dict[str, Any]) -> dict[str, Any]:
        result = dict(row)
        participant = self._participant_source(conn, fundacion_id, result.get("participante_origen") or "master_ninos", int(result.get("participante_id") or 0))
        result["participante"] = participant
        result["nombre_participante"] = participant.get("nombre") if participant else f"Participante #{result.get('participante_id')}"
        result["documento_participante"] = participant.get("documento") if participant else None
        result["caracterizacion"] = parse_json(result.get("caracterizacion_json"), {})
        return result

    def sync_family_records(self, fundacion_id: int, expediente_uca_id: int | None, unit_name: str | None, unit_id: int | None, user: dict[str, Any]) -> dict[str, Any]:
        now = now_iso(); created = 0; updated = 0
        with self.connect() as conn:
            if not self._table_exists(conn, "master_ninos"):
                return {"creados": 0, "actualizados": 0, "mensaje": "Base Maestra no disponible."}
            cols = self._columns(conn, "master_ninos")
            where = [] ; params: list[Any] = []
            if "fundacion_id" in cols:
                where.append("fundacion_id=?"); params.append(fundacion_id)
            if "activo" in cols:
                where.append("COALESCE(activo,1)=1")
            sql = "SELECT * FROM master_ninos" + (" WHERE " + " AND ".join(where) if where else "")
            rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
            expected = normalize_text(unit_name)
            for participant in rows:
                source_unit = self._field(participant, "unidad_servicio", "unidad", "uds", "uca")
                if expected and normalize_text(source_unit) != expected and normalize_text(self._field(participant, "codigo_unidad")) != expected:
                    continue
                participant_id = int(participant.get("id") or 0)
                if not participant_id:
                    continue
                existing = conn.execute("SELECT id FROM fcr_expedientes_familiares WHERE fundacion_id=? AND participante_origen='master_ninos' AND participante_id=?", (fundacion_id, participant_id)).fetchone()
                values = (
                    expediente_uca_id, unit_id, source_unit or unit_name, unit_key(self._field(participant, "codigo_unidad", default=source_unit or unit_name)),
                    self._field(participant, "nombre_cuidador", "cuidador_principal", "madre_padre_cuidador"),
                    self._field(participant, "parentesco_cuidador", "parentesco"), self._field(participant, "telefono", "telefono_contacto", "celular"),
                    self._field(participant, "correo", "email"), self._field(participant, "direccion", "direccion_residencia"), user.get("id"), now,
                )
                if existing:
                    conn.execute("""UPDATE fcr_expedientes_familiares SET expediente_uca_id=COALESCE(?,expediente_uca_id),unidad_id=COALESCE(?,unidad_id),unidad_nombre=COALESCE(?,unidad_nombre),unidad_clave=COALESCE(?,unidad_clave),cuidador_principal=COALESCE(cuidador_principal,?),parentesco=COALESCE(parentesco,?),telefono_principal=COALESCE(telefono_principal,?),correo=COALESCE(correo,?),direccion=COALESCE(direccion,?),actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?""", (*values, fundacion_id, int(existing[0])))
                    updated += 1
                else:
                    conn.execute("""INSERT INTO fcr_expedientes_familiares
                    (fundacion_id,expediente_uca_id,unidad_id,unidad_nombre,unidad_clave,participante_origen,participante_id,cuidador_principal,parentesco,telefono_principal,correo,direccion,creado_por,actualizado_por,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,'master_ninos',?,?,?,?,?,?,?,?,?,?)""",
                    (fundacion_id, expediente_uca_id, unit_id, source_unit or unit_name, unit_key(self._field(participant, "codigo_unidad", default=source_unit or unit_name)), participant_id,
                     self._field(participant, "nombre_cuidador", "cuidador_principal", "madre_padre_cuidador"), self._field(participant, "parentesco_cuidador", "parentesco"),
                     self._field(participant, "telefono", "telefono_contacto", "celular"), self._field(participant, "correo", "email"), self._field(participant, "direccion", "direccion_residencia"),
                     user.get("id"), user.get("id"), now, now)); created += 1
            self.audit(fundacion_id, user, "SINCRONIZAR_EXPEDIENTES_FAMILIARES", "fcr_expedientes_familiares", None, {"creados": created, "actualizados": updated, "unidad": unit_name}, conn)
            conn.commit()
        return {"creados": created, "actualizados": updated}

    def list_family_records(self, fundacion_id: int, filters: dict[str, Any] | None = None, limit: int = 1000) -> list[dict[str, Any]]:
        filters = filters or {}; sql = "SELECT * FROM fcr_expedientes_familiares WHERE fundacion_id=?"; params: list[Any] = [fundacion_id]
        if filters.get("unidad"):
            q = f"%{normalize_text(filters['unidad'])}%"; sql += " AND (UPPER(unidad_nombre) LIKE ? OR UPPER(unidad_clave) LIKE ?)"; params.extend([q, q])
        if filters.get("estado"):
            sql += " AND estado=?"; params.append(safe_state(filters["estado"]))
        sql += " ORDER BY unidad_nombre,id DESC LIMIT ?"; params.append(max(1, min(int(limit), 5000)))
        with self.connect() as conn:
            return [self._resolve_family(conn, fundacion_id, dict(row)) for row in conn.execute(sql, params).fetchall()]

    def family_record(self, fundacion_id: int, record_id: int) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM fcr_expedientes_familiares WHERE fundacion_id=? AND id=?", (fundacion_id, record_id)).fetchone()
            if not row:
                raise LookupError("Expediente familiar no encontrado.")
            result = self._resolve_family(conn, fundacion_id, dict(row))
            result["compromisos"] = self._commitments_conn(conn, fundacion_id, {"expediente_familiar_id": record_id})
            result["alertas"] = self._alerts_conn(conn, fundacion_id, {"expediente_familiar_id": record_id})
            result["actividades"] = [dict(x) for x in conn.execute("SELECT a.* FROM fcr_actividades a JOIN fcr_asistencias s ON s.actividad_id=a.id WHERE a.fundacion_id=? AND s.expediente_familiar_id=? GROUP BY a.id ORDER BY COALESCE(a.fecha_programada,a.fecha_creacion) DESC", (fundacion_id, record_id)).fetchall()]
            return result

    def update_family_record(self, fundacion_id: int, record_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        allowed = {"cuidador_principal", "parentesco", "telefono_principal", "correo", "direccion", "contacto_alterno_nombre", "contacto_alterno_parentesco", "contacto_alterno_telefono", "autoridad_tradicional", "autoridad_tradicional_telefono", "caracterizacion", "observaciones", "estado"}
        updates: list[str] = []; values: list[Any] = []
        for field in allowed:
            if field not in data: continue
            db_field = "caracterizacion_json" if field == "caracterizacion" else field
            value = json_dump(data[field] or {}) if field == "caracterizacion" else (safe_state(data[field]) if field == "estado" else data[field])
            updates.append(f"{db_field}=?"); values.append(value)
        if updates:
            now = now_iso(); values.extend([user.get("id"), now, fundacion_id, record_id])
            with self.connect() as conn:
                if not conn.execute("SELECT id FROM fcr_expedientes_familiares WHERE fundacion_id=? AND id=?", (fundacion_id, record_id)).fetchone(): raise LookupError("Expediente familiar no encontrado.")
                conn.execute(f"UPDATE fcr_expedientes_familiares SET {','.join(updates)},actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?", values)
                self.audit(fundacion_id, user, "ACTUALIZAR_EXPEDIENTE_FAMILIAR", "fcr_expedientes_familiares", record_id, {"campos": sorted(set(data) & allowed)}, conn); conn.commit()
        return self.family_record(fundacion_id, record_id)

    def create_activity(self, fundacion_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        title = str(data.get("titulo") or "").strip(); activity_type = safe_state(data.get("tipo"), "ACOMPANAMIENTO_FAMILIAR")
        if not title: raise ValueError("El título de la actividad es obligatorio.")
        now = now_iso(); unit_name = str(data.get("unidad_nombre") or data.get("unidad") or "").strip()
        with self.connect() as conn:
            cur = conn.execute("""INSERT INTO fcr_actividades
            (fundacion_id,expediente_uca_id,unidad_id,unidad_nombre,unidad_clave,tipo,titulo,objetivo,metodologia,lugar,fecha_programada,fecha_ejecucion,fecha_limite_cierre,estado,profesional_id,profesional_nombre,participantes_esperados,resultados,conclusiones_profesionales,compromisos_generales,observaciones,creado_por,actualizado_por,fecha_creacion,fecha_actualizacion)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,'PROGRAMADA',?,?,?,?,?,?,?,?,?,?,?)""",
            (fundacion_id, data.get("expediente_uca_id"), data.get("unidad_id"), unit_name or None, unit_key(data.get("unidad_clave") or unit_name), activity_type, title[:500], str(data.get("objetivo") or "")[:4000] or None,
             str(data.get("metodologia") or "")[:4000] or None, str(data.get("lugar") or "")[:500] or None, str(data.get("fecha_programada") or "")[:10] or None,
             str(data.get("fecha_ejecucion") or "")[:10] or None, str(data.get("fecha_limite_cierre") or data.get("fecha_programada") or "")[:10] or None,
             data.get("profesional_id") or user.get("id"), str(data.get("profesional_nombre") or user.get("username") or "")[:250], int(data.get("participantes_esperados") or 0),
             None, None, None, str(data.get("observaciones") or "")[:4000] or None, user.get("id"), user.get("id"), now, now))
            activity_id = int(cur.lastrowid)
            for record_id in data.get("expedientes_familiares") or []:
                family = conn.execute("SELECT * FROM fcr_expedientes_familiares WHERE fundacion_id=? AND id=?", (fundacion_id, int(record_id))).fetchone()
                if not family: continue
                resolved = self._resolve_family(conn, fundacion_id, dict(family))
                conn.execute("""INSERT OR IGNORE INTO fcr_asistencias
                (fundacion_id,actividad_id,expediente_familiar_id,nombre_asistente,tipo_asistente,documento_referencia,telefono,asistio,registrado_por,fecha_registro)
                VALUES(?,?,?,?,?,?,?,0,?,?)""",
                (fundacion_id, activity_id, int(record_id), resolved.get("cuidador_principal") or resolved.get("nombre_participante"), "FAMILIA", resolved.get("documento_participante"), resolved.get("telefono_principal"), user.get("id"), now))
            self.audit(fundacion_id, user, "CREAR_ACTIVIDAD_FAMILIAS", "fcr_actividades", activity_id, {"tipo": activity_type, "unidad": unit_name}, conn); conn.commit()
        docs = self.prepare_activity_documents(fundacion_id, activity_id, user, types=("ACTA", "LISTADO_ASISTENCIA"))
        self._upsert_motor_activity(fundacion_id, activity_id, user)
        result = self.activity_detail(fundacion_id, activity_id); result["documentos_preparados"] = docs["documentos"]
        return result

    def list_activities(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}; sql = "SELECT * FROM fcr_actividades WHERE fundacion_id=?"; params: list[Any] = [fundacion_id]
        for field in ("estado", "tipo"):
            if filters.get(field): sql += f" AND {field}=?"; params.append(safe_state(filters[field]))
        if filters.get("unidad"):
            q=f"%{normalize_text(filters['unidad'])}%"; sql += " AND (UPPER(unidad_nombre) LIKE ? OR UPPER(unidad_clave) LIKE ?)"; params.extend([q,q])
        if filters.get("profesional_id"):
            sql += " AND profesional_id=?"; params.append(int(filters["profesional_id"]))
        sql += " ORDER BY COALESCE(fecha_programada,fecha_creacion) DESC,id DESC LIMIT 2000"
        with self.connect() as conn:
            rows=[dict(r) for r in conn.execute(sql,params).fetchall()]
            for row in rows:
                row["asistencias"] = int(conn.execute("SELECT COUNT(*) FROM fcr_asistencias WHERE fundacion_id=? AND actividad_id=? AND asistio=1", (fundacion_id,row["id"])).fetchone()[0] or 0)
                row["documentos"] = int(conn.execute("SELECT COUNT(*) FROM fcr_documentos_generados WHERE fundacion_id=? AND actividad_id=?", (fundacion_id,row["id"])).fetchone()[0] or 0)
            return rows

    def activity_detail(self, fundacion_id: int, activity_id: int) -> dict[str, Any]:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_actividades WHERE fundacion_id=? AND id=?",(fundacion_id,activity_id)).fetchone()
            if not row: raise LookupError("Actividad no encontrada.")
            result=dict(row); attendees=[]
            for raw in conn.execute("SELECT * FROM fcr_asistencias WHERE fundacion_id=? AND actividad_id=? ORDER BY nombre_asistente",(fundacion_id,activity_id)).fetchall():
                item=dict(raw)
                if item.get("expediente_familiar_id"):
                    family=conn.execute("SELECT * FROM fcr_expedientes_familiares WHERE fundacion_id=? AND id=?",(fundacion_id,item["expediente_familiar_id"])).fetchone()
                    if family:item["familia"] = self._resolve_family(conn,fundacion_id,dict(family))
                attendees.append(item)
            result["asistencias"]=attendees
            result["compromisos"]=self._commitments_conn(conn,fundacion_id,{"actividad_id":activity_id})
            result["documentos"]=[dict(x) for x in conn.execute("SELECT * FROM fcr_documentos_generados WHERE fundacion_id=? AND actividad_id=? ORDER BY fecha_generacion DESC",(fundacion_id,activity_id)).fetchall()]
            result["evidencias"]=[dict(x) for x in conn.execute("SELECT * FROM fcr_evidencias WHERE fundacion_id=? AND actividad_id=? AND activo=1 ORDER BY fecha_carga DESC",(fundacion_id,activity_id)).fetchall()]
            return result

    def update_activity(self,fundacion_id:int,activity_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        allowed={"titulo","objetivo","metodologia","lugar","fecha_programada","fecha_ejecucion","fecha_limite_cierre","estado","profesional_id","profesional_nombre","participantes_esperados","resultados","conclusiones_profesionales","compromisos_generales","observaciones"};updates=[];values=[]
        for field in allowed:
            if field in data:
                value=safe_state(data[field]) if field=="estado" else data[field];updates.append(f"{field}=?");values.append(value)
        if updates:
            now=now_iso();values.extend([user.get("id"),now,fundacion_id,activity_id])
            with self.connect() as conn:
                if not conn.execute("SELECT id FROM fcr_actividades WHERE fundacion_id=? AND id=?",(fundacion_id,activity_id)).fetchone():raise LookupError("Actividad no encontrada.")
                conn.execute(f"UPDATE fcr_actividades SET {','.join(updates)},actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",values)
                self.audit(fundacion_id,user,"ACTUALIZAR_ACTIVIDAD_FAMILIAS","fcr_actividades",activity_id,{"campos":sorted(set(data)&allowed)},conn);conn.commit()
        self._upsert_motor_activity(fundacion_id,activity_id,user)
        return self.activity_detail(fundacion_id,activity_id)

    def update_attendance(self,fundacion_id:int,activity_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        attendance_id=int(data.get("id") or 0);now=now_iso()
        with self.connect() as conn:
            if attendance_id:
                conn.execute("UPDATE fcr_asistencias SET asistio=?,firma_referencia=?,observaciones=?,registrado_por=?,fecha_registro=? WHERE fundacion_id=? AND actividad_id=? AND id=?",(int(bool(data.get("asistio"))),str(data.get("firma_referencia") or "")[:2000] or None,str(data.get("observaciones") or "")[:2000] or None,user.get("id"),now,fundacion_id,activity_id,attendance_id))
            else:
                name=str(data.get("nombre_asistente") or "").strip()
                if not name:raise ValueError("El nombre del asistente es obligatorio.")
                conn.execute("""INSERT INTO fcr_asistencias(fundacion_id,actividad_id,expediente_familiar_id,nombre_asistente,tipo_asistente,documento_referencia,telefono,asistio,firma_referencia,observaciones,registrado_por,fecha_registro) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(fundacion_id,activity_id,data.get("expediente_familiar_id"),name[:300],safe_state(data.get("tipo_asistente"),"FAMILIA"),str(data.get("documento_referencia") or "")[:100] or None,str(data.get("telefono") or "")[:100] or None,int(bool(data.get("asistio"))),str(data.get("firma_referencia") or "")[:2000] or None,str(data.get("observaciones") or "")[:2000] or None,user.get("id"),now))
            total=int(conn.execute("SELECT COUNT(*) FROM fcr_asistencias WHERE fundacion_id=? AND actividad_id=? AND asistio=1",(fundacion_id,activity_id)).fetchone()[0] or 0);conn.execute("UPDATE fcr_actividades SET participantes_asistentes=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",(total,user.get("id"),now,fundacion_id,activity_id));self.audit(fundacion_id,user,"REGISTRAR_ASISTENCIA","fcr_actividades",activity_id,{"asistentes":total},conn);conn.commit()
        return self.activity_detail(fundacion_id,activity_id)

    def create_commitment(self,fundacion_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        title=str(data.get("titulo") or "").strip()
        if not title:raise ValueError("El compromiso requiere título.")
        now=now_iso();unit_name=str(data.get("unidad_nombre") or "").strip()
        with self.connect() as conn:
            cur=conn.execute("""INSERT INTO fcr_compromisos(fundacion_id,actividad_id,expediente_familiar_id,expediente_uca_id,unidad_nombre,unidad_clave,titulo,descripcion,responsable_id,responsable_nombre,fecha_compromiso,fecha_limite,estado,prioridad,porcentaje,creado_por,actualizado_por,fecha_creacion,fecha_actualizacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,'PENDIENTE',?,0,?,?,?,?)""",(fundacion_id,data.get("actividad_id"),data.get("expediente_familiar_id"),data.get("expediente_uca_id"),unit_name or None,unit_key(data.get("unidad_clave") or unit_name),title[:500],str(data.get("descripcion") or "")[:4000] or None,data.get("responsable_id"),str(data.get("responsable_nombre") or "")[:250] or None,str(data.get("fecha_compromiso") or date.today().isoformat())[:10],str(data.get("fecha_limite") or "")[:10] or None,safe_state(data.get("prioridad"),"MEDIA"),user.get("id"),user.get("id"),now,now));cid=int(cur.lastrowid);self.audit(fundacion_id,user,"CREAR_COMPROMISO","fcr_compromisos",cid,{"actividad_id":data.get("actividad_id")},conn);conn.commit()
        self._upsert_motor_commitment(fundacion_id,cid,user)
        return self.commitment_detail(fundacion_id,cid)

    def _commitments_conn(self,conn:sqlite3.Connection,fundacion_id:int,filters:dict[str,Any]|None=None)->list[dict[str,Any]]:
        filters=filters or {};sql="SELECT * FROM fcr_compromisos WHERE fundacion_id=?";params:list[Any]=[fundacion_id]
        for field in ("actividad_id","expediente_familiar_id"):
            if filters.get(field):sql+=f" AND {field}=?";params.append(int(filters[field]))
        if filters.get("estado"):sql+=" AND estado=?";params.append(safe_state(filters["estado"]))
        if filters.get("unidad"):
            q=f"%{normalize_text(filters['unidad'])}%";sql+=" AND (UPPER(unidad_nombre) LIKE ? OR UPPER(unidad_clave) LIKE ?)";params.extend([q,q])
        sql+=" ORDER BY COALESCE(fecha_limite,'9999-12-31'),id DESC LIMIT 3000";return [dict(x) for x in conn.execute(sql,params).fetchall()]

    def list_commitments(self,fundacion_id:int,filters:dict[str,Any]|None=None)->list[dict[str,Any]]:
        with self.connect() as conn:return self._commitments_conn(conn,fundacion_id,filters)

    def commitment_detail(self,fundacion_id:int,commitment_id:int)->dict[str,Any]:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_compromisos WHERE fundacion_id=? AND id=?",(fundacion_id,commitment_id)).fetchone()
            if not row:raise LookupError("Compromiso no encontrado.")
            result=dict(row);result["seguimientos"]=[dict(x) for x in conn.execute("SELECT * FROM fcr_seguimientos WHERE fundacion_id=? AND compromiso_id=? ORDER BY fecha DESC,id DESC",(fundacion_id,commitment_id)).fetchall()];return result

    def add_followup(self,fundacion_id:int,commitment_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        result=str(data.get("resultado") or "").strip()
        if not result:raise ValueError("El resultado del seguimiento es obligatorio.")
        current=self.commitment_detail(fundacion_id,commitment_id);pct=max(0,min(100,float(data.get("porcentaje_reportado") or current.get("porcentaje") or 0)));now=now_iso()
        with self.connect() as conn:
            cur=conn.execute("""INSERT INTO fcr_seguimientos(fundacion_id,compromiso_id,expediente_familiar_id,fecha,resultado,porcentaje_reportado,proxima_accion,fecha_proximo_seguimiento,evidencia_referencia,creado_por,fecha_creacion) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",(fundacion_id,commitment_id,current.get("expediente_familiar_id"),str(data.get("fecha") or date.today().isoformat())[:10],result[:8000],pct,str(data.get("proxima_accion") or "")[:4000] or None,str(data.get("fecha_proximo_seguimiento") or "")[:10] or None,str(data.get("evidencia_referencia") or "")[:2000] or None,user.get("id"),now));fid=int(cur.lastrowid);state="PENDIENTE_VALIDACION" if pct>=100 else "EN_SEGUIMIENTO";conn.execute("UPDATE fcr_compromisos SET porcentaje=?,estado=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",(pct,state,user.get("id"),now,fundacion_id,commitment_id));self.audit(fundacion_id,user,"AGREGAR_SEGUIMIENTO_COMPROMISO","fcr_seguimientos",fid,{"compromiso_id":commitment_id,"porcentaje":pct},conn);conn.commit()
        self._upsert_motor_commitment(fundacion_id,commitment_id,user);return self.commitment_detail(fundacion_id,commitment_id)

    def close_commitment(self,fundacion_id:int,commitment_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        observation=str(data.get("observaciones_cierre") or "").strip()
        if not observation:raise ValueError("El cierre exige una observación validada.")
        current=self.commitment_detail(fundacion_id,commitment_id)
        if float(current.get("porcentaje") or 0)<100:raise ValueError("El compromiso debe registrar 100% antes de validarlo.")
        now=now_iso()
        with self.connect() as conn:
            conn.execute("UPDATE fcr_compromisos SET estado='CERRADO',fecha_cierre=?,cierre_validado_por=?,observaciones_cierre=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",(now[:10],user.get("id"),observation[:4000],user.get("id"),now,fundacion_id,commitment_id));self.audit(fundacion_id,user,"CERRAR_COMPROMISO","fcr_compromisos",commitment_id,{},conn);conn.commit()
        self._upsert_motor_commitment(fundacion_id,commitment_id,user);return self.commitment_detail(fundacion_id,commitment_id)

    def create_network(self,fundacion_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        name=str(data.get("nombre") or "").strip();kind=safe_state(data.get("tipo_actor"),"INSTITUCION")
        if not name:raise ValueError("El nombre de la red o institución es obligatorio.")
        now=now_iso()
        with self.connect() as conn:
            cur=conn.execute("""INSERT INTO fcr_redes_apoyo(fundacion_id,nombre,tipo_actor,territorio,municipio,direccion,contacto_nombre,telefono,correo,servicios_json,rutas_json,horario,observaciones,activo,creado_por,actualizado_por,fecha_creacion,fecha_actualizacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?,?,?)""",(fundacion_id,name[:500],kind,str(data.get("territorio") or "")[:500] or None,str(data.get("municipio") or "")[:250] or None,str(data.get("direccion") or "")[:500] or None,str(data.get("contacto_nombre") or "")[:250] or None,str(data.get("telefono") or "")[:100] or None,str(data.get("correo") or "")[:250] or None,json_dump(data.get("servicios") or []),json_dump(data.get("rutas") or []),str(data.get("horario") or "")[:500] or None,str(data.get("observaciones") or "")[:4000] or None,user.get("id"),user.get("id"),now,now));rid=int(cur.lastrowid);self.audit(fundacion_id,user,"CREAR_RED_APOYO","fcr_redes_apoyo",rid,{"tipo":kind},conn);conn.commit()
        return self.network(fundacion_id,rid)

    def network(self,fundacion_id:int,network_id:int)->dict[str,Any]:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_redes_apoyo WHERE fundacion_id=? AND id=?",(fundacion_id,network_id)).fetchone()
            if not row:raise LookupError("Red de apoyo no encontrada.")
            result=dict(row);result["servicios"]=parse_json(result.get("servicios_json"),[]);result["rutas"]=parse_json(result.get("rutas_json"),[]);return result

    def list_networks(self,fundacion_id:int)->list[dict[str,Any]]:
        with self.connect() as conn:
            rows=[]
            for row in conn.execute("SELECT * FROM fcr_redes_apoyo WHERE fundacion_id=? AND activo=1 ORDER BY tipo_actor,nombre",(fundacion_id,)).fetchall():
                item=dict(row);item["servicios"]=parse_json(item.get("servicios_json"),[]);item["rutas"]=parse_json(item.get("rutas_json"),[]);rows.append(item)
            return rows

    def verify_network(self,fundacion_id:int,network_id:int,user:dict[str,Any])->dict[str,Any]:
        with self.connect() as conn:
            now=now_iso();conn.execute("UPDATE fcr_redes_apoyo SET verificado_por=?,fecha_verificacion=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",(user.get("id"),now[:10],user.get("id"),now,fundacion_id,network_id));self.audit(fundacion_id,user,"VERIFICAR_RED_APOYO","fcr_redes_apoyo",network_id,{},conn);conn.commit()
        return self.network(fundacion_id,network_id)

    def create_alert(self,fundacion_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        description=str(data.get("descripcion") or "").strip()
        if not description:raise ValueError("La alerta requiere descripción.")
        now=now_iso();unit_name=str(data.get("unidad_nombre") or "").strip()
        with self.connect() as conn:
            cur=conn.execute("""INSERT INTO fcr_alertas(fundacion_id,expediente_familiar_id,actividad_id,expediente_uca_id,unidad_nombre,unidad_clave,tipo,nivel,descripcion,estado,entidad_ruta_id,fecha_identificacion,fecha_activacion_ruta,fecha_proximo_seguimiento,responsable_id,responsable_nombre,creado_por,actualizado_por,fecha_creacion,fecha_actualizacion) VALUES(?,?,?,?,?,?,?,?,?,'ABIERTA',?,?,?,?,?,?,?,?,?,?)""",(fundacion_id,data.get("expediente_familiar_id"),data.get("actividad_id"),data.get("expediente_uca_id"),unit_name or None,unit_key(data.get("unidad_clave") or unit_name),safe_state(data.get("tipo"),"ACOMPANAMIENTO"),safe_state(data.get("nivel"),"MEDIO"),description[:8000],data.get("entidad_ruta_id"),str(data.get("fecha_identificacion") or date.today().isoformat())[:10],str(data.get("fecha_activacion_ruta") or "")[:10] or None,str(data.get("fecha_proximo_seguimiento") or "")[:10] or None,data.get("responsable_id") or user.get("id"),str(data.get("responsable_nombre") or user.get("username") or "")[:250],user.get("id"),user.get("id"),now,now));aid=int(cur.lastrowid);self.audit(fundacion_id,user,"CREAR_ALERTA_FAMILIAR","fcr_alertas",aid,{"nivel":data.get("nivel")},conn);conn.commit()
        self._upsert_motor_alert(fundacion_id,aid,user);return self.alert(fundacion_id,aid)

    def _alerts_conn(self,conn:sqlite3.Connection,fundacion_id:int,filters:dict[str,Any]|None=None)->list[dict[str,Any]]:
        filters=filters or {};sql="SELECT * FROM fcr_alertas WHERE fundacion_id=?";params:list[Any]=[fundacion_id]
        if filters.get("expediente_familiar_id"):sql+=" AND expediente_familiar_id=?";params.append(int(filters["expediente_familiar_id"]))
        if filters.get("estado"):sql+=" AND estado=?";params.append(safe_state(filters["estado"]))
        if filters.get("unidad"):
            q=f"%{normalize_text(filters['unidad'])}%";sql+=" AND (UPPER(unidad_nombre) LIKE ? OR UPPER(unidad_clave) LIKE ?)";params.extend([q,q])
        sql+=" ORDER BY CASE nivel WHEN 'CRITICO' THEN 0 WHEN 'ALTO' THEN 1 WHEN 'MEDIO' THEN 2 ELSE 3 END,COALESCE(fecha_proximo_seguimiento,'9999-12-31'),id DESC LIMIT 2000";return [dict(x) for x in conn.execute(sql,params).fetchall()]

    def list_alerts(self,fundacion_id:int,filters:dict[str,Any]|None=None)->list[dict[str,Any]]:
        with self.connect() as conn:return self._alerts_conn(conn,fundacion_id,filters)

    def alert(self,fundacion_id:int,alert_id:int)->dict[str,Any]:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_alertas WHERE fundacion_id=? AND id=?",(fundacion_id,alert_id)).fetchone()
            if not row:raise LookupError("Alerta no encontrada.")
            return dict(row)

    def update_alert(self,fundacion_id:int,alert_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        allowed={"tipo","nivel","descripcion","estado","entidad_ruta_id","fecha_activacion_ruta","fecha_proximo_seguimiento","responsable_id","responsable_nombre"};updates=[];values=[]
        for field in allowed:
            if field in data:
                value=safe_state(data[field]) if field in {"tipo","nivel","estado"} else data[field]
                if field=="estado" and value in {"CERRADA","CERRADO"}:raise PermissionError("El cierre requiere validación explícita y evidencia.")
                updates.append(f"{field}=?");values.append(value)
        if updates:
            now=now_iso();values.extend([user.get("id"),now,fundacion_id,alert_id])
            with self.connect() as conn:
                conn.execute(f"UPDATE fcr_alertas SET {','.join(updates)},actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",values);self.audit(fundacion_id,user,"ACTUALIZAR_ALERTA_FAMILIAR","fcr_alertas",alert_id,{"campos":sorted(set(data)&allowed)},conn);conn.commit()
        self._upsert_motor_alert(fundacion_id,alert_id,user);return self.alert(fundacion_id,alert_id)

    def close_alert(self,fundacion_id:int,alert_id:int,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        result=str(data.get("resultado_cierre") or "").strip();evidence=str(data.get("evidencia_cierre") or "").strip()
        if not result or not evidence:raise ValueError("El cierre exige resultado y evidencia.")
        now=now_iso()
        with self.connect() as conn:
            if not conn.execute("SELECT id FROM fcr_alertas WHERE fundacion_id=? AND id=?",(fundacion_id,alert_id)).fetchone():raise LookupError("Alerta no encontrada.")
            conn.execute("UPDATE fcr_alertas SET estado='CERRADA',fecha_cierre=?,cerrado_por=?,resultado_cierre=?,evidencia_cierre=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",(now[:10],user.get("id"),result[:8000],evidence[:2000],user.get("id"),now,fundacion_id,alert_id));self.audit(fundacion_id,user,"CERRAR_ALERTA_FAMILIAR","fcr_alertas",alert_id,{},conn);conn.commit()
        self._upsert_motor_alert(fundacion_id,alert_id,user);return self.alert(fundacion_id,alert_id)

    def _tenant_path(self,fundacion_id:int,*parts:str)->Path:
        path=tenant_storage_root(self.data_dir,fundacion_id)/"familias_redes"
        path=path.joinpath(*parts)
        path.mkdir(parents=True,exist_ok=True)
        return path

    def add_evidence(self,fundacion_id:int,file_obj:Any,data:dict[str,Any],user:dict[str,Any])->dict[str,Any]:
        if not file_obj or not getattr(file_obj,"filename",None):raise ValueError("Selecciona un archivo.")
        original=str(file_obj.filename);safe=secure_filename(original) or "evidencia";folder=self._tenant_path(fundacion_id,"evidencias");token=datetime.now().strftime("%Y%m%d_%H%M%S_%f");target=folder/f"{token}_{safe}";file_obj.save(target)
        size=target.stat().st_size
        if size<=0:target.unlink(missing_ok=True);raise ValueError("El archivo está vacío.")
        sha=file_sha256(target);now=now_iso();mime=getattr(file_obj,"mimetype",None) or mimetypes.guess_type(original)[0] or "application/octet-stream"
        with self.connect() as conn:
            cur=conn.execute("""INSERT INTO fcr_evidencias(fundacion_id,actividad_id,compromiso_id,alerta_id,expediente_familiar_id,tipo,titulo,nombre_original,nombre_guardado,ruta_archivo,mime_type,tamano_bytes,sha256,version,activo,cargado_por,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,1,1,?,?)""",(fundacion_id,data.get("actividad_id"),data.get("compromiso_id"),data.get("alerta_id"),data.get("expediente_familiar_id"),safe_state(data.get("tipo"),"SOPORTE"),str(data.get("titulo") or original)[:500],original[:500],target.name,str(target),mime,size,sha,user.get("id"),now));eid=int(cur.lastrowid);self.audit(fundacion_id,user,"CARGAR_EVIDENCIA_FAMILIAS","fcr_evidencias",eid,{"sha256":sha},conn);conn.commit();row=conn.execute("SELECT * FROM fcr_evidencias WHERE id=?",(eid,)).fetchone();return dict(row)

    def evidence_path(self,fundacion_id:int,evidence_id:int)->tuple[Path,str,str]|None:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_evidencias WHERE fundacion_id=? AND id=? AND activo=1",(fundacion_id,evidence_id)).fetchone()
            if not row:return None
            path=Path(row["ruta_archivo"]).resolve()
            try:path.relative_to(tenant_storage_root(self.data_dir,fundacion_id).resolve())
            except ValueError:return None
            if not path.is_file() or file_sha256(path)!=row["sha256"]:return None
            return path,row["nombre_original"],row["mime_type"] or "application/octet-stream"

    def evidence(self, fundacion_id: int, evidence_id: int) -> dict[str, Any]:
        """Retorna metadatos de una evidencia y la UCA de su entidad padre."""

        with self.connect() as conn:
            row = conn.execute(
                "SELECT * FROM fcr_evidencias WHERE fundacion_id=? AND id=? AND activo=1",
                (fundacion_id, evidence_id),
            ).fetchone()
            if not row:
                raise LookupError("Evidencia no encontrada.")
            item = dict(row)
            unit_name = None
            if item.get("actividad_id"):
                parent = conn.execute(
                    "SELECT unidad_nombre FROM fcr_actividades WHERE fundacion_id=? AND id=?",
                    (fundacion_id, item["actividad_id"]),
                ).fetchone()
                unit_name = parent[0] if parent else None
            elif item.get("compromiso_id"):
                parent = conn.execute(
                    "SELECT unidad_nombre FROM fcr_compromisos WHERE fundacion_id=? AND id=?",
                    (fundacion_id, item["compromiso_id"]),
                ).fetchone()
                unit_name = parent[0] if parent else None
            elif item.get("alerta_id"):
                parent = conn.execute(
                    "SELECT unidad_nombre FROM fcr_alertas WHERE fundacion_id=? AND id=?",
                    (fundacion_id, item["alerta_id"]),
                ).fetchone()
                unit_name = parent[0] if parent else None
            elif item.get("expediente_familiar_id"):
                parent = conn.execute(
                    "SELECT unidad_nombre FROM fcr_expedientes_familiares WHERE fundacion_id=? AND id=?",
                    (fundacion_id, item["expediente_familiar_id"]),
                ).fetchone()
                unit_name = parent[0] if parent else None
            item["unidad_nombre"] = unit_name
            return item

    def prepare_activity_documents(self,fundacion_id:int,activity_id:int,user:dict[str,Any],types:tuple[str,...]=( "ACTA","LISTADO_ASISTENCIA","INFORME"))->dict[str,Any]:
        activity=self.activity_detail(fundacion_id,activity_id);folder=self._tenant_path(fundacion_id,"documentos");token=datetime.now().strftime("%Y%m%d_%H%M%S");docs=[]
        for doc_type in types:
            kind=safe_state(doc_type)
            if kind=="LISTADO_ASISTENCIA":path=folder/f"listado_asistencia_actividad_{activity_id}_{token}.xlsx";self._write_attendance(path,activity);mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:path=folder/f"{kind.lower()}_actividad_{activity_id}_{token}.pdf";self._write_activity_pdf(path,activity,kind);mime="application/pdf"
            with self.connect() as conn:
                sha=file_sha256(path);now=now_iso();cur=conn.execute("""INSERT INTO fcr_documentos_generados(fundacion_id,actividad_id,tipo_documento,nombre_archivo,ruta_archivo,mime_type,tamano_bytes,sha256,estado,version_plantilla,generado_por,fecha_generacion) VALUES(?,?,?,?,?,?,?,?, 'BORRADOR','INTERNA-1',?,?)""",(fundacion_id,activity_id,kind,path.name,str(path),mime,path.stat().st_size,sha,user.get("id"),now));doc_id=int(cur.lastrowid);self.audit(fundacion_id,user,"GENERAR_DOCUMENTO_FAMILIAS","fcr_documentos_generados",doc_id,{"tipo":kind,"actividad_id":activity_id},conn);conn.commit();docs.append({"id":doc_id,"tipo_documento":kind,"nombre_archivo":path.name,"mime_type":mime,"tamano_bytes":path.stat().st_size,"sha256":sha,"estado":"BORRADOR","fecha_generacion":now})
        return {"documentos":docs}

    def _write_activity_pdf(self,path:Path,activity:dict[str,Any],kind:str)->None:
        styles=getSampleStyleSheet();doc=SimpleDocTemplate(str(path),pagesize=A4,rightMargin=1.7*cm,leftMargin=1.7*cm,topMargin=1.5*cm,bottomMargin=1.5*cm);title="ACTA DE ACTIVIDAD" if kind=="ACTA" else "BORRADOR DE INFORME DE ACTIVIDAD";story=[Paragraph(title,styles["Title"]),Paragraph("BORRADOR PARA REVISIÓN Y APROBACIÓN PROFESIONAL",styles["Heading2"]),Spacer(1,10)]
        rows=[["Campo","Información"],["Tipo",str(activity.get("tipo") or "")],["Título",str(activity.get("titulo") or "")],["UCA",str(activity.get("unidad_nombre") or "")],["Fecha programada",str(activity.get("fecha_programada") or "")],["Fecha de ejecución",str(activity.get("fecha_ejecucion") or "Pendiente")],["Lugar",str(activity.get("lugar") or "Pendiente")],["Profesional",str(activity.get("profesional_nombre") or "")],["Objetivo",str(activity.get("objetivo") or "Pendiente de diligenciar")],["Metodología",str(activity.get("metodologia") or "Pendiente de diligenciar")],["Resultados",str(activity.get("resultados") or "Pendiente de diligenciar por el profesional")],["Conclusiones profesionales",str(activity.get("conclusiones_profesionales") or "Pendiente de diligenciar por el profesional")],["Compromisos",str(activity.get("compromisos_generales") or "Pendiente de diligenciar")]]
        table=Table([[Paragraph(str(c),styles["BodyText"]) for c in row] for row in rows],colWidths=[4.5*cm,12.5*cm]);table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#155e75")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.35,colors.grey),("VALIGN",(0,0),(-1,-1),"TOP"),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]));story.append(table);story.append(Spacer(1,12));story.append(Paragraph("Este documento no inventa resultados ni reemplaza el análisis profesional. Debe revisarse, ajustarse y aprobarse antes de uso oficial.",styles["Italic"]));doc.build(story)

    def _write_attendance(self,path:Path,activity:dict[str,Any])->None:
        wb=Workbook();ws=wb.active;ws.title="Listado de asistencia";ws.append(["LISTADO DE ASISTENCIA - BORRADOR"]);ws.merge_cells("A1:H1");ws["A1"].font=Font(bold=True,size=14,color="FFFFFF");ws["A1"].fill=PatternFill("solid",fgColor="155E75");ws.append(["Actividad",activity.get("titulo"),"Tipo",activity.get("tipo"),"UCA",activity.get("unidad_nombre"),"Fecha",activity.get("fecha_programada")]);ws.append([]);ws.append(["N°","Nombre","Tipo asistente","Documento referencia","Teléfono","Asistió","Firma/referencia","Observaciones"])
        for index,item in enumerate(activity.get("asistencias") or [],1):ws.append([index,item.get("nombre_asistente"),item.get("tipo_asistente"),item.get("documento_referencia"),item.get("telefono"),"SÍ" if item.get("asistio") else "NO",item.get("firma_referencia"),item.get("observaciones")])
        for cell in ws[4]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="334155");cell.alignment=Alignment(wrap_text=True)
        widths=[6,30,20,20,18,10,25,35]
        for i,width in enumerate(widths,1):ws.column_dimensions[chr(64+i)].width=width
        ws.freeze_panes="A5";wb.save(path)

    def list_documents(self,fundacion_id:int,activity_id:int|None=None)->list[dict[str,Any]]:
        sql="""SELECT d.*, a.unidad_nombre, a.unidad_clave
                 FROM fcr_documentos_generados d
                 LEFT JOIN fcr_actividades a
                   ON a.fundacion_id=d.fundacion_id AND a.id=d.actividad_id
                 WHERE d.fundacion_id=?""";params:list[Any]=[fundacion_id]
        if activity_id:sql+=" AND actividad_id=?";params.append(activity_id)
        sql+=" ORDER BY d.fecha_generacion DESC,d.id DESC LIMIT 1000"
        with self.connect() as conn:return [dict(x) for x in conn.execute(sql,params).fetchall()]

    def document(self, fundacion_id: int, document_id: int) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute(
                """SELECT d.*, a.unidad_nombre, a.unidad_clave
                     FROM fcr_documentos_generados d
                     LEFT JOIN fcr_actividades a
                       ON a.fundacion_id=d.fundacion_id AND a.id=d.actividad_id
                     WHERE d.fundacion_id=? AND d.id=?""",
                (fundacion_id, document_id),
            ).fetchone()
            if not row:
                raise LookupError("Documento no encontrado.")
            return dict(row)

    def document_path(self,fundacion_id:int,document_id:int)->tuple[Path,str,str]|None:
        with self.connect() as conn:
            row=conn.execute("SELECT * FROM fcr_documentos_generados WHERE fundacion_id=? AND id=?",(fundacion_id,document_id)).fetchone()
            if not row:return None
            path=Path(row["ruta_archivo"]).resolve()
            try:path.relative_to(tenant_storage_root(self.data_dir,fundacion_id).resolve())
            except ValueError:return None
            if not path.is_file() or file_sha256(path)!=row["sha256"]:return None
            return path,row["nombre_archivo"],row["mime_type"] or "application/octet-stream"

    def review_document(self,fundacion_id:int,document_id:int,action:str,user:dict[str,Any])->dict[str,Any]:
        action=safe_state(action);mapping={"REVISAR":"PENDIENTE_APROBACION","APROBAR":"APROBADO","DEVOLVER":"DEVUELTO"}
        if action not in mapping:raise ValueError("Acción inválida.")
        now=now_iso();sets="estado=?";values:list[Any]=[mapping[action]]
        if action=="REVISAR":sets+=",revisado_por=?,fecha_revision=?";values.extend([user.get("id"),now])
        if action=="APROBAR":sets+=",aprobado_por=?,fecha_aprobacion=?";values.extend([user.get("id"),now])
        values.extend([fundacion_id,document_id])
        with self.connect() as conn:
            conn.execute(f"UPDATE fcr_documentos_generados SET {sets} WHERE fundacion_id=? AND id=?",values);row=conn.execute("SELECT * FROM fcr_documentos_generados WHERE fundacion_id=? AND id=?",(fundacion_id,document_id)).fetchone()
            if not row:raise LookupError("Documento no encontrado.")
            self.audit(fundacion_id,user,f"{action}_DOCUMENTO_FAMILIAS","fcr_documentos_generados",document_id,{},conn);conn.commit();return dict(row)

    def dashboard(self,fundacion_id:int,user:dict[str,Any],filters:dict[str,Any]|None=None)->dict[str,Any]:
        filters=filters or {};professional_only=user.get("rol")=='PSICOSOCIAL' and not filters.get("vista_coordinacion");activity_filters={"unidad":filters.get("unidad")}
        if professional_only:activity_filters["profesional_id"]=user.get("id")
        activities=self.list_activities(fundacion_id,activity_filters);families=self.list_family_records(fundacion_id,{"unidad":filters.get("unidad")},limit=5000);commitments=self.list_commitments(fundacion_id,{"unidad":filters.get("unidad")});alerts=self.list_alerts(fundacion_id,{"unidad":filters.get("unidad")});networks=self.list_networks(fundacion_id);today=date.today().isoformat();open_commit=[x for x in commitments if safe_state(x.get("estado")) not in COMPLETED_STATES];overdue=[x for x in open_commit if x.get("fecha_limite") and str(x["fecha_limite"])<today];open_alerts=[x for x in alerts if safe_state(x.get("estado")) not in {"CERRADA","CERRADO"}]
        by_unit=defaultdict(lambda:{"familias":0,"actividades":0,"compromisos_abiertos":0,"alertas_abiertas":0})
        for f in families:by_unit[f.get("unidad_nombre") or "Sin UCA"]["familias"]+=1
        for a in activities:by_unit[a.get("unidad_nombre") or "Sin UCA"]["actividades"]+=1
        for c in open_commit:by_unit[c.get("unidad_nombre") or "Sin UCA"]["compromisos_abiertos"]+=1
        for a in open_alerts:by_unit[a.get("unidad_nombre") or "Sin UCA"]["alertas_abiertas"]+=1
        units=[{"unidad":name,**item} for name,item in sorted(by_unit.items())]
        return {"vista":"PROFESIONAL" if professional_only else "COORDINACION","resumen":{"expedientes_familiares":len(families),"actividades":len(activities),"actividades_pendientes_cierre":sum(1 for x in activities if safe_state(x.get("estado")) not in COMPLETED_STATES),"compromisos_abiertos":len(open_commit),"compromisos_vencidos":len(overdue),"alertas_abiertas":len(open_alerts),"redes_activas":len(networks),"documentos_borrador":sum(1 for x in self.list_documents(fundacion_id) if safe_state(x.get("estado"))=="BORRADOR")},"por_uca":units,"actividades":activities[:100],"compromisos":commitments[:100],"alertas":alerts[:100],"redes":networks[:100]}

    def prepare_summary_package(self,fundacion_id:int,user:dict[str,Any],filters:dict[str,Any]|None=None)->dict[str,Any]:
        data=self.dashboard(fundacion_id,user,filters);folder=self._tenant_path(fundacion_id,"reportes");token=datetime.now().strftime("%Y%m%d_%H%M%S");zip_path=folder/f"familias_redes_{token}.zip";xlsx=folder/f"familias_redes_{token}.xlsx";self._write_summary_xlsx(xlsx,data)
        with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as archive:
            archive.write(xlsx,xlsx.name);archive.writestr("00_RESUMEN.json",json.dumps(data,ensure_ascii=False,indent=2,default=str));archive.writestr("LEEME.txt","Paquete de gestión de familias, comunidad y redes. Los documentos son borradores y requieren revisión humana.\n")
        return {"ruta":str(zip_path),"nombre_archivo":zip_path.name,"sha256":file_sha256(zip_path),"tamano_bytes":zip_path.stat().st_size}

    def _write_summary_xlsx(self,path:Path,data:dict[str,Any])->None:
        wb=Workbook();ws=wb.active;ws.title="Resumen";ws.append(["GESTIÓN INTEGRAL DE FAMILIAS, COMUNIDAD Y REDES"]);ws.merge_cells("A1:D1");ws["A1"].font=Font(bold=True,size=14,color="FFFFFF");ws["A1"].fill=PatternFill("solid",fgColor="0F766E");ws.append(["Producto","BORRADOR PARA REVISIÓN HUMANA"])
        for key,val in data.get("resumen",{}).items():ws.append([key.replace("_"," ").title(),val])
        wa=wb.create_sheet("Actividades");wa.append(["ID","Tipo","Título","UCA","Fecha","Estado","Profesional","Asistentes"])
        for item in data.get("actividades",[]):wa.append([item.get("id"),item.get("tipo"),item.get("titulo"),item.get("unidad_nombre"),item.get("fecha_programada"),item.get("estado"),item.get("profesional_nombre"),item.get("asistencias")])
        wc=wb.create_sheet("Compromisos");wc.append(["ID","Título","UCA","Responsable","Fecha límite","Estado","Avance"])
        for item in data.get("compromisos",[]):wc.append([item.get("id"),item.get("titulo"),item.get("unidad_nombre"),item.get("responsable_nombre"),item.get("fecha_limite"),item.get("estado"),item.get("porcentaje")])
        we=wb.create_sheet("Alertas");we.append(["ID","Tipo","Nivel","UCA","Descripción","Estado","Próximo seguimiento"])
        for item in data.get("alertas",[]):we.append([item.get("id"),item.get("tipo"),item.get("nivel"),item.get("unidad_nombre"),item.get("descripcion"),item.get("estado"),item.get("fecha_proximo_seguimiento")])
        for sheet in wb.worksheets:
            for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="334155")
            sheet.freeze_panes="A2"
        wb.save(path)

    def package_path(self,fundacion_id:int,path_value:str)->Path|None:
        path=Path(path_value).resolve()
        try:path.relative_to(tenant_storage_root(self.data_dir,fundacion_id).resolve())
        except ValueError:return None
        return path if path.is_file() else None

    def _upsert_motor(self, fundacion_id: int, source_table: str, source_id: int, item: dict[str, Any], user: dict[str, Any]) -> None:
        with self.connect() as conn:
            if not self._table_exists(conn, "mgp_tareas"):
                return
            key = mgp_source_key(source_table, source_id)
            now = now_iso()
            unit_name = str(item.get("unidad_nombre") or "").strip() or None
            priority = safe_state(item.get("prioridad"), "MEDIA")
            task_state = safe_state(item.get("estado"), "PENDIENTE")
            score = 90 if priority in {"CRITICO", "CRITICA"} else (70 if priority in {"ALTO", "ALTA"} else 40)
            values = (
                fundacion_id, item.get("expediente_uca_id"), unit_name, mgp_unit_key(unit_name),
                "FAMILIAS_REDES", source_table, source_id, key,
                item.get("tipo_tarea") or "ACTIVIDAD_FAMILIAS", "FAMILIA_COMUNIDAD_REDES",
                item.get("titulo") or "Actividad familias", item.get("descripcion"),
                item.get("fecha_inicio"), item.get("fecha_limite"), item.get("fecha_finalizacion"),
                task_state, priority, score, item.get("responsable_id"), item.get("responsable_nombre"),
                0, 0, 0, json_dump({"origen": "familias_redes", "source_id": source_id}),
                1, user.get("id"), user.get("id"), now, now,
            )
            legacy_key = f"{source_table}:{source_id}"
            if legacy_key != key:
                conn.execute("DELETE FROM mgp_tareas WHERE fundacion_id=? AND fuente_tabla=? AND fuente_clave=?", (fundacion_id, source_table, legacy_key))
            conn.execute(
                """
                INSERT INTO mgp_tareas
                (fundacion_id,expediente_id,unidad_nombre,unidad_clave,fuente_modulo,fuente_tabla,fuente_id,fuente_clave,
                 tipo_tarea,componente,titulo,descripcion,fecha_inicio,fecha_limite,fecha_finalizacion,estado,prioridad,
                 puntaje_prioridad,responsable_id,responsable_nombre,requiere_evidencia,evidencias_total,bloqueada,
                 metadata_json,activa,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(fundacion_id,fuente_tabla,fuente_clave) DO UPDATE SET
                  expediente_id=excluded.expediente_id,unidad_nombre=excluded.unidad_nombre,unidad_clave=excluded.unidad_clave,
                  fuente_modulo=excluded.fuente_modulo,fuente_id=excluded.fuente_id,tipo_tarea=excluded.tipo_tarea,
                  componente=excluded.componente,titulo=excluded.titulo,descripcion=excluded.descripcion,
                  fecha_inicio=excluded.fecha_inicio,fecha_limite=excluded.fecha_limite,fecha_finalizacion=excluded.fecha_finalizacion,
                  estado=excluded.estado,prioridad=excluded.prioridad,puntaje_prioridad=excluded.puntaje_prioridad,
                  responsable_id=excluded.responsable_id,responsable_nombre=excluded.responsable_nombre,
                  metadata_json=excluded.metadata_json,activa=1,actualizada_por=excluded.actualizada_por,
                  fecha_actualizacion=excluded.fecha_actualizacion
                """,
                values,
            )
            conn.commit()

    def _upsert_motor_activity(self,fundacion_id:int,activity_id:int,user:dict[str,Any])->None:
        with self.connect() as conn:row=conn.execute("SELECT * FROM fcr_actividades WHERE fundacion_id=? AND id=?",(fundacion_id,activity_id)).fetchone()
        if row:self._upsert_motor(fundacion_id,"fcr_actividades",activity_id,{"expediente_uca_id":row["expediente_uca_id"],"unidad_nombre":row["unidad_nombre"],"tipo_tarea":"ACTIVIDAD_FAMILIAS","titulo":f"{str(row['tipo']).replace('_',' ').title()}: {row['titulo']}","descripcion":row["objetivo"],"fecha_inicio":row["fecha_programada"],"fecha_limite":row["fecha_limite_cierre"] or row["fecha_programada"],"fecha_finalizacion":row["fecha_ejecucion"],"estado":row["estado"],"prioridad":"MEDIA","responsable_id":row["profesional_id"],"responsable_nombre":row["profesional_nombre"]},user)

    def _upsert_motor_commitment(self,fundacion_id:int,commitment_id:int,user:dict[str,Any])->None:
        current=self.commitment_detail(fundacion_id,commitment_id);self._upsert_motor(fundacion_id,"fcr_compromisos",commitment_id,{"expediente_uca_id":current.get("expediente_uca_id"),"unidad_nombre":current.get("unidad_nombre"),"tipo_tarea":"COMPROMISO_FAMILIAR","titulo":current.get("titulo"),"descripcion":current.get("descripcion"),"fecha_inicio":current.get("fecha_compromiso"),"fecha_limite":current.get("fecha_limite"),"fecha_finalizacion":current.get("fecha_cierre"),"estado":current.get("estado"),"prioridad":current.get("prioridad"),"responsable_id":current.get("responsable_id"),"responsable_nombre":current.get("responsable_nombre")},user)

    def _upsert_motor_alert(self,fundacion_id:int,alert_id:int,user:dict[str,Any])->None:
        current=self.alert(fundacion_id,alert_id);self._upsert_motor(fundacion_id,"fcr_alertas",alert_id,{"expediente_uca_id":current.get("expediente_uca_id"),"unidad_nombre":current.get("unidad_nombre"),"tipo_tarea":"ALERTA_FAMILIAR","titulo":f"Alerta {current.get('tipo')}","descripcion":current.get("descripcion"),"fecha_inicio":current.get("fecha_identificacion"),"fecha_limite":current.get("fecha_proximo_seguimiento"),"fecha_finalizacion":current.get("fecha_cierre"),"estado":current.get("estado"),"prioridad":"CRITICA" if current.get("nivel")=="CRITICO" else current.get("nivel"),"responsable_id":current.get("responsable_id"),"responsable_nombre":current.get("responsable_nombre")},user)
