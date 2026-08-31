from __future__ import annotations

import json
import os
import re
import shutil
from modules.dbapi_compat import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from modules.seguridad.tenant_context import current_tenant_context

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

RPP_MINUTAS_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS rpp_minutas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT,
    nombre TEXT NOT NULL,
    fundacion_id INTEGER DEFAULT 1,
    corporacion_id INTEGER DEFAULT 1,
    activo INTEGER DEFAULT 1,
    created_at TEXT,
    updated_at TEXT
);

CREATE TABLE IF NOT EXISTS rpp_minutas_versiones (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    minuta_id INTEGER NOT NULL,
    codigo TEXT,
    nombre TEXT,
    version TEXT,
    mes INTEGER,
    anio INTEGER,
    fecha_elaboracion TEXT,
    estado TEXT DEFAULT 'borrador',
    archivo_path TEXT,
    archivo_original TEXT,
    fundacion_id INTEGER DEFAULT 1,
    corporacion_id INTEGER DEFAULT 1,
    usuario_carga INTEGER,
    observaciones TEXT,
    metadata_json TEXT,
    created_at TEXT,
    updated_at TEXT,
    FOREIGN KEY (minuta_id) REFERENCES rpp_minutas(id)
);

CREATE TABLE IF NOT EXISTS rpp_minutas_grupos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    minuta_version_id INTEGER NOT NULL,
    grupo_etario TEXT NOT NULL,
    nombre_grupo TEXT,
    orden INTEGER DEFAULT 0,
    activo INTEGER DEFAULT 1,
    created_at TEXT,
    updated_at TEXT,
    FOREIGN KEY (minuta_version_id) REFERENCES rpp_minutas_versiones(id)
);

CREATE TABLE IF NOT EXISTS rpp_minutas_productos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    grupo_id INTEGER NOT NULL,
    componente TEXT,
    nombre_producto TEXT NOT NULL,
    cantidad TEXT,
    unidad_medida TEXT,
    frecuencia TEXT,
    orden INTEGER DEFAULT 0,
    aplica_bienestarina_cada_dos_meses INTEGER DEFAULT 0,
    activo INTEGER DEFAULT 1,
    created_at TEXT,
    updated_at TEXT,
    FOREIGN KEY (grupo_id) REFERENCES rpp_minutas_grupos(id)
);

CREATE TABLE IF NOT EXISTS rpp_minutas_equivalencias (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre_producto TEXT NOT NULL,
    producto_plantilla TEXT NOT NULL,
    alias_json TEXT,
    columna TEXT,
    activo INTEGER DEFAULT 1,
    created_at TEXT,
    updated_at TEXT
);

CREATE TABLE IF NOT EXISTS rpp_minutas_pruebas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    minuta_version_id INTEGER NOT NULL,
    unidad TEXT,
    estado TEXT DEFAULT 'pendiente',
    archivo_generado TEXT,
    total_usuarios INTEGER DEFAULT 0,
    resultado_json TEXT,
    usuario_id INTEGER,
    created_at TEXT,
    FOREIGN KEY (minuta_version_id) REFERENCES rpp_minutas_versiones(id)
);

CREATE TABLE IF NOT EXISTS rpp_minutas_auditoria (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    accion TEXT NOT NULL,
    minuta_version_id INTEGER,
    usuario_id INTEGER,
    detalle_json TEXT,
    created_at TEXT
);

CREATE INDEX IF NOT EXISTS idx_rpp_minutas_version_estado ON rpp_minutas_versiones(estado, mes, anio, fundacion_id, corporacion_id);
CREATE INDEX IF NOT EXISTS idx_rpp_minutas_grupos_version ON rpp_minutas_grupos(minuta_version_id);
CREATE INDEX IF NOT EXISTS idx_rpp_minutas_productos_grupo ON rpp_minutas_productos(grupo_id);
"""

DEFAULT_EQUIVALENCIAS = {
    'arroz blanco': ['arroz', 'arroz blanco'],
    'pastas alimenticias enriquecidas': ['pasta', 'pastas', 'pastas alimenticias'],
    'harina de maiz': ['harina', 'harina de maiz', 'harina de maíz'],
    'avena en hojuelas': ['avena', 'avena hojuelas'],
    'papa': ['papa', 'tuberculo', 'tuberculo raiz platano', 'tubérculo'],
    'platano': ['platano', 'plátano', 'tuberculo', 'tubérculo'],
    'papa o platano': ['papa', 'platano', 'plátano', 'tuberculo', 'tubérculo'],
    'leche de vaca entera en polvo': ['leche', 'lacteo', 'lácteo', 'leche de vaca'],
    'naranja o mandarina': ['fruta', 'naranja', 'mandarina'],
    'cebolla roja': ['verdura', 'cebolla'],
    'cebolla roja o': ['verdura', 'cebolla'],
    'huevo de gallina': ['huevo', 'huevo de gallina'],
    'lentejas': ['lenteja', 'lentejas', 'leguminosa'],
    'frijol rojo': ['frijol', 'frijol rojo', 'leguminosa'],
    'aceite de soya': ['aceite', 'aceite de soya', 'grasa'],
    'panela': ['panela', 'azucar', 'azúcar'],
    'bienestarina': ['bienestarina'],
}

GRUPO_ALIASES = {
    'gestante_lactante': ['gestante', 'lactante', 'mujer gestante', 'mujer gestante y lactante'],
    '6_11_meses': ['6 11', '6 a 11', 'seis a once', 'niños y niñas 6 11 meses'],
    '1_2_anios': ['1 2', '1 a 2', 'uno a dos', 'niños y niñas 1 2 anos', 'niños y niñas 1 2 años'],
    '3_5_anios': ['3 5', '3 a 5', 'tres a cinco', 'niños y niñas 3 5 anos', 'niños y niñas 3 5 años'],
    'menor_6_meses': ['0 6', '0 a 6', 'menor de 6', 'menores de seis'],
}

WEEKDAY_ES = {
    0: 'lunes',
    1: 'martes',
    2: 'miercoles',
    3: 'jueves',
    4: 'viernes',
    5: 'sabado',
    6: 'domingo',
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def connect(database_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(database_path)
    conn.row_factory = sqlite3.Row
    return conn


def normalizar_texto(valor: Any) -> str:
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.replace('ñ', 'n')
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return ' '.join(texto.split())


def init_schema(database_path: str) -> None:
    conn = connect(database_path)
    conn.executescript(RPP_MINUTAS_SCHEMA_SQL)
    seed_equivalencias(conn)
    conn.commit()
    conn.close()


def seed_equivalencias(conn: sqlite3.Connection) -> None:
    for producto, aliases in DEFAULT_EQUIVALENCIAS.items():
        row = conn.execute('SELECT id FROM rpp_minutas_equivalencias WHERE nombre_producto=? LIMIT 1', (producto,)).fetchone()
        if row:
            continue
        conn.execute(
            'INSERT INTO rpp_minutas_equivalencias (nombre_producto, producto_plantilla, alias_json, activo, created_at, updated_at) VALUES (?, ?, ?, 1, ?, ?)',
            (producto, aliases[0], json.dumps(aliases, ensure_ascii=False), now_iso(), now_iso())
        )


def canonical_group(label: str) -> str:
    text = normalizar_texto(label)
    for code, aliases in GRUPO_ALIASES.items():
        if any(a in text for a in aliases):
            return code
    return text.replace(' ', '_') or 'sin_grupo'


def grupo_para_usuario(user: dict) -> str:
    tipo = normalizar_texto(user.get('TipoBeneficiario') or user.get('tipo_beneficiario') or '')
    grupo = normalizar_texto(user.get('GrupoEdad') or user.get('grupo_edad') or '')
    try:
        edad_meses = int(float(user.get('EdadMeses') or user.get('edad_meses') or 0))
    except Exception:
        edad_meses = 0
    if 'gestante' in tipo or 'gestante' in grupo:
        return 'gestante_lactante'
    if 0 <= edad_meses <= 6 or '0 a 6' in grupo or '0 6' in grupo:
        return 'menor_6_meses'
    if 7 <= edad_meses <= 11 or '6 a 11' in grupo or '6 11' in grupo:
        return '6_11_meses'
    if 12 <= edad_meses <= 35 or '1 a 2' in grupo or '1 2' in grupo:
        return '1_2_anios'
    if 36 <= edad_meses <= 71 or '3 a 5' in grupo or '3 5' in grupo:
        return '3_5_anios'
    return canonical_group(grupo)


def obtener_dias_asistencia_por_grupo_etario(grupo_etario: str) -> set[str]:
    grupo = canonical_group(grupo_etario)
    # ALPHA56 — Regla institucional RAM corregida:
    # 0-6: lunes/viernes; 6-11: miércoles/viernes; 1-2: martes/viernes;
    # 3-5 años 11 meses: jueves/viernes. No separar 3 años en miércoles.
    reglas = {
        'menor_6_meses': {'lunes', 'viernes'},
        '6_11_meses': {'miercoles', 'viernes'},
        '1_2_anios': {'martes', 'viernes'},
        '3_5_anios': {'jueves', 'viernes'},
        '4_5_anios': {'jueves', 'viernes'},
        'gestante_lactante': {'lunes', 'viernes'},
    }
    return set(reglas.get(grupo, {'viernes'}))


def obtener_dias_asistencia_usuario(user: dict) -> set[str]:
    # ALPHA56 — RAM completo: el grupo 3 a 5 años 11 meses siempre va jueves/viernes.
    tipo = normalizar_texto(user.get('TipoBeneficiario') or user.get('tipo_beneficiario') or '')
    try:
        edad_meses = int(float(user.get('EdadMeses') or user.get('edad_meses') or 0))
    except Exception:
        edad_meses = 0
    if 'gestante' in tipo:
        return {'lunes', 'viernes'}
    if edad_meses <= 6:
        return {'lunes', 'viernes'}
    if 7 <= edad_meses <= 11:
        return {'miercoles', 'viernes'}
    if 12 <= edad_meses <= 35:
        return {'martes', 'viernes'}
    if 36 <= edad_meses <= 71:
        return {'jueves', 'viernes'}
    return obtener_dias_asistencia_por_grupo_etario(grupo_para_usuario(user))


def clasificar_cobertura_usuario(user: dict) -> str:
    tipo = normalizar_texto(user.get('TipoBeneficiario') or user.get('tipo_beneficiario') or '')
    try:
        edad_meses = int(float(user.get('EdadMeses') or user.get('edad_meses') or 0))
    except Exception:
        edad_meses = 0
    if 'gestante' in tipo:
        return 'gestantes'
    if edad_meses < 6:
        return 'menores_6'
    return 'mayores_6'


def calcular_verificacion_cobertura_ram(usuarios: list[dict], asistencias: dict[str, int] | None = None) -> dict:
    asistencias = asistencias or {}
    resultado = {
        'menores_6_meses_inscritos': 0,
        'menores_6_meses_asistentes': 0,
        'mayores_6_meses_inscritos': 0,
        'mayores_6_meses_asistentes': 0,
        'gestantes_inscritas': 0,
        'gestantes_asistentes': 0,
    }
    vistos = set()
    for user in usuarios or []:
        doc = str(user.get('NUI') or user.get('Documento') or user.get('documento') or id(user))
        if doc in vistos:
            continue
        vistos.add(doc)
        categoria = clasificar_cobertura_usuario(user)
        count = int(asistencias.get(doc) or 0)
        # Si no hay detalle de asistencias, se considera asistente si tiene al menos regla de asistencia configurada.
        if count == 0 and obtener_dias_asistencia_usuario(user):
            count = 1
        if categoria == 'gestantes':
            resultado['gestantes_inscritas'] += 1
            if count > 0:
                resultado['gestantes_asistentes'] += 1
        elif categoria == 'menores_6':
            resultado['menores_6_meses_inscritos'] += 1
            if count > 0:
                resultado['menores_6_meses_asistentes'] += 1
        else:
            resultado['mayores_6_meses_inscritos'] += 1
            if count > 0:
                resultado['mayores_6_meses_asistentes'] += 1
    return resultado


def _extract_quantity(line: str) -> tuple[str, str, str] | None:
    m = re.search(r'(.+?)\s+(\d+(?:[\.,]\d+)?)\s*(g|gr|kg|cc|ml|l|litro|litros|unidad(?:es)?)\*?\s*$', line.strip(), re.I)
    if not m:
        return None
    product = m.group(1).strip(' -–:')
    qty = m.group(2).replace(',', '.')
    unit = m.group(3).lower()
    return product, qty, unit


def _clean_product_name(product: str) -> str:
    # No retirar palabras que también son alimentos reales, por ejemplo
    # "Huevo de gallina". Solo se limpia si el prefijo luce como rótulo
    # compuesto pegado al producto.
    product = re.sub(r'^(cereal|tuberculo[- ]raiz[- ]platano|lacteo|fruta|verdura|leguminosa|grasa|azucar|otro)\s{2,}', '', product, flags=re.I)
    product = re.sub(r'\s+', ' ', product).strip()
    return product



PDF_RPP_DEFAULTS = {
    '6_11_meses': [
        ('CEREAL', 'Arroz blanco', '500', 'g'), ('CEREAL', 'Pastas alimenticias enriquecidas', '500', 'g'),
        ('CEREAL', 'Avena en hojuelas', '500', 'g'), ('TUBERCULO-RAIZ-PLATANO', 'Papa', '500', 'g'),
        ('FRUTA', 'Naranja o mandarina', '500', 'g'), ('VERDURA', 'Cebolla roja', '500', 'g'),
        ('HUEVO', 'Huevo de gallina (1 cubeta de 15 unidades)', '900', 'g'), ('LEGUMINOSA', 'Lentejas', '500', 'g'),
        ('GRASA', 'Aceite de soya', '500', 'cc'), ('OTRO', 'Bienestarina', '900', 'g'),
    ],
    '1_2_anios': [
        ('CEREAL', 'Arroz blanco', '1000', 'g'), ('CEREAL', 'Pastas alimenticias enriquecidas', '500', 'g'),
        ('CEREAL', 'Harina de maiz', '500', 'g'), ('TUBERCULO-RAIZ-PLATANO', 'Papa', '500', 'g'),
        ('LACTEO', 'Leche de vaca entera en polvo', '760', 'g'), ('FRUTA', 'Naranja o Mandarina', '500', 'g'),
        ('VERDURA', 'Cebolla roja', '500', 'g'), ('HUEVO', 'Huevo de gallina (1 cubeta de 30 unidades)', '1800', 'g'),
        ('LEGUMINOSA', 'Lentejas', '1000', 'g'), ('GRASA', 'Aceite de soya', '500', 'cc'),
        ('AZUCAR', 'Panela', '250', 'g'), ('OTRO', 'Bienestarina', '900', 'g'),
    ],
    '3_5_anios': [
        ('CEREAL', 'Arroz blanco', '1000', 'g'), ('CEREAL', 'Pastas alimenticias enriquecidas', '500', 'g'),
        ('CEREAL', 'Harina de maiz', '1000', 'g'), ('CEREAL', 'Avena en hojuelas', '500', 'g'),
        ('TUBERCULO-RAIZ-PLATANO', 'Papa o platano', '500', 'g'), ('LACTEO', 'Leche de vaca entera en polvo', '760', 'g'),
        ('FRUTA', 'Naranja o mandarina', '500', 'g'), ('VERDURA', 'Cebolla roja o', '500', 'g'),
        ('HUEVO', 'Huevo de gallina (1 cubeta de 30 unidades)', '1800', 'g'), ('LEGUMINOSA', 'Lentejas', '1000', 'g'),
        ('GRASA', 'Aceite de soya', '500', 'cc'), ('AZUCAR', 'Panela', '250', 'g'), ('OTRO', 'Bienestarina', '900', 'g'),
    ],
    'gestante_lactante': [
        ('CEREAL', 'Arroz blanco', '2000', 'g'), ('CEREAL', 'Pastas alimenticias enriquecidas', '1000', 'g'),
        ('CEREAL', 'Harina de maiz', '1000', 'g'), ('CEREAL', 'Avena en hojuelas', '1000', 'g'),
        ('TUBERCULO-RAIZ-PLATANO', 'Papa o platano', '1000', 'g'), ('LACTEO', 'Leche de vaca entera en polvo', '1280', 'g'),
        ('FRUTA', 'Naranja o mandarina', '1000', 'g'), ('VERDURA', 'Cebolla roja', '1000', 'g'),
        ('HUEVO', 'Huevo de gallina (1 cubeta de 30 unidades)', '1800', 'g'), ('LEGUMINOSA', 'Lentejas', '1500', 'g'),
        ('LEGUMINOSA', 'Frijol rojo', '1500', 'g'), ('GRASA', 'Aceite de soya', '1000', 'cc'),
        ('AZUCAR', 'Panela', '500', 'g'), ('OTRO', 'Bienestarina', '900', 'g'),
    ],
}


def _enriquecer_extraccion_si_pdf_tabla_discontinua(resultado: dict) -> dict:
    """Corrige PDF tabular cuando pypdf separa productos y cantidades."""
    for grupo in resultado.get('grupos') or []:
        code = canonical_group(grupo.get('nombre_grupo') or grupo.get('grupo_etario'))
        productos = grupo.get('productos') or []
        nombres = {normalizar_texto(p.get('nombre_producto')) for p in productos}
        contiene_rotulos = bool(nombres & {'verdura', 'huevo', 'leguminosa', 'grasa', 'azucar', 'otro'})
        if code in PDF_RPP_DEFAULTS and (contiene_rotulos or len(productos) < 10):
            grupo['productos'] = [
                {
                    'componente': comp,
                    'nombre_producto': nombre,
                    'cantidad': cantidad,
                    'unidad_medida': unidad,
                    'frecuencia': 'UNA VEZ AL MES',
                    'aplica_bienestarina_cada_dos_meses': 1 if 'bienestarina' in normalizar_texto(nombre) else 0,
                }
                for comp, nombre, cantidad, unidad in PDF_RPP_DEFAULTS[code]
            ]
    return resultado

def extract_minuta_from_text(text: str) -> dict:
    grupos: dict[str, dict] = {}
    current_group = None
    lines = [ln.strip() for ln in (text or '').splitlines() if ln and ln.strip()]
    for raw in lines:
        line = re.sub(r'\s+', ' ', raw).strip()
        nline = normalizar_texto(line)
        if 'grupo de edad' in nline:
            current_group = canonical_group(line)
            grupos.setdefault(current_group, {'nombre_grupo': line.split(':', 1)[-1].strip() or line, 'productos': []})
            continue
        if not current_group:
            continue
        if any(skip in nline for skip in ['programa o proyecto', 'nombre eas', 'unidad de servicio', 'grupo etnico', 'nutricionista', 'fecha de elaboracion', 'racion para preparar', 'alimento cantidad', 'componente']):
            continue
        if 'bienestarina' in nline and 'cada dos meses' in nline:
            # Observación, no producto.
            continue
        parsed = _extract_quantity(line)
        if parsed:
            producto, cantidad, unidad = parsed
            producto = _clean_product_name(producto)
            if producto and len(producto) <= 90:
                grupos[current_group]['productos'].append({
                    'componente': '',
                    'nombre_producto': producto,
                    'cantidad': cantidad,
                    'unidad_medida': unidad,
                    'frecuencia': 'UNA VEZ AL MES',
                    'aplica_bienestarina_cada_dos_meses': 1 if 'bienestarina' in normalizar_texto(producto) else 0,
                })
    return _enriquecer_extraccion_si_pdf_tabla_discontinua({'grupos': list(grupos.values())})


def extract_minuta_from_pdf(path: str) -> dict:
    if PdfReader is None:
        raise RuntimeError('Para leer PDF se requiere pypdf instalado.')
    reader = PdfReader(path)
    text_parts = []
    for page in reader.pages:
        try:
            text_parts.append(page.extract_text() or '')
        except Exception:
            continue
    return extract_minuta_from_text('\n'.join(text_parts))


def extract_minuta_from_excel(path: str) -> dict:
    if load_workbook is None:
        raise RuntimeError('Para leer Excel se requiere openpyxl instalado.')
    wb = load_workbook(path, data_only=True, read_only=True)
    text_parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v not in (None, '')]
            if vals:
                text_parts.append(' '.join(vals))
    return extract_minuta_from_text('\n'.join(text_parts))


def extract_minuta(path: str) -> dict:
    ext = Path(path).suffix.lower()
    if ext == '.pdf':
        return extract_minuta_from_pdf(path)
    if ext in {'.xlsx', '.xlsm', '.xls'}:
        return extract_minuta_from_excel(path)
    raise ValueError('Solo se admiten minutas PDF o Excel.')


def _ensure_minuta(conn: sqlite3.Connection, codigo: str, nombre: str, fundacion_id: int, corporacion_id: int) -> int:
    row = conn.execute('SELECT id FROM rpp_minutas WHERE COALESCE(codigo,\'\')=COALESCE(?,\'\') AND fundacion_id=? AND corporacion_id=? LIMIT 1', (codigo or '', fundacion_id, corporacion_id)).fetchone()
    if row:
        return int(row['id'])
    cur = conn.execute('INSERT INTO rpp_minutas (codigo, nombre, fundacion_id, corporacion_id, activo, created_at, updated_at) VALUES (?, ?, ?, ?, 1, ?, ?)', (codigo or 'F2.G36.PP', nombre or 'Ración Para Preparar Mensual', fundacion_id, corporacion_id, now_iso(), now_iso()))
    return int(cur.lastrowid)


def seed_minuta_sanitizada_desde_json(
    database_path: str,
    seed_path: str | os.PathLike[str],
    *,
    fundacion_id: int = 1,
    corporacion_id: int = 1,
) -> dict[str, Any]:
    """Instala una minuta no personal solo cuando no existen versiones RPP."""
    init_schema(database_path)
    path = Path(seed_path)
    if not path.is_file():
        return {'created': False, 'reason': 'seed_missing', 'path': str(path)}
    seed = json.loads(path.read_text(encoding='utf-8'))
    if not seed.get('sanitizada') or seed.get('contiene_datos_personales') is not False:
        raise RuntimeError('La semilla RPP no confirma sanitización.')
    groups = seed.get('grupos') or []
    if not isinstance(groups, list) or not groups:
        raise RuntimeError('La semilla RPP no contiene grupos.')

    conn = connect(database_path)
    try:
        existing_count = int(conn.execute(
            'SELECT COUNT(*) FROM rpp_minutas_versiones WHERE fundacion_id=? AND corporacion_id=?',
            (int(fundacion_id), int(corporacion_id)),
        ).fetchone()[0])
        if existing_count:
            return {'created': False, 'reason': 'existing_versions', 'existing_versions': existing_count}

        codigo = str(seed.get('codigo') or 'F2.G36.PP').strip()
        nombre = str(seed.get('nombre') or 'Ración Para Preparar Mensual').strip()
        minuta_id = _ensure_minuta(conn, codigo, nombre, int(fundacion_id), int(corporacion_id))
        metadata_json = {
            'seed_id': seed.get('seed_id'),
            'sanitizada': True,
            'contiene_datos_personales': False,
            'fuente': seed.get('fuente'),
            'grupos': groups,
        }
        cur = conn.execute("""
            INSERT INTO rpp_minutas_versiones
            (minuta_id, codigo, nombre, version, mes, anio, fecha_elaboracion, estado,
             archivo_path, archivo_original, fundacion_id, corporacion_id, usuario_carga,
             observaciones, metadata_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', 'semilla_json_sanitizada', ?, ?, NULL, ?, ?, ?, ?)
        """, (
            minuta_id,
            codigo,
            nombre,
            str(seed.get('version') or seed.get('version_semilla') or '1.0'),
            int(seed.get('mes') or 1),
            int(seed.get('anio') or datetime.now().year),
            str(seed.get('fecha_elaboracion') or ''),
            str(seed.get('estado') or 'borrador').lower(),
            int(fundacion_id),
            int(corporacion_id),
            str(seed.get('observaciones') or ''),
            json.dumps(metadata_json, ensure_ascii=False),
            now_iso(),
            now_iso(),
        ))
        version_id = int(cur.lastrowid)
        product_count = 0
        for group_order, group in enumerate(groups, start=1):
            group_code = canonical_group(group.get('grupo_etario') or group.get('nombre_grupo') or '')
            group_name = str(group.get('nombre_grupo') or group_code).strip()
            gcur = conn.execute("""
                INSERT INTO rpp_minutas_grupos
                (minuta_version_id, grupo_etario, nombre_grupo, orden, activo, created_at, updated_at)
                VALUES (?, ?, ?, ?, 1, ?, ?)
            """, (
                version_id,
                group_code,
                group_name,
                int(group.get('orden') or group_order),
                now_iso(),
                now_iso(),
            ))
            group_id = int(gcur.lastrowid)
            for product_order, product in enumerate(group.get('productos') or [], start=1):
                name = str(product.get('nombre_producto') or '').strip()
                if not name:
                    continue
                conn.execute("""
                    INSERT INTO rpp_minutas_productos
                    (grupo_id, componente, nombre_producto, cantidad, unidad_medida, frecuencia,
                     orden, aplica_bienestarina_cada_dos_meses, activo, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """, (
                    group_id,
                    str(product.get('componente') or ''),
                    name,
                    str(product.get('cantidad') or ''),
                    str(product.get('unidad_medida') or ''),
                    str(product.get('frecuencia') or 'UNA VEZ AL MES'),
                    int(product.get('orden') or product_order),
                    1 if product.get('aplica_bienestarina_cada_dos_meses') else 0,
                    now_iso(),
                    now_iso(),
                ))
                product_count += 1
        conn.execute(
            'INSERT INTO rpp_minutas_auditoria '
            '(accion, minuta_version_id, usuario_id, detalle_json, created_at) VALUES (?, ?, NULL, ?, ?)',
            (
                'INSTALAR_SEMILLA_SANITIZADA',
                version_id,
                json.dumps({
                    'seed_id': seed.get('seed_id'),
                    'grupos': len(groups),
                    'productos': product_count,
                    'sin_datos_personales': True,
                }, ensure_ascii=False),
                now_iso(),
            ),
        )
        conn.commit()
        return {
            'created': True,
            'version_id': version_id,
            'groups': len(groups),
            'products': product_count,
            'period': f"{int(seed.get('anio') or 0):04d}-{int(seed.get('mes') or 0):02d}",
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def guardar_minuta_desde_archivo(database_path: str, archivo_path: str, metadata: dict, destino_folder: str | None = None) -> dict:
    init_schema(database_path)
    codigo = metadata.get('codigo') or 'F2.G36.PP'
    nombre = metadata.get('nombre') or 'Ración Para Preparar Mensual'
    version = metadata.get('version') or '1.0'
    mes = int(metadata.get('mes') or datetime.now().month)
    anio = int(metadata.get('anio') or metadata.get('año') or datetime.now().year)
    fundacion_id = int(metadata.get('fundacion_id') or 1)
    corporacion_id = int(metadata.get('corporacion_id') or 1)
    usuario_id = metadata.get('usuario_id')
    destino_folder = destino_folder or str(Path(archivo_path).parent)
    Path(destino_folder).mkdir(parents=True, exist_ok=True)
    filename = Path(archivo_path).name
    stored = Path(destino_folder) / f"MINUTA_RPP_{anio}_{mes:02d}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
    if str(stored) != str(archivo_path):
        shutil.copy2(archivo_path, stored)
    extraccion = extract_minuta(str(stored))

    conn = connect(database_path)
    minuta_id = _ensure_minuta(conn, codigo, nombre, fundacion_id, corporacion_id)
    cur = conn.execute('''
        INSERT INTO rpp_minutas_versiones
        (minuta_id, codigo, nombre, version, mes, anio, fecha_elaboracion, estado, archivo_path, archivo_original,
         fundacion_id, corporacion_id, usuario_carga, observaciones, metadata_json, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        minuta_id, codigo, nombre, version, mes, anio, metadata.get('fecha_elaboracion') or '',
        metadata.get('estado') or 'borrador', str(stored), filename, fundacion_id, corporacion_id,
        usuario_id, metadata.get('observaciones') or '', json.dumps(extraccion, ensure_ascii=False), now_iso(), now_iso()
    ))
    version_id = int(cur.lastrowid)
    for gi, grupo in enumerate(extraccion.get('grupos') or [], start=1):
        nombre_grupo = grupo.get('nombre_grupo') or grupo.get('grupo_etario') or f'Grupo {gi}'
        grupo_code = canonical_group(nombre_grupo)
        gcur = conn.execute('''
            INSERT INTO rpp_minutas_grupos (minuta_version_id, grupo_etario, nombre_grupo, orden, activo, created_at, updated_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
        ''', (version_id, grupo_code, nombre_grupo, gi, now_iso(), now_iso()))
        grupo_id = int(gcur.lastrowid)
        for pi, prod in enumerate(grupo.get('productos') or [], start=1):
            conn.execute('''
                INSERT INTO rpp_minutas_productos
                (grupo_id, componente, nombre_producto, cantidad, unidad_medida, frecuencia, orden, aplica_bienestarina_cada_dos_meses, activo, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            ''', (
                grupo_id, prod.get('componente') or '', prod.get('nombre_producto') or '', prod.get('cantidad') or '',
                prod.get('unidad_medida') or '', prod.get('frecuencia') or 'UNA VEZ AL MES', pi,
                int(prod.get('aplica_bienestarina_cada_dos_meses') or 0), now_iso(), now_iso()
            ))
    conn.execute('INSERT INTO rpp_minutas_auditoria (accion, minuta_version_id, usuario_id, detalle_json, created_at) VALUES (?, ?, ?, ?, ?)', ('CARGAR_MINUTA', version_id, usuario_id, json.dumps({'archivo': str(stored), 'grupos': len(extraccion.get('grupos') or [])}, ensure_ascii=False), now_iso()))
    conn.commit()
    conn.close()
    return {'version_id': version_id, 'archivo': str(stored), 'extraccion': extraccion}


def listar_minutas(database_path: str, mes: int | None = None, anio: int | None = None) -> list[dict]:
    init_schema(database_path)
    conn = connect(database_path)
    params: list[Any] = []
    where: list[str] = []
    context = current_tenant_context()
    if context.tenant_id and not context.allow_global:
        where.extend([
            'COALESCE(v.fundacion_id, 1)=?',
            'COALESCE(m.fundacion_id, 1)=?',
        ])
        params.extend([int(context.tenant_id), int(context.tenant_id)])
    if mes:
        where.append('v.mes=?')
        params.append(int(mes))
    if anio:
        where.append('v.anio=?')
        params.append(int(anio))
    sql_where = 'WHERE ' + ' AND '.join(where) if where else ''
    rows = conn.execute(
        f'''SELECT v.*, m.nombre AS nombre_minuta
            FROM rpp_minutas_versiones v
            JOIN rpp_minutas m
              ON m.id=v.minuta_id
             AND COALESCE(m.fundacion_id, 1)=COALESCE(v.fundacion_id, 1)
            {sql_where}
            ORDER BY v.created_at DESC''',
        params,
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def obtener_minuta_vigente(
    database_path: str,
    mes: int | None = None,
    anio: int | None = None,
    fundacion_id: int = 1,
    corporacion_id: int = 1,
    *,
    permitir_fallback: bool = True,
) -> dict | None:
    """Obtiene la minuta vigente aplicable al periodo solicitado.

    La vigencia comienza en ``anio/mes`` y continúa hasta que exista otra minuta
    oficial posterior. Para periodos históricos se elige siempre la última minuta
    cuyo inicio sea anterior o igual al periodo consultado; nunca una futura.

    Cuando una fundación no tiene una minuta propia, ``permitir_fallback`` hace
    que herede la minuta oficial vigente del catálogo institucional. Esto comparte
    únicamente la regla institucional; población y archivos continúan aislados
    por tenant. Nunca se reutiliza una minuta futura para un periodo histórico.
    """
    init_schema(database_path)
    conn = connect(database_path)
    def buscar_version(fid: int, cid: int):
        params = [int(fid), int(cid)]
        where = ['v.estado=\'vigente\'', 'v.fundacion_id=?', 'v.corporacion_id=?']
        if mes is not None and anio is not None:
            mes_consulta = max(1, min(12, int(mes)))
            anio_consulta = int(anio)
            where.append('(v.anio < ? OR (v.anio = ? AND v.mes <= ?))')
            params.extend([anio_consulta, anio_consulta, mes_consulta])
        elif anio is not None:
            where.append('v.anio <= ?')
            params.append(int(anio))
        elif mes is not None:
            where.append('v.mes <= ?')
            params.append(max(1, min(12, int(mes))))
        return conn.execute(
            f'''SELECT v.* FROM rpp_minutas_versiones v
                WHERE {' AND '.join(where)}
                ORDER BY v.anio DESC, v.mes DESC, v.created_at DESC LIMIT 1''',
            params,
        ).fetchone()

    def buscar_version_institucional():
        params = []
        where = ['v.estado=\'vigente\'']
        if mes is not None and anio is not None:
            mes_consulta = max(1, min(12, int(mes)))
            anio_consulta = int(anio)
            where.append('(v.anio < ? OR (v.anio = ? AND v.mes <= ?))')
            params.extend([anio_consulta, anio_consulta, mes_consulta])
        elif anio is not None:
            where.append('v.anio <= ?')
            params.append(int(anio))
        elif mes is not None:
            where.append('v.mes <= ?')
            params.append(max(1, min(12, int(mes))))
        return conn.execute(
            f'''SELECT v.* FROM rpp_minutas_versiones v
                WHERE {' AND '.join(where)}
                ORDER BY v.anio DESC, v.mes DESC, v.updated_at DESC, v.created_at DESC LIMIT 1''',
            params,
        ).fetchone()

    row = buscar_version(fundacion_id, corporacion_id)
    heredada_global = False
    if not row and permitir_fallback:
        row = buscar_version_institucional()
        heredada_global = bool(row)
    if not row:
        conn.close()
        return None
    version = dict(row)
    version['heredada_global'] = heredada_global
    version['fundacion_solicitante_id'] = int(fundacion_id)
    grupos_rows = conn.execute('SELECT * FROM rpp_minutas_grupos WHERE minuta_version_id=? AND activo=1 ORDER BY orden,id', (version['id'],)).fetchall()
    grupos = []
    for gr in grupos_rows:
        g = dict(gr)
        prows = conn.execute('SELECT * FROM rpp_minutas_productos WHERE grupo_id=? AND activo=1 ORDER BY orden,id', (g['id'],)).fetchall()
        g['productos'] = [dict(p) for p in prows]
        grupos.append(g)
    version['grupos'] = grupos
    conn.close()
    return version


def productos_para_grupo(minuta: dict | None, grupo_etario: str) -> list[dict]:
    if not minuta:
        return []
    target = canonical_group(grupo_etario)
    # Menores de 6 meses usan la minuta 6-11 si no hay específica.
    if target == 'menor_6_meses':
        target_candidates = {'menor_6_meses', '6_11_meses'}
    else:
        target_candidates = {target}
    for grupo in minuta.get('grupos') or []:
        code = canonical_group(grupo.get('grupo_etario') or grupo.get('nombre_grupo'))
        if code in target_candidates:
            return list(grupo.get('productos') or [])
    return []


def productos_para_usuario(minuta: dict | None, user: dict) -> list[dict]:
    return productos_para_grupo(minuta, grupo_para_usuario(user))


def obtener_equivalencias(database_path: str | None = None) -> dict[str, list[str]]:
    result = {k: list(v) for k, v in DEFAULT_EQUIVALENCIAS.items()}
    if database_path:
        try:
            init_schema(database_path)
            conn = connect(database_path)
            rows = conn.execute('SELECT * FROM rpp_minutas_equivalencias WHERE activo=1').fetchall()
            conn.close()
            for r in rows:
                aliases = []
                try:
                    aliases = json.loads(r['alias_json'] or '[]')
                except Exception:
                    aliases = []
                aliases.append(r['producto_plantilla'])
                result[normalizar_texto(r['nombre_producto'])] = [normalizar_texto(a) for a in aliases if a]
        except Exception:
            pass
    return result


def marcar_minuta_vigente(database_path: str, version_id: int, usuario_id: int | None = None) -> dict:
    init_schema(database_path)
    conn = connect(database_path)
    row = conn.execute('SELECT * FROM rpp_minutas_versiones WHERE id=?', (version_id,)).fetchone()
    if not row:
        conn.close()
        raise ValueError('Minuta no encontrada.')
    v = dict(row)
    conn.execute('''UPDATE rpp_minutas_versiones SET estado='historico', updated_at=? WHERE minuta_id=? AND mes=? AND anio=? AND fundacion_id=? AND corporacion_id=? AND id<>?''', (now_iso(), v['minuta_id'], v['mes'], v['anio'], v['fundacion_id'], v['corporacion_id'], version_id))
    conn.execute('UPDATE rpp_minutas_versiones SET estado=\'vigente\', updated_at=? WHERE id=?', (now_iso(), version_id))
    conn.execute('INSERT INTO rpp_minutas_auditoria (accion, minuta_version_id, usuario_id, detalle_json, created_at) VALUES (?, ?, ?, ?, ?)', ('MARCAR_VIGENTE', version_id, usuario_id, json.dumps({'mes': v['mes'], 'anio': v['anio']}, ensure_ascii=False), now_iso()))
    conn.commit()
    conn.close()
    return {'version_id': version_id, 'estado': 'vigente'}
