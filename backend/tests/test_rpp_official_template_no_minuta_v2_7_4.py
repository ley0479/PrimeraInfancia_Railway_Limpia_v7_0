"""RPP se genera por categoría desde plantilla oficial, sin Minuta Patrón."""
from pathlib import Path
import sys
import tempfile

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from modules.plantillas_oficiales import generar_desde_plantilla_oficial  # noqa: E402


app_source = (ROOT / "backend/app.py").read_text(encoding="utf-8")
frontend = (ROOT / "frontend/js/app.js").read_text(encoding="utf-8")
presenter_css = (ROOT / "frontend/css/elian-presenter.css").read_text(encoding="utf-8")
avatar_css = (ROOT / "frontend/css/ian-avatar.css").read_text(encoding="utf-8")
env_example = (ROOT / ".env.example").read_text(encoding="utf-8")

start = app_source.index("def descargar_rpp_por_categoria")
endpoint = app_source[start:app_source.index("@app.route('/api/descargar-archivo", start)]
assert "obtener_minuta_vigente" not in endpoint
assert "requiereMinutaVigente" not in endpoint
assert "RPP_SOURCE_MODE" in endpoint
assert "source_mode='official_template'" in endpoint
assert "RPP_REQUIRE_MINUTA_PATRON=false" in env_example
assert "RPP_ENABLE_MINUTA_ENRICHMENT=false" in env_example
assert "RPP_SOURCE_MODE=official_template" in env_example
assert "Authorization': `Bearer ${token}`" in frontend
assert "&mes=${encodeURIComponent(periodo.mes)}&anio=${encodeURIComponent(periodo.anio)}" in frontend
assert ".elian-presenter { position: fixed; inset: 0; z-index: 10043; pointer-events: none; }" in presenter_css
assert ".ian-tour-avatar{position:fixed" in avatar_css and "pointer-events:none" in avatar_css

categories = {
    "GESTANTES_0_6_MESES": {"NUI": "G-1", "Nombre": "Registro prueba", "GrupoEdad": "D"},
    "6_11_MESES": {"NUI": "E-1", "Nombre": "Registro prueba", "EdadMeses": 8},
    "1_2_ANOS": {"NUI": "F-1", "Nombre": "Registro prueba", "EdadMeses": 24},
    "3_5_ANOS": {"NUI": "G-2", "Nombre": "Registro prueba", "EdadMeses": 48},
}

with tempfile.TemporaryDirectory() as temp:
    base = Path(temp)
    official = base / "templates" / "oficiales"
    official.mkdir(parents=True)
    template = official / "plantilla_rpp_oficial.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "RPP OFICIAL"
    ws["A1"] = "PLANTILLA ORIGINAL"
    ws["B40"] = "=SUM(1,2)"
    ws.print_area = "A1:AA42"
    wb.create_sheet("INSTRUCCIONES")["A1"] = "Conservar"
    wb.save(template)

    for category, participant in categories.items():
        output = base / f"RPP_UNIDAD_PRUEBA_{category}_2026_08.xlsx"
        generar_desde_plantilla_oficial(
            "rpp",
            {"metadata": {"unidad": "UNIDAD PRUEBA", "mes": 8, "anio": 2026}, "usuarios": [participant]},
            output,
            base / "templates",
        )
        assert output.is_file() and output.stat().st_size > 0
        generated = load_workbook(output, data_only=False)
        assert generated.sheetnames == ["RPP OFICIAL", "INSTRUCCIONES"]
        assert generated["RPP OFICIAL"]["B40"].value == "=SUM(1,2)"
        assert "$A$1:$AA$42" in str(generated["RPP OFICIAL"].print_area)
        generated.close()

print("RPP_OFFICIAL_TEMPLATE_NO_MINUTA_V2_7_4_PASS")
