import sqlite3
import sys
import tempfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / 'backend'))
from generador_formatos import GeneradorFormatos


def main():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / 'qap6.db'
        conn = sqlite3.connect(db)
        conn.executescript('''
          CREATE TABLE sn_valoraciones(id INTEGER PRIMARY KEY,activo INTEGER,fundacion_id INTEGER,
            documento TEXT,nombre_completo TEXT,unidad TEXT,periodo TEXT,fecha_valoracion TEXT,
            peso_kg REAL,talla_cm REAL,diagnostico_global TEXT);
          INSERT INTO sn_valoraciones VALUES(1,1,1,'DOC1','PERSONA MAESTRA','UDS UNO','2026-08','2026-08-01',10,80,'RIESGO');
          INSERT INTO sn_valoraciones VALUES(2,1,1,'DOC1','PERSONA MAESTRA','UDS UNO','2026-08','2026-08-20',11,81,'ADECUADO');
          INSERT INTO sn_valoraciones VALUES(3,1,2,'OTRO','OTRA CORPORACION','UDS UNO','2026-08','2026-08-20',99,99,'OTRO');
        ''')
        conn.commit(); conn.close()
        path = GeneradorFormatos(str(db), str(root), str(root)).generar_nutricion(8, 2026, 'UDS UNO')
        wb = load_workbook(path, data_only=True); ws = wb['NUTRICION']
        rows = list(ws.iter_rows(min_row=2, values_only=True)); wb.close()
        assert len(rows) == 1, rows
        assert rows[0][0] == 'PERSONA MAESTRA' and rows[0][2] == 11, rows
    print('REPORTE_NUTRICION_CANONICO_QAP6_PASS')


if __name__ == '__main__':
    main()
