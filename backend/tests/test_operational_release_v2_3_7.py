from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from models import Schema
from modules.plantillas_oficiales import generar_desde_plantilla_oficial, iter_plantillas_oficiales_para_generacion
from services.rpp_minutas_service import obtener_minuta_vigente, seed_minuta_sanitizada_desde_json
from services.seed_sync import sync_managed_seed_tree
from services.uds_catalog import catalog_summary, ensure_catalog_units_sqlite, migrate_demo_units_sqlite, normalize_unit


def assert_true(value, message):
    if not value:
        raise AssertionError(message)


def run():
    checks = []
    summary = catalog_summary()
    assert_true(summary["total_unidades"] == 32, "El catálogo UDS no contiene 32 unidades")
    assert_true(normalize_unit("UCA EYAZAKE") == "EYASAKE", "Alias EYAZAKE no normalizado")
    assert_true(normalize_unit("UNIDAD DEMO 32") == "VIVE LA PAZ 2", "Migración demo 32 incorrecta")
    checks.append("Catálogo UDS y alias")

    with tempfile.TemporaryDirectory(prefix="pi-uds-rpp-") as temp:
        db = Path(temp) / "database.sqlite3"
        conn = sqlite3.connect(db)
        conn.executescript(Schema.get_schema_sql())
        conn.execute("INSERT INTO beneficiarios (documento,nombres,apellidos,fecha_nacimiento,unidad,fecha_ingreso,fecha_carga) VALUES (?,?,?,?,?,?,?)", ('0001','PERSONA','PRUEBA','2025-01-01','UNIDAD DEMO 01','2026-01-01','2026-08-03T00:00:00'))
        conn.commit()
        conn.close()
        migration = migrate_demo_units_sqlite(str(db))
        ensure = ensure_catalog_units_sqlite(str(db), fundacion_id=1)
        conn = sqlite3.connect(db)
        unit = conn.execute("SELECT unidad FROM beneficiarios LIMIT 1").fetchone()[0]
        count = conn.execute("SELECT COUNT(*) FROM unidades").fetchone()[0]
        conn.close()
        assert_true(unit == "BAJO PACURITA", "La UDS demo no migró al catálogo")
        assert_true(count >= 32, "No se sembraron las 32 UDS")
        assert_true(migration["updated_scalar_values"] >= 1 and ensure["creadas"] >= 1, "Reportes de UDS incompletos")

        seed_file = BACKEND / "seed_data" / "config" / "rpp_minuta_base_2026_05.json"
        first = seed_minuta_sanitizada_desde_json(str(db), seed_file)
        second = seed_minuta_sanitizada_desde_json(str(db), seed_file)
        minute = obtener_minuta_vigente(str(db), mes=5, anio=2026)
        minute_june = obtener_minuta_vigente(str(db), mes=6, anio=2026)
        assert_true(
            minute_june is not None and minute_june.get('id') == minute.get('id'),
            "La minuta de mayo debe continuar vigente en junio si no existe otra posterior",
        )
        assert_true(
            obtener_minuta_vigente(str(db), mes=4, anio=2026) is None,
            "Una minuta futura nunca debe reutilizarse para un periodo anterior",
        )
        conn = sqlite3.connect(db)
        conn.execute("UPDATE rpp_minutas_versiones SET fundacion_id=8, corporacion_id=3")
        conn.commit()
        conn.close()
        minute_tenant = obtener_minuta_vigente(str(db), mes=6, anio=2026, fundacion_id=25, corporacion_id=1)
        assert_true(
            minute_tenant is not None and minute_tenant.get('id') == minute.get('id') and minute_tenant.get('heredada_global') is True,
            "Una fundación nueva debe heredar la minuta institucional aunque no esté registrada como 1/1",
        )
        conn = sqlite3.connect(db)
        equivalence_count = conn.execute("SELECT COUNT(*) FROM rpp_minutas_equivalencias WHERE activo=1").fetchone()[0]
        conn.close()
        assert_true(first.get("created") is True, "No se creó la minuta RPP")
        assert_true(second.get("created") is False, "La semilla RPP no fue idempotente")
        assert_true(len(minute.get("grupos") or []) == 4, "La minuta RPP no tiene 4 grupos")
        assert_true(sum(len(g.get("productos") or []) for g in minute["grupos"]) == 49, "La minuta RPP no tiene 49 productos")
        assert_true(equivalence_count == 17, f"Se esperaban 17 equivalencias RPP y hay {equivalence_count}")
    checks.append("Migración UDS y semilla RPP idempotente")

    seed_root = BACKEND / "seed_data" / "templates_originales"
    july = iter_plantillas_oficiales_para_generacion(seed_root, mes=7, anio=2026)
    august = iter_plantillas_oficiales_para_generacion(seed_root, mes=8, anio=2026)
    july_ram = [x for x in july if x.get("tipo") == "ram"]
    august_ram = [x for x in august if x.get("tipo") == "ram"]
    assert_true(len(july_ram) == 1 and str(july_ram[0].get("version")) == "2", "Julio no seleccionó RAM V2")
    assert_true(len(august_ram) == 1 and str(august_ram[0].get("version")) == "3", "Agosto no seleccionó RAM V3")

    synthetic_user = {
        "tipo_documento": "RC",
        "numero_documento": "00000001",
        "primer_nombre": "PRUEBA",
        "primer_apellido": "CONTROL",
        "fecha_nacimiento": "2025-01-01",
        "fecha_ingreso": "2026-01-01",
    }
    with tempfile.TemporaryDirectory(prefix="pi-ram-") as temp:
        out2 = Path(temp) / "ram_v2.xlsx"
        out3 = Path(temp) / "ram_v3.xlsx"
        generar_desde_plantilla_oficial(
            "ram",
            {"metadata": {"anio": 2026, "mes_numero": 7, "unidad": "BAJO PACURITA"}, "usuarios": [synthetic_user]},
            out2,
            seed_root,
        )
        generar_desde_plantilla_oficial(
            "ram",
            {"metadata": {"anio": 2026, "mes_numero": 8, "unidad": "BAJO PACURITA"}, "usuarios": [synthetic_user]},
            out3,
            seed_root,
        )
        assert_true(out2.exists() and out2.stat().st_size > 0, "RAM V2 no se generó")
        assert_true(out3.exists() and out3.stat().st_size > 0, "RAM V3 no se generó")
        wb2 = load_workbook(out2, read_only=True)
        wb3 = load_workbook(out3, read_only=True)
        try:
            assert_true("FORMATO RAM V2 HISTORICO" in wb2.sheetnames, "Hoja RAM V2 ausente")
            assert_true("FORMATO RAM" in wb3.sheetnames, "Hoja RAM V3 ausente")
        finally:
            wb2.close()
            wb3.close()

    historical = seed_root / "oficiales" / "plantilla_ram_oficial_v2_historica.xlsx"
    historical_wb = load_workbook(historical, data_only=False)
    historical_ws = historical_wb["FORMATO RAM V2 HISTORICO"]
    for cell in ("A4", "A5", "D5", "F5", "H5", "U5", "A6", "D6", "F6", "I6", "A7", "F7", "K7", "A8", "F8", "T8"):
        assert_true("_" in str(historical_ws[cell].value or ""), f"El encabezado {cell} no quedó neutralizado")
    for row in range(15, 35):
        for col in range(2, 8):
            assert_true(historical_ws.cell(row=row, column=col).value in (None, ""), f"RAM V2 contiene datos en fila {row}, columna {col}")
    historical_wb.close()
    checks.append("RAM V2/V3 por vigencia y plantilla histórica sanitizada")

    with tempfile.TemporaryDirectory(prefix="pi-seed-sync-") as temp:
        base = Path(temp)
        src = base / "src"
        dst = base / "dst"
        data = base / "data"
        backups = base / "backups"
        src.mkdir()
        (src / "managed.txt").write_text("v1", encoding="utf-8")
        (src / "seed_manifest.json").write_text(json.dumps({"plantillas": []}), encoding="utf-8")
        first = sync_managed_seed_tree(src, dst, data_dir=data, backups_dir=backups)
        assert_true("managed.txt" in first["copied"], "Primera sincronización no copió")
        (src / "managed.txt").write_text("v2", encoding="utf-8")
        second = sync_managed_seed_tree(src, dst, data_dir=data, backups_dir=backups)
        assert_true("managed.txt" in second["updated"], "Semilla administrada no se actualizó")
        (dst / "managed.txt").write_text("personalizada", encoding="utf-8")
        (src / "managed.txt").write_text("v3", encoding="utf-8")
        third = sync_managed_seed_tree(src, dst, data_dir=data, backups_dir=backups)
        assert_true("managed.txt" in third["preserved_custom"], "Plantilla personalizada no se preservó")
    checks.append("Sincronización gestionada con respaldo y preservación")

    app_text = (BACKEND / "app.py").read_text(encoding="utf-8")
    html_text = (ROOT / "frontend" / "index.html").read_text(encoding="utf-8")
    js_text = (ROOT / "frontend" / "js" / "modules" / "acceso-compartido.js").read_text(encoding="utf-8")
    assert_true("def resolver_url_publica" in app_text, "Falta resolución de URL pública")
    assert_true("/api/acceso/storage-health" in app_text, "Falta diagnóstico de volumen")
    assert_true("/api/formatos/diagnostico" in app_text, "Falta preflight de formatos")
    assert_true("Enlace público de la plataforma" in html_text, "UI no muestra enlace Railway")
    assert_true("data.esProduccion" in js_text, "Frontend no diferencia producción")
    assert_true("accesoProbarAlmacenamiento" in js_text, "Acceso Compartido no comprueba /data")
    checks.append("Acceso Railway, volumen y preflight")

    print(json.dumps({"ok": True, "checks": checks}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    run()
