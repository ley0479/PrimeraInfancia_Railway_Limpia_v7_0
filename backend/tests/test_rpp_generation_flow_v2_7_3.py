"""Regresión integral del flujo RPP: sesión, periodo, clasificación y Excel oficial."""
from pathlib import Path
import sys
import tempfile

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from modules.plantillas_oficiales import (  # noqa: E402
    _grupo_edad_marker,
    generar_desde_plantilla_oficial,
    iter_plantillas_oficiales_para_generacion,
)


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def main():
    frontend = (ROOT / "frontend/js/app.js").read_text(encoding="utf-8")
    app_source = (ROOT / "backend/app.py").read_text(encoding="utf-8")

    helper_start = frontend.index("async function descargarArchivoFormatoAlpha63")
    helper_end = frontend.index("async function descargarBienestarinaAlpha62", helper_start)
    helper = frontend[helper_start:helper_end]
    require("const token = authToken()" in helper, "La descarga RPP no obtiene la sesión")
    require("'Authorization': `Bearer ${token}`" in helper, "La descarga RPP no envía Authorization")
    require("&mes=${encodeURIComponent(periodo.mes)}&anio=${encodeURIComponent(periodo.anio)}" in frontend,
            "RPP no envía el periodo seleccionado")

    endpoint_start = app_source.index("def descargar_rpp_por_categoria")
    endpoint = app_source[endpoint_start:app_source.index("@app.route('/api/descargar-archivo", endpoint_start)]
    for contract in (
        "iter_plantillas_oficiales_para_generacion",
        "obtener_minuta_vigente",
        "_alpha59_obtener_usuarios_unidad",
        "_alpha64_generar_rpp_resiliente(unidad, grupo, mes=mes, anio=anio)",
    ):
        require(contract in endpoint, f"Falta validación RPP: {contract}")

    require(_grupo_edad_marker({"GrupoEdad": "", "EdadMeses": None}) == "",
            "Una edad ausente todavía se clasifica como cero meses")
    require(_grupo_edad_marker({"EdadMeses": 8}) == "E", "No clasifica 6 a 11 meses")
    require(_grupo_edad_marker({"EdadMeses": 24}) == "F", "No clasifica 1 a 2 años")
    require(_grupo_edad_marker({"EdadMeses": 48}) == "G", "No clasifica 3 a 5 años")

    with tempfile.TemporaryDirectory() as tmp:
        base = Path(tmp)
        templates = base / "templates"
        official = templates / "oficiales"
        official.mkdir(parents=True)
        template_path = official / "plantilla_rpp_oficial.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "PLANTILLA O FORMADE RPP"
        ws["B40"] = "=SUM(1,2)"
        ws.print_area = "A1:AA42"
        instructions = wb.create_sheet("INSTRUCCIONES")
        instructions["A1"] = "Hoja oficial que debe conservarse"
        wb.save(template_path)
        wb.save(official / "plantilla_rpp_oficial_v2026.xlsx")
        available = iter_plantillas_oficiales_para_generacion(templates, mes=8, anio=2026)
        require(any(item.get("tipo") == "rpp" for item in available), f"No detectó plantilla RPP: {available}")

        output = base / "RPP_PRUEBA_2026_08.xlsx"
        generar_desde_plantilla_oficial(
            "rpp",
            {
                "metadata": {"unidad": "UCA PRUEBA", "mes": 8, "anio": 2026},
                "usuarios": [{"NUI": "PRUEBA-1", "Nombre": "Usuario Prueba", "EdadMeses": 8}],
            },
            output,
            templates,
        )
        require(output.is_file() and output.stat().st_size > 0, "RPP no creó un archivo real")
        generated = load_workbook(output, data_only=False)
        require(generated.sheetnames == ["PLANTILLA O FORMADE RPP", "INSTRUCCIONES"],
                "RPP alteró las hojas oficiales")
        require(generated["PLANTILLA O FORMADE RPP"]["B40"].value == "=SUM(1,2)",
                "RPP alteró una fórmula oficial")
        require("$A$1:$AA$42" in str(generated["PLANTILLA O FORMADE RPP"].print_area),
                "RPP alteró el área de impresión oficial")
        generated.close()

    print("RPP_GENERATION_FLOW_V2_7_3_PASS")


if __name__ == "__main__":
    main()
